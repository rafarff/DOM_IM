# -*- coding: utf-8 -*-
"""
gerar_planilha.py — Script ÚNICO da Fase 1 (Inteligência de Mercado DOM)
Padrão congelado em 14/04/2026. Ver PADRAO.md no mesmo diretório.

Regras:
  - Não criar variantes deste script (v11, v12, etc.). Versiona-se OS DADOS, não o código.
  - Não adicionar colunas ao template sem atualizar antes o PADRAO.md e obter aprovação.
  - Gerar: Planilha_Mestre_Panorama_vX.Y.xlsx em /00_ESTUDO_CONSOLIDADO/

Uso: python3 gerar_planilha.py
"""
import os, glob, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ═══════════════════════════════════════════════════════════════
# PARÂMETROS GLOBAIS
# ═══════════════════════════════════════════════════════════════
VERSION = "10.8"
DATE_STR = "03/05/2026"
# v10.8 — (03/05/2026): WEB RESEARCH BATCH (Rafael 03/05). +4 destravados de Total
# via fontes oficiais e imprensa, +7 enriquecidos parciais, +3 correções de bairro:
#   Destravados (Total apurado por web/site_oficial):
#     - LIV Residence (Alfa): Total=75, mono 3D, área 90,83-100,23m². Site Alfa.
#     - Residencial Ana Vitória (Castelucci): Total=30 casas 83m² 2D;3D. Site Castelucci.
#     - Mount Solaro (Berg+Gonçalves): Total=50 (20+10+20), book multi 2D+3D. Site Gonçalves.
#       Bairro corrigido São Luís→Península.
#     - Village Prime Eldorado (Canopus): Total=400, mono 2D 43,5m², 5 torres. Imirante 31/10.
#   Enriquecidos parciais (Total ainda falta mas dados preenchidos):
#     - Legacy Residence (Alfa): tipologia 4D, área 175-180m².
#     - Reserva Península (Sá Cav.): tipologia 4D, área 127-171m². Bairro Península.
#     - Varandas Grand Park (Franere): tipologia 3D, área 74-87m². Bairro Calhau.
#     - Villa Adagio (Lua Nova): tipologia 2D, área 48,90m². Bairro Iguaíba.
#     - Village Reserva II (Canopus): tipologia 2D 41m². Bairro Cohatrac.
#     - Village Del Ville II (Canopus): tipologia 2D 42-43m². Bairro Iguaíba.
#     - Villa di Carpi (Castelucci): área 49-52m² (3 plantas).
#   Correções de bairro: Mount Solaro→Península, Varandas→Calhau, Villa Adagio→Iguaíba.
#   Achado fora-do-escopo: Lagoon Residence (Lua Nova) é em Santo Amaro do Maranhão
#   (cidade satélite, não SLZ-Grande SLZ) — flag pra Rafael decidir manter ou tirar.
# Cobertura Composição: 29/44 → ~33/44 = 75%. Bloqueados: 17 → ~13.
# v10.7 — (03/05/2026): +2 destravados de pendencias_TOTAL.md (Rafael 03/05):
#   1. Dom Antônio: Total=12 (Rafael), 12 casas iguais 136,2m², mono-tipologia 3D.
#      §3.7 nível 5.1 aplica automático: 12u 3D 136,2m². Origem total = informado_manualmente.
#   2. Edifício Dom Ricardo: Total=30 (Rafael), 10 andares × 3 col por andar.
#      Composição via book (fonte forte nível 3): 10u 2D 71,92m² + 20u 3D 84,96-85,75m².
#      Book diz "100% VENDIDO" → estoque manual 0.0. Total tipologia declarado em
#      BOOK_TOTAL_OVERRIDE (compute_total_per_tipologia).
# Carteira: 27/44 → 29/44 cobertura Composição = 66%. Bloqueados: 17 → 15.
# v10.6 — (03/05/2026): VIRADA ESTRUTURAL §3.7 v2 (PADRAO v6.2)
#   1. Consolidação multi-torre (regra A): Vernazza N+S → "Vernazza Residenza";
#      Giardino Fiore+Luce → "Giardino Residenza". Carteira 46→44.
#   2. Composição obrigatória — invariante Σ Total tipologia = E_RAW.Total.
#      Aba Composição expandida 10→11 col (+ "Total tipologia" entre Tipologia e
#      Disponíveis). Total tipologia computado em runtime:
#        - Mono-tipologia E_RAW: Total tipologia = E_RAW.Total (trivial)
#        - Multi-tipologia origem `tabela_local_completa`: Σ disp já = Total
#        - Multi-tipologia origem `tabela_local_parcial`: pro-rata (sufixo _pro_rata)
#   3. Hierarquia §3.7.A ganha NÍVEL 5 `estimativa_distribuição` (sub-regras
#      5.1-5.4). Aplicado automaticamente a empreend. com Total mas sem C_RAW.
#      Sub-regras: 5.1 mono / 5.2 multi+área / 5.3 multi sem área / 5.4 sem tipologia.
#   4. Total é a âncora: estimativas auto-fecham com Total; fontes fortes
#      (níveis 1-4) que não fechem geram WARN sem alterar Total.
#   5. Empreend. sem Total apurado → `pendencias_TOTAL.md`, NÃO recebem
#      estimativa de Composição.
# v5.0 — (25/04/2026): MUDANÇA ESTRUTURAL — adoção do PADRAO v2.0.
# +Coluna Tipo (Vertical/Horizontal/Misto) inserida como col. 5. 24 → 25 colunas.
# +Hiali e DOM Incorporação como incorporadoras monitoradas (14 → 16). Tracking da
# .xlsx no git habilitado. Total: 45 empreendimentos, 16 incorporadoras.
# v5.1 — (27/04/2026): DOIS GRUPOS de mudança:
#
# (a) NOVOS DADOS de tabelas locais:
# +Vila Coimbra (Castelucci): tabela 03/2026 → 124,63m², ticket R$1.019k-1.082k,
# 36+ casas Araçagi. +Giardino Fiore (Alfa): 6/45 → 87% vendido → Últimas unidades.
# +Giardino Luce (Alfa): 5/60 → 92%, dorms corrigido para 3 suítes. +Monte Meru
# (Berg): tabela ABR/2026 → 135m², ticket R$1.932k-1.944k. +The View (Delman):
# tabela 27.04 substitui 24.04 → ticket_min sobe para R$559k (apto 409 vendido),
# 90 aptos disponíveis de ~110 = ~20 vendas em 3 dias.
#
# (b) CORREÇÃO SISTÊMICA de Mês lançamento (PADRAO §1 — col 9 deve ser MM/AAAA):
# 13 entries com pasta XX_MMAAAA tinham datas imprecisas no E_RAW (ex: Lagoon
# como ~2024 quando pasta dizia 042026). Atualizadas para MM/AAAA preciso.
# 8 entries sem pasta-data: estimadas como 06/AAAA + ⚠ T-36. parse_launch_date
# em ambos os scripts agora REJEITA AAAA puro (vai pro fim da lista).
# Validação assert no início do script bloqueia E_RAW com mês fora do padrão.
# v5.2 — (27/04/2026): +Edifício Bossa (Mota Machado) atualizado a partir da tabela
# local 04/2026: tickets R$2,85-3,71M, áreas 191-196m², R$/m² médio R$16.663 (faixa
# 14,9-19,2k), 36 aptos disponíveis de 60 (~60% estoque, 40% vendido), entrega
# 09/2030. Endereço completo. Segmento reclassifica auto pelo R$/m².
# +Fix reclassificar_status: preserva "Lançamento" como decisão de tempo (< 6m de
# venda), não força "Em comercialização" só por estoque > 15%.
# +Fix áreas em horizontais (Dom Lucas + Dom José): valores armazenados como
# construído/terreno misturados — corrigido para área CONSTRUÍDA em min e max
# (uma única tipologia em ambos). Terreno migrado para Observações. Convenção
# nova no PADRAO v2.0 §1 nota.
# v5.3 — (27/04/2026): RECALIBRAÇÃO de faixas de Segmento (PADRAO §4.2 v2.2):
# Médio 6-8k (antes 6-9k), Médio-alto 8-10k (antes 9-13k), Alto 10-15k (antes
# 13-18k), Luxo >15k (antes >18k). Motivo: classificação antiga punha Bossa,
# The View, Vernazza, Giardino (R$ 14-17k) em Alto, contradizendo posicionamento
# de marca real. Distribuição da carteira após recalibragem: 2 Médio, 3 Médio-
# alto, 6 Alto, 12 Luxo, 22 sem dados. 8 entries com Segmento hardcoded
# divergente do cálculo: trocadas para None (auto-classifica) — evita drift
# futuro entre intent e fórmula.
# v6.0 — (27/04/2026): MUDANÇA ESTRUTURAL — coluna Status REMOVIDA (PADRAO v3.0).
# 25 → 24 colunas. Motivo: classificação subjetiva, parcialmente derivada de outros
# campos (estoque, data). Removido: STATUS list, função reclassificar_status, todas
# as 45 entries do E_RAW perderam o 7º campo. Indices internos do script todos
# shiftados em -1 a partir da col 7. PADRAO §4.3 marcado como removido.
# Adicionalmente: filtro "ativo no ciclo" no HTML também eliminado — Panorama mostra
# TODOS os 45 empreendimentos mapeados (decisão Rafael). build_panorama.should_include
# sempre retorna True. As 3 abas (Panorama, Pendências, Dados Completos) compartilham
# o mesmo universo, diferenciando-se apenas pela visualização.
# v6.1 — (27/04/2026): Tipologia padronizada (PADRAO §4.6). Col 13 vira enum:
# Studio | 1D | 2D | 3D | 4D | Lote (separar por ;). Suíte conta como dormitório.
# Texto descritivo antigo migrado para Observações com prefixo "Tipologia detalhada:".
# Tooltip ℹ na col Tipologia (HTML) extrai esse prefixo das Observações.
# v6.2 — (27/04/2026): apenas regra de UI (sem mudança nos DADOS). Pin no mapa
# passa a aparecer SÓ quando Endereço (col 3) é "completo" (Rua/Av./Plus Code).
# Endereços com só bairro ficam fora do mapa, sinalizados com badge ◌ "sem endereço".
# A pendência "endereço" entra automaticamente na lista da Tabela B. Implementação
# em build_panorama.py: tem_endereco_completo() + on_map=True só com ambos
# (endereço_ok + geocoda). PADRAO §1 col 3 atualizada e v3.2 emitida.
# v6.3 — (27/04/2026): MAPA REMOVIDO do HTML (PADRAO v3.3). Decisão do Rafael —
# pins aproximados por bairro estavam confundindo mais que ajudando. Visual: aba
# Panorama foca nas tabelas A e B sem o mapa interativo. Site identidade DOM
# (logo embutido + fundo cinza-dark) + filtro Tipo + senha JS (DOM2026) + remoção
# de "2025/2026" e legenda Origem data. Funções tem_endereco_completo() e
# geocode_bairro() preservadas no build_panorama.py para futuro retorno. Pendência
# "endereço" continua útil na Tabela B até reativarmos o mapa.
# v6.4 — (27/04/2026): MUDANÇA DUPLA. (a) Tipo "Misto" removido (não havia entries).
# +Tipo "Loteamento" formalizado (PADRAO 3.4 §4.5). Villagio Treviso reclassificado
# de Horizontal → Loteamento. +Novo empreend. Lua Nova: Golden Green Beach (lotes
# de luxo no Calhau, R$2,65-4,4M, terrenos 453-682m²). Atenção: R$/m² em loteamento
# é m² de TERRENO, não comparável com construído. (b) Aba Dashboard adicionada ao
# HTML com 6 cards de análise usando Chart.js CDN. Logo DOM atualizado (versão
# minimalista do INBOX, só ícone). 45 → 46 empreendimentos.
# v6.5 — (28/04/2026): Atualização Inbox 28/04. (a) The View — tabela 28/04
# (versão 0000001 - 1.1) substitui v2_2704: 93 aptos disponíveis residenciais +
# 1 loja. Ticket_min CAI para R$ 539.969 (apto 414, 36,45m²) — vs R$559k em
# 27/04 (apto que era min foi vendido OU apto 414 voltou ao mercado). Ticket_max
# inalterado (R$1.504k, apto 1601). Estimativa ~49% vendido (assumindo ~182
# unidades totais = 13 andares × 14 aptos). Cobertura 17º pavto com 5 aptos
# diferenciados (54-80m²). +1 loja (LOJA 21, 48m², R$890k).
# (b) Edifício Bossa — sem mudança em tickets/áreas (tabela 04/2026 idêntica
# md5 à arquivada em 24/04 → eram duplicatas no Inbox). Enriquecimento de
# observações: arquiteto Nasser Hissa Arquitetos Associados (parceiro), nota
# de vizinhança com The View (mesma quadra 02 da Av. Holandeses, lotes 07 e 08).
# Endereço refinado para "Loteamento Calhau".
# (c) Fix estrutural — BASE e SKILL_ASSETS deixam de ser hardcoded com session
# ID antiga (intelligent-festive-lamport). Agora são derivados via pathlib
# relativa ao __file__, tornando script auto-suficiente em qualquer sessão.
# Logo DOM copiado para _PADRAO_FASE_1/assets/.
# v6.6 — (29/04/2026): TRÊS GRUPOS de mudança.
# (a) UI HTML: logo aumentado 60→90px (presença de marca). +Coluna "% Vendido"
# na Tabela A da aba Panorama com tooltip rico de origem (orig_estoque,
# orig_precos, data_verif, ref §3.3). Coloração leve: ≥85% dourado bold,
# 60-85% neutro, <60% cinza claro.
# (b) AUDITORIA de gaps tabela arquivada vs E_RAW: 26 empreend. com tabela
# em /TABELA/, 22 já completos. 4 gaps reais identificados.
# (c) Preenchimento de 3 dos 4 gaps a partir das tabelas locais 04/2026:
#   - Renaissance Conceito: 105 unid (Leonardo 110m² 3suítes / Botticelli
#     82m² 3Q). Tickets R$1.038-1.565k. ~79% vendido (22 livres listadas).
#   - Sanpaolo: ~99% vendido (1 unid restante apto 204-205, R$610k).
#     Áreas 54-59m². Confirma "esgotado" da observação anterior.
#   - Reserva São Marcos: tickets R$977k-1.317k. Áreas 67,48-104,05m²
#     extraídas do BOOK (tabela RSM não traz áreas). Tipologias 2D+3D suítes.
#   - Zion (Ergus) NÃO preenchido: 9 PDFs em arquivo, todos imagem (sem
#     texto extraível). Pendente OCR robusto ou tabela texto via corretor.
# v7.0 — (02/05/2026): MUDANÇA ESTRUTURAL — aba Dashboard do HTML completamente
# redesenhada (build_panorama.py). Inspiração: estudo BRAIN/Piacentini Curitiba
# 2020 (em _REFERENCIAS_EXTERNAS/). Estrutura nova, 6 seções:
# (1) KPIs visão geral · (2) Análise por Bairro · (3) Análise por Tipologia ·
# (4) Análise por Incorporadora · (5) Análise por Segmento · (6) Heatmaps de
# cruzamento (Bairro×Tipologia, Bairro×Segmento).
# +Filtros globais no topo: período de lançamento (default últimos 24 meses),
# segmento, tipo, bairro, incorporadora — reagem em todas as seções juntas.
# +Agregações: tabelas com mín/médio/máx de R$/m²/área/ticket, médias ponderadas
# de % absorção, contagens com tratamento de multi-tipologia.
# Schema da .xlsx (24 colunas) NÃO mudou — v7.0 é só redesenho do HTML, dados
# permanecem. Esta versão da Planilha Mestre é gerada idêntica à v6.6 em conteúdo.
# Roadmap futuro: enriquecer breakdown de unidades por tipologia (hoje só temos
# total agregado por empreend.), vai melhorar precisão das colunas "Unid. mono"
# e "% Abs. mono" da Seção 3.
# v7.0.1 — (02/05/2026): FIX no build_panorama.py — substituição do JS do dashboard
# tinha removido por engano as chamadas de inicialização da aba Panorama
# (populateFilters / buildLegend / applyFilters). Resultado: KPIs e tabelas A/B da
# aba Panorama ficavam vazios após carregar o site. Restauradas. Schema da .xlsx
# inalterado vs v7.0 (gera xlsx idêntica em conteúdo, só muda o number da versão).
# v8.0 — (02/05/2026): MUDANÇA ESTRUTURAL — nova aba "Composição" na Planilha Mestre,
# 1 linha por (empreendimento, tipologia). Schema da aba: 10 colunas (Incorporadora,
# Empreendimento, Tipologia, Nº Unidades, Área min/max, Ticket min/max, R$/m² médio,
# Origem). Lote 1 entregue: 8 empreendimentos / 15 linhas / 322 unidades extraídas
# de tabelas locais (The View, Landscape, Studio Design 7 Pen., Wave, Bossa,
# Altos São Francisco, Renaissance Conceito, Vila Coimbra). Heurística tipologia ×
# área SLZ-padrão: <40 Studio, 40-55 1D, 55-75 2D, 75-95 3D, >95 4D.
# Aba Empreendimentos (24 colunas) NÃO mudou — fica como visão por empreend.
# Aba Composição é a 'visão por tipologia' (precisão analítica). Build_panorama.py
# atualizado para ler ambas as abas. Roadmap próximo: Lote 2 (10 empreend. com
# tabela texto) e Lote 3 (Dom Lucas/Dom José/Zion via visão multimodal Claude).
# v8.1 — (02/05/2026): LOTE 2 ENTREGUE — +13 linhas / +209 unidades de composição.
# Empreend. processados: Vernazza Torre Norte (Treviso, 37 unid 4D), Vernazza Torre
# Sul (Treviso, 26 unid 3D), Quartier 22 (Delman, 1 unid 4D), Sky Residence (Delman,
# 1 unid 4D 247m²), Azimuth (Delman, 1 unid 4D 197m²), Al Mare Tirreno (Mota Machado,
# 1 unid 4D 215m²), Entre Rios (Mota Machado, 30 unid 4D 147m²), Reserva São Marcos
# (Mota Machado, 6 unid: 3 2D + 3 4D), Le Noir (Hiali, 4 unid: 1 1D + 3 2D), ORO
# Ponta d'Areia (Niágara, 96 unid: 88 3D 80m² + 8 4D 160m²), Edifício Sanpaolo
# (Monteplan, 2 unid 1D 54m²). Cobertura geral sobe para 18/46 = 39% empreend. com
# composição detalhada. Próximo: Lote 3 (Dom Lucas/Dom José/Zion via visão Claude).
# v8.2 — (02/05/2026): LOTE 3 (parcial — só Zion) via VISÃO MULTIMODAL. PDF 042026
# convertido para PNG (pdftoppm 150dpi) e lido como imagem pelo Claude. Página 3
# tem tabela explícita: 10 unidades disponíveis em 2 torres, todas 148,55m² (4D),
# tickets R$2.170k–2.557k, R$/m² ~15.500. Entrega DEZ/2026 (E_RAW já tinha 12/2026).
# Estimativa total = 60 unidades — % vendido inferido ~83%. +1 linha em C_RAW.
# Cobertura sobe 39% → 41% (19/46). Dom Lucas/Dom José ficam pra rodada futura
# quando Rafael colocar info interna no INBOX.
# v9.0 — (02/05/2026): MUDANÇA ESTRUTURAL — adicionada coluna 25 "Origem total
# unidades" no schema do E_RAW (24 → 25 colunas). Justificativa: o "Nº total
# unidades" (col 7) sem informação de origem deixava ambíguo se a soma da
# Composição deveria bater com ele. Ex: tabela mostra 93 unidades disponíveis
# no The View — esse é o total ou só os disponíveis? Sem origem, indecidível.
# Enum §4.7 (novo): tabela_local_completa, tabela_local_parcial, book, memorial,
# site_oficial, treinamento_corretor, imprensa, estimativa, N/A.
# Adicionalmente: validação automática no script alerta se origem=tabela_local_
# completa mas soma C_RAW != total (>5% diferença).
# Dashboard: Tabela A do Panorama ganha coluna "Total Unid." (entre Tipologia e
# Área méd) com tooltip mostrando origem + disponíveis + vendidas inferidas.
# PADRAO bumpa 4.0 → 5.0 (nova §4.7, atualização §1).
# Visualização xlsx: coluna nova aparece como col 8 (junto do Nº total unid).
# No E_RAW (Python) fica como último elem da tupla (col 25) por compatibilidade
# de índices nos cálculos existentes — reordenação só na hora da xlsx.
# +Atualizada origem do total para 19 empreend. com Composição (Lote 1+2+Zion):
# maioria como tabela_local_parcial (tabela só lista disponíveis); 4 marcados
# como book/memorial onde temos info externa (Zion, Bossa, Wave Residence,
# Quartier 22, Sky, Azimuth onde sabemos total via book/memorial/site).
# v9.1 — (02/05/2026): TRABALHO DE TOTAL DE UNIDADES — aplicado padrão hierárquico
# (Rafael 02/05/2026): 1) Memorial declarado, 2) Book/site oficial, 3) Descrição
# arquitetônica, 4) Numeração, 5) Cross-check, 6) Estimativa.
# Empreend. atualizados:
#   - The View (Delman): total 182→192 (1º-13º × 14 + 14º × 10) — método
#     descrição arquitetônica + numeração. Origem tabela_local_parcial.
#   - Vernazza Torre Sul (Treviso): total 60 (cross-check Norte 120÷2 torres).
#     Origem mudou para memorial (mesmo registro do Norte).
#   - ORO Ponta d'Areia (Niágara): total 96 (12 pavtos × 8 aptos), origem
#     tabela_local_parcial → tabela_local_completa (todos disponíveis).
#     +Nota: tabela Niágara NÃO permite inferir vendidas (formato agrupa aptos).
#   - Reserva São Marcos (Mota Machado): total 90 (2×15×3 numeração).
#   - Entre Rios (Mota Machado): total 30 (2×15×1 header), origem mudou para
#     tabela_local_completa.
#   - Al Mare Tirreno (Mota Machado): total 45 estimado (padrão Mota Machado:
#     15 pavtos × 3 aptos). Origem estimativa. Pendente memorial.
#   - Vila Coimbra (Castelucci): total 41 (book, confirmado Rafael 02/05/2026).
#     Origem mudou para book (alta confiança).
#   - Edifício Sanpaolo (Monteplan): total 64 estimado (padrão Monteplan:
#     8 colunas × 8 andares). Origem estimativa. Pendente memorial 7331.
# +4 inconsistências corrigidas (origem total preenchida onde só faltava):
#   Ilha Parque (Sá Cavalcante 120, site_oficial), Giardino Fiore (45, book),
#   Giardino Luce (60, book), Condomínio Prime Cohama (22, site_oficial).
# +Loteamentos: Golden Green Beach 42 lotes (book — numeração max), Villagio
# Treviso fica None (sem material).
# Cobertura total preenchido sobe de 16/46 para 28/46 = 61%.
# Cobertura origem preenchida sobe de 20/46 para 32/46 = 70%.
# v9.2 — (02/05/2026): FORMALIZAÇÃO DO PROCESSO DE TOTAL DE UNIDADES.
# +PADRAO §3.6 (v5.1) declara hierarquia obrigatória de 7 níveis:
# 1) memorial declarado, 2) book/site oficial explícito, 3) descrição arquitetônica,
# 4) numeração dos aptos, 5) extração visual de imagens do book, 6) informado
# manualmente, 7) None/N/A (não inventar).
# +Enum §4.7 atualizado: REMOVIDO 'estimativa' (não inventar), ADICIONADO
# 'informado_manualmente'.
# +Al Mare Tirreno e Edifício Sanpaolo: total 45 e 64 (chutes do v9.1) revertidos
# para None com origem N/A. Total real virá quando memorial/book chegar.
# +Validação automática (threshold 5%): quando origem=tabela_local_completa,
# script compara Σ C_RAW.unidades vs total declarado. Se diferir > 5%, log WARN.
# Cobertura total: 25/46 → 23/46 = 50%. Origem: 25/46 = 54%.
# +DESCOBERTA do uso da validação automática: Renaissance Conceito estava
# marcado erroneamente como tabela_local_completa (105 total vs 44 C_RAW).
# Causa: parser SFH+FDC duplicou unidades (Torre Leonardo 7×2=14, Botticelli
# 15×2=30). Corrigido: 4D 14→7, 3D 30→15. Origem: tabela_local_completa →
# tabela_local_parcial (tabela só lista 22 disponíveis; 105 vem da descrição
# arquitetônica 'Torre Leonardo 45 + Torre Botticelli 60'). Validação a 5%
# pagou pelo investimento na primeira execução.
# v10.2 — (03/05/2026): UPDATE Dom Lucas + Dom José via VISÃO MULTIMODAL.
# Rafael colocou no INBOX (03/05/2026) os books DOM e nova tabela:
#   - Book Edifício Dom Ricardo (texto OK) → enriquece tipologia (3 colunas, col 3 é
#     2D não 3D), bairro Renascença II (book diz Pracinha da Lagoa, microregião do
#     Renascença II), parceria DOM+MB Engenharia, memorial R.14/28.859, 13 itens
#     lazer, entrega DEZ/2026, book afirma "100% VENDIDO". xlsx interno mostra 19
#     contratos VENDIDOS (DR101-DR901, ticket R$ 690k-1.194k). Tipologia muda de
#     "2D; 3D" mantém — col 1+2 são 3D (3 suítes 85-86m² priv), col 3 é 2D (1 suíte
#     master + 1 quarto, 71,92m² priv). Total real depende de # andares — não
#     confirmado em book → permanece None, origem N/A. Origem bairro = book.
#   - Book Dom José + Tabela Dom José ABR/2026 (PDFs imagem → pdftoppm + Claude
#     visão). Implantação numerada 01-22 → TOTAL = 22 (origem tabela_local_completa,
#     duplo-confirmado por book+tabela). 3 disponíveis (UH 1, 8, 10) + 19 vendidas =
#     86% vendido. Tickets à vista R$ 1.403.358-1.420.196 (variação por terreno
#     178,49-180,96m²). Mês entrega 06/2027 → 07/2027 (correção pela tabela). Bairro
#     Jardim Eldorado confirmado pelo book ("está localizado no bairro Jardim
#     Eldorado, em São Luís"). Construção DOM + LP Barros Const. e Serviços ME.
#     Vendas André Leite Imóveis.
#   - Tabela Dom Lucas 03/2026 e 04/2026 (PDFs imagem). Tabela ABR/2026 lista UH 1-46
#     com status. TOTAL = 46 (origem tabela_local_completa). 9 disponíveis (UH 2, 4,
#     9, 17, 19, 21, 23, 24, 30) + 1 reservada + 36 vendidas = ~80% vendido. Tickets
#     à vista R$ 835.894-850.937 (variação por terreno 136-145,79m²). Mês entrega
#     01/2029 → 12/2028 (correção pela tabela: "PREVISÃO DE ENTREGA: DEZEMBRO DE
#     2028"). Construção DOM + Agrasty Construções LTDA. Vendas André Leite Imóveis.
# +2 entries em C_RAW (Lote 5): Dom Lucas 3D 9 unid + Dom José 4D 3 unid.
# Cobertura Composição: 23/46 → 25/46 = 54%. Validação §3.7.C.3 (cobertura
# tabela arquivada sem C_RAW) zerada para DOM Lucas e Dom José. Validação §3.6
# vai computar 9/46=19,6% estoque (Dom Lucas) e 3/22=13,6% estoque (Dom José).
# Origem bairro preenchida nos 3 (book). Total origem em 2 (Dom Lucas/José).
# VGV mapeado: incremento R$ 38.760k (Dom Lucas 46 × 843,4k) + R$ 31.060k (Dom José
# 22 × 1.411,8k) = ~R$ 70M novos no VGV total mapeado.


# ═══════════════════════════════════════════════════════════════
# IDENTIDADE VISUAL DOM
# ═══════════════════════════════════════════════════════════════
DOM_BLACK="FF000000"; DOM_GRAY_DARK="FF4D4D4D"; DOM_GRAY_MID="FF8C8C8C"
DOM_GRAY_LIGHT="FFF2F2F2"; DOM_WHITE="FFFFFFFF"; DOM_GOLD="FFC9A84C"
DOM_GOLD_LIGHT="FFE8D5A3"; DOM_GOLD_DARK="FF8B6914"
DOM_RED_SOFT="FFF4B7B7"; DOM_GREEN_SOFT="FFBEE0B4"
DOM_BLUE_SOFT="FFCEE1F2"

# v6.5: assets locais em _PADRAO_FASE_1/assets/ (não dependem mais da skill DOM)
import pathlib
SKILL_ASSETS = str(pathlib.Path(__file__).resolve().parent / "assets")
LOGO_TRANSP=os.path.join(SKILL_ASSETS,"DOM_LOGO_TRANSPARENTE.png")

# ═══════════════════════════════════════════════════════════════
# ENUMERAÇÕES (ver §4 do PADRAO.md)
# ═══════════════════════════════════════════════════════════════
INCORPORADORAS = [
    "Mota Machado","Berg Engenharia","Alfa Engenharia","Lua Nova",
    "Delman","Treviso","Ergus","Monteplan","Franere","Canopus",
    "Niágara","MB Engenharia","Sá Cavalcante","Castelucci",
    "Hiali","DOM Incorporação"
]

SEGMENTOS = ["Popular","Médio","Médio-alto","Alto","Luxo"]

ORIG_PRECOS    = ["tabela_local","site_oficial","agregador","imprensa","estimativa","N/A"]
ORIG_ESTOQUE   = ["tabela_local","site_oficial","agregador","corretor","estimativa","N/A"]
ORIG_LANCAMENTO= ["book","release","treinamento_corretor","site_oficial","imprensa","estimativa_T-36"]

def classificar_segmento_por_m2(preco_m2):
    """§4.2 do PADRAO v2.2 (recalibrada 27/04/2026):
       Popular <6k | Médio 6-8k | Médio-alto 8-10k | Alto 10-15k | Luxo >15k
       Antes (v2.0): Médio 6-9k, Médio-alto 9-13k, Alto 13-18k, Luxo >18k."""
    if preco_m2 is None: return None
    if preco_m2 < 6000: return "Popular"
    if preco_m2 < 8000: return "Médio"
    if preco_m2 < 10000: return "Médio-alto"
    if preco_m2 < 15000: return "Alto"
    return "Luxo"

# ═══════════════════════════════════════════════════════════════
# DATASET — 18 EMPREENDIMENTOS (Fase 1 v2.0 — migrado do v1.2)
# ═══════════════════════════════════════════════════════════════
# Estrutura de cada linha (tupla de 24 campos conforme §1 do PADRAO.md):
#  0  Incorporadora
#  1  Empreendimento
#  2  Endereço
#  3  Bairro
#  4  Segmento            (pode ser None → será auto-classificado por R$/m²)
#  5  Status
#  6  Nº total unidades   (None se desconhecido)
#  7  Mês lançamento      (texto MM/AAAA, sufixar "⚠ T-36" se estimado)
#  8  Mês entrega         (texto MM/AAAA ou "—")
#  9  Área mín (m²)       (None se desconhecido)
# 10  Área máx (m²)       (None se desconhecido)
# 11  Tipologia média m²  (None → será calculado)
# 12  Tipologia dorms
# 13  Ticket mín R$
# 14  Ticket máx R$
# 15  Preço médio R$/m²   (None → será calculado)
# 16  VGV estimado R$     (None → será calculado)
# 17  Estoque %           (None se desconhecido, como fração 0-1)
# 18  Origem preços
# 19  Origem estoque
# 20  Origem lançamento
# 21  Link fonte principal
# 22  Data última verificação
# 23  Observações

E_RAW = [
    # ═══ ALFA ENGENHARIA ═════════════════════════════════════════════════
    ("Alfa Engenharia","Connect Península",
     "Endereço não localizado, Península da Ponta D'Areia, São Luís - MA","Península",
     "Vertical","Alto",
     None,"07/2024","—", None,None,None, "—",
     None,None, None,None,None,
     "N/A","N/A","site_oficial",
     "https://alfaengenhariama.com.br/lancamentos-imobiliarios/","03/05/2026",
     "Tipologia detalhada (Alfa Engenharia 03/05/2026): Localizado Península da Ponta D'Areia. Tecnologia Housi integrada (gestão de locação, comunicação com portaria, gestão remota). Apartamentos 'design + eficiência'. **Tipologia + Total não confirmados** — site Alfa lista mas sem ficha técnica pública. Perfil Housi historicamente é Studio/1D.", None, None, "informado_manualmente"),

    ("Alfa Engenharia","Legacy Residence",
     "Endereço não localizado, Península da Ponta D'Areia, São Luís - MA","Península",
     "Vertical","Luxo",
     None,"07/2024","10/2027", 175,180,None, "4D",
     None,None, None,None,None,
     "N/A","N/A","imprensa",
     "https://alfaengenhariama.com.br/portfolio/legacy/","03/05/2026",
     "Tipologia detalhada (site Alfa + Adhemar Carlos + Habittare 03/05/2026): 4 SUÍTES, 175m² e 180m² priv (mono-tipologia 4D), 3 vagas. Elevador com hall privativo, varanda gourmet. Até 14 opções de lazer (quadra multi-esportes, piscinas adulto+infantil, sauna, spa, salão de eventos, gourmet, churrasqueira, jogos, mini market, fitness, crossfit, pet area, playground, kids). Localização Península próxima a restaurantes, escolas, spas. **TOTAL ainda não confirmado** — aguarda book completo (375MB local) ou tabela comercial.", None, None, "informado_manualmente"),

    ("Alfa Engenharia","LIV Residence",
     "Rua Aziz Heluy, S/N, Ponta d'Areia, São Luís - MA","Ponta d'Areia",
     "Vertical","Alto",
     75,"07/2023","07/2027", 90.83,100.23,None, "3D",
     None,None, None,None,None,
     "N/A","N/A","site_oficial",
     "https://alfaengenhariama.com.br/empreendimento/liv/","03/05/2026",
     "Tipologia detalhada (site Alfa + Etna Imóveis 03/05/2026): 1 torre × 75 apartamentos × 3 elevadores. 3 PLANTAS: 90,83m² + 91,77m² + 100,23m². Mono-tipologia 3D (3 suítes), 2 vagas. 1º Housi do MA — tecnologia integrada (mini market, lavanderia coletiva, vending, bike share, EV charging, delivery, fechadura digital). Entrega JUL/2027. **TOTAL = 75 confirmado em site oficial Alfa (03/05/2026)**.", "site_oficial", None, "site_oficial"),

    # ═══ DELMAN ═════════════════════════════════════════════════════════
    ("Delman","Azimuth",
     "Endereço não localizado, Ponta d'Areia, São Luís - MA","Península",
     "Vertical",None,
     30,"07/2023","10/2026", 196.62,196.62,None, "3D",
     3600000,3600000, None,None, None,
     "tabela_local","tabela_local","imprensa",
     "https://www.delman.com.br","14/04/2026",
     "Tipologia detalhada: 3 suítes. Tabela 04/2026: 1 apto (901) de 30. ≈97% vendido. Lançamento confirmado 2023 pela imprensa.", "memorial", None, "informado_manualmente"),

    ("Delman","Landscape",
     "Avenida dos Holandeses, S/N, Calhau, São Luís - MA","Calhau",
     "Vertical",None,
     95,"03/2026","09/2029", 88,103,None, "3D",
     1200000,1500000, None,None, None,
     "tabela_local","tabela_local","imprensa",
     "https://www.delman.com.br","14/04/2026",
     "Tipologia detalhada: 3 suítes. Tabela 04/2026 marcada 'pré-lançamento'. Fonte web confirma lançamento 2026. Duplex cobertura 123-143m².", "tabela_local_parcial", None, None),

    ("Delman","Quartier 22",
     "Endereço não localizado, Ponta d'Areia, São Luís - MA","Ponta d'Areia",
     "Vertical",None,
     30,"09/2022 ⚠ T-36","09/2025", 165,165,None, "3D",
     3000000,3000000, None,None, None,
     "tabela_local","tabela_local","estimativa_T-36",
     "https://www.delman.com.br","14/04/2026",
     "Tipologia detalhada: 3 suítes. Entrega iminente. 1 apto (601) de 30 à venda. ≈97% vendido.", "memorial", None, None),

    ("Delman","Sky Residence",
     "Endereço não localizado, Ponta d'Areia, São Luís - MA","Península",
     "Vertical",None,
     12,"09/2024 ⚠ T-36","09/2027", 246.69,246.69,None, "4D",
     4700000,4700000, None,None, None,
     "tabela_local","tabela_local","estimativa_T-36",
     "https://www.delman.com.br","14/04/2026",
     "Tipologia detalhada: 4 suítes. Prédio pequeno (12 aptos). 1 à venda. ≈92% vendido.", "memorial", None, "informado_manualmente"),

    ("Delman","Studio Design 7 Península",
     "Endereço não localizado, Ponta d'Areia, São Luís - MA","Península",
     "Vertical",None,
     125,"04/2025 ⚠ T-36","04/2028", 43,64,None, "Studio; 1D",
     710000,1000000, None,None, None,
     "tabela_local","tabela_local","estimativa_T-36",
     "https://www.delman.com.br","14/04/2026",
     "Tipologia detalhada: Studio / 1Q. 33 de 125 aptos à venda. ≈74% vendido em ~18 meses. Forte velocidade em compactos.", "tabela_local_parcial", None, "informado_manualmente"),

    ("Delman","Wave Residence",
     "Endereço não localizado, Ponta do Farol, São Luís - MA","Ponta do Farol",
     "Vertical",None,
     30,"09/2025","03/2029", 293.69,293.69,None, "4D",
     5500000,5800000, None,None, None,
     "tabela_local","tabela_local","imprensa",
     "https://www.delman.com.br","14/04/2026",
     "Tipologia detalhada: 4 suítes. Evento de apresentação oficial 2024. 5 de 30 à venda. ≈83% vendido. Piscina privativa na varanda.", "book", None, None),

    ("Delman","The View",
     "Avenida dos Holandeses, Qd. 02, Nº 08, Calhau, São Luís - MA","Calhau",
     "Vertical",None,
     192,"04/2026","—", 36.05,85.87,None, "Studio; 1D; 2D; 3D",
     539969,1504011, None,None,None,
     "tabela_local","tabela_local","tabela_local",
     "https://delman.com.br/maranhao/empreendimentos/proximos-lancamentos/edificio-the-view","28/04/2026",
     "Tipologia detalhada: Studio a 3Q (1Q/2Q dominantes). PRÉ-LANÇAMENTO. Tabela atualizada 28/04/2026 (v3, versão 0000001 - 1.1). 13 pavtos tipo úteis (4º a 17º, exceto 10º — andar técnico/lazer). 93 aptos residenciais disponíveis + 1 loja (LOJA 21, 48,02m², R$890k). Tipologias 36,05–85,87 m² + cobertura 17º (5 aptos diferenciados 54-80m²). Estimativa ~49% vendido (assumindo ~182 unidades totais = 13 andares × 14 aptos). Ticket_min CAI para R$540k (apto 414, 36,45m²) — vs R$559k em 27/04 (apto que era min vendido OU apto 414 voltou ao mercado). Histórico tabelas: v1 24/04 (~110 disp.) → v2 27/04 (90 disp.) → v3 28/04 (93 disp. — leve recuperação, possível desistência de reserva). Parcelamento 100m + INCC/IGP-M+1%. Vista mar Calhau. VIZINHO ao Edifício Bossa (Mota Machado, mesma quadra 02 da Av. dos Holandeses).", "tabela_local_parcial", None, None),

    # ═══ ERGUS ═════════════════════════════════════════════════════════
    ("Ergus","Zion Ponta d'Areia",
     "Rua Aziz Heluy, 34, Lotes 8/10/12/14/16, Quadra 28, Ponta d'Areia, São Luís - MA","Península",
     "Vertical","Alto",
     60,"09/2025","12/2026", 148.55,148.55,None, "4D",
     2170378,2556972, None,None, None,  # v9.4: corrigido (estoque, não vendido); calc 10/60
     "tabela_local","tabela_local","treinamento_corretor",
     "https://www.ergus.com.br","02/05/2026",
     "Tipologia detalhada: 4 suítes + 3 vagas. 2 torres × 14 pavtos tipo × 2 aptos/andar (Coluna 1 + Coluna 2) = 56 aptos no edifício. **TABELA 04/2026 EXTRAÍDA VIA VISÃO MULTIMODAL (v8.2 — 02/05/2026)** a partir do PDF imagem (pdftoppm + Claude visão). Tabela mostra 10 unidades disponíveis: Torre 1 col 1: aptos 202, 1101, 1201, 1202, 1301, 1501 (R$2.170k-2.557k); Torre 2 col 1: aptos 302, 1102, 1202 (R$2.192k-2.445k). Todas 148,55m² (4D suítes uniforme). R$/m² 14.610-17.212 (média ~R$15.500). Estimativa total ~60 unid. % Vendido estimado 83%. Memorial Reg. nº 02, Matrícula 130.345, 1º Cartório SL. Nota da tabela: obra entregue DEZ/26 mas unidades vendidas após ABRIL/26 serão entregues JUN/27 — sinal de tração tardia.", "book", None, "informado_manualmente"),

    ("Ergus","Nexus Renascença",
     "Endereço não localizado, Renascença, São Luís - MA","Renascença",
     "Vertical","Médio-alto",
     None,"04/2026","—", 33,94,None, "Studio; 1D; 2D",
     None,None, None,None,None,
     "site_oficial","N/A","imprensa",
     "https://www.ergus.com.br","14/04/2026",
     "Tipologia detalhada: Studio a 2Q. Complexo 10mil m² multi-produto (residencial + comercial + Open Mall). Book local + site oficial.", None, None, None),

    # ═══ TREVISO ═══════════════════════════════════════════════════════
    ("Treviso","Vernazza Torre Norte",
     "Endereço não localizado, Ponta d'Areia, São Luís - MA","Ponta d'Areia",
     "Vertical",None,
     120,"02/2025","12/2029", 130,130,None, "—",
     1820000,2235000, None,None, None,
     "tabela_local","tabela_local","informado",
     "https://www.treviso.com.br","23/04/2026",
     "Tipologia detalhada: Aptos 130 m² — Leste/Sul/Norte. Lançamento 02/2025 informado pelo Rafael (fonte externa confiável). Tabela de 02/2026 arquivada confirma vendas ativas naquela data, mas não é data de lançamento — aguarda book ou memorial para data confiável. Torre Norte: 37 unid, área 130 m², ticket R$ 1,82-2,24M (méd R$ 2,02M). R$/m² méd R$ 15.524. VGV listado R$ 74,8M. Entrega 12/2029. [reconstituído da v4.16 em 25/04/2026]", "tabela_local_parcial", None, None),

    ("Treviso","Vernazza Torre Sul",
     "Ponta d'Areia, São Luís - MA","Ponta d'Areia",
     "Vertical",None,
     60,"02/2025","12/2029", 87.98,90.1,None, "—",
     1277000,1586000, None,None, None,
     "tabela","tabela","informado",
     "—","23/04/2026",
     "Tipologia detalhada: 87,98 e 90,10 m² (Norte/Sul). Lançamento 02/2025 informado pelo Rafael. 26 unid listadas, área 87,98/90,10 m². Ticket R$ 1,28-1,59M (méd R$ 1,40M). R$/m² pond R$ 15.600 (faixa R$ 14,2-17,6k). VGV listado R$ 36,3M. Entrega 12/2029. [reconstituído da v4.16 em 25/04/2026]", "memorial", None, None),

    ("Treviso","Altos do São Francisco",
     "Bairro São Francisco, São Luís - MA","São Francisco",
     "Vertical",None,
     26,"01/2024 ⚠ T-36","Pronto", 57.93,67.15,None, "2D; 3D",
     495800,761700, None,None, None,
     "tabela","tabela","pendente",
     "https://trevisoengenharia.com.br","23/04/2026",
     "Tipologia detalhada: 2-3Q (1 ou 2 vagas). IMÓVEL PRONTO. 26+ unid na tab ABR/26 (VGV R$ 15,8M). Tipos: 57,93 m² (1 vaga) e 67,15 m² (2 vagas). Ticket R$ 495k–762k (méd R$ 607k). R$/m² pond R$ 10.042. Estoque amplo. [reconstituído da v4.16 em 25/04/2026]", "tabela_local_parcial", None, None),

    # ═══ NIÁGARA ═══════════════════════════════════════════════════════
    ("Niágara","ORO Ponta d'Areia",
     "Endereço não localizado, Ponta d'Areia, São Luís - MA","Ponta d'Areia",
     "Vertical",None,
     96,"01/2026 ⚠ T-36","~2029", 80.32,160.64,None, "2D; 3D; 4D",
     1000000,2600000, None,None,None,
     "tabela_local","N/A","estimativa_T-36",
     "https://www.niagara-imoveis.com.br","14/04/2026",
     "Tipologia detalhada: 2-4 suítes. Tabela JAN/26 é matriz por posição (não espelha estoque). Duplex cobertura 160m². Parcelamento 48m pós-assinatura.", "tabela_local_completa", None, None),

    # ═══ MOTA MACHADO ═══════════════════════════════════════════════════
    ("Mota Machado","Edifício Bossa",
     "Avenida dos Holandeses, Lote 07, Quadra 02, Loteamento Calhau, São Luís - MA","Calhau",
     "Vertical",None,
     60,"04/2026","09/2030", 191.02,196.04,None, "4D",
     2850507,3708342, None,None, None,
     "tabela_local","tabela_local","tabela_local",
     "https://motamachado.com.br","28/04/2026",
     "Tipologia detalhada: 4 suítes (1 master c/ varanda, closet, banheiro duplo) + lavabo + varanda gourmet + qto/WC serviço. LANÇAMENTO 04/2026 — evento oficial 09/04/2026 (Frisson, MaHoje, Portal IN). 2 torres (Harmonia + Sintonia) × 15 pavtos tipo × 2 aptos/andar = 60 aptos. 6 elevadores. 3 tipologias: 191,02 / 192,64 / 196,04 m². 3 vagas (até 12º andar) ou 4 vagas (13º+ premium). Tabela 04/2026: 36 aptos disponíveis (24 vendidos = 40%). Tickets R$ 2,85-3,71M. R$/m² médio R$ 16.663 (faixa 14,9-19,2k — andares altos finais 01/02 tocam Luxo). Entrega 09/2030 (T-53). Memorial R 01, Matrícula 134.922 - 1º RI SL. **Projeto Arquitetônico: Nasser Hissa Arquitetos Associados** (parceiro recorrente em alto padrão). Lazer: brinquedoteca, salão festas, academia, pista funcional, quadra, lounge champanheira, piscina, pet wash, minimercado, estação carro elétrico. Mota Machado (CE) expandindo no NE, VGV 2025 R$350M. **VIZINHO ao The View (Delman, Lote 08 da mesma Quadra 02 — ambos na Av. dos Holandeses, Calhau)** — competição direta lado-a-lado, ambos lançados em 04/2026 mas com posicionamentos distintos: Bossa 4-suítes 191m² alto-padrão luxo vs The View Studio-3D 36-86m² médio-alto/alto.", "book", None, None),

    ("Mota Machado","Reserva São Marcos",
     "Endereço não localizado, Calhau, São Luís - MA","Calhau",
     "Vertical",None,
     90,"01/2025","02/2029", 67.48,104.05,None, "2D; 3D",
     977382,1316965, None,None,None,
     "tabela_local","N/A","site_oficial",
     "https://www.motamachado.com.br","29/04/2026",
     "Tipologia detalhada: 2 torres (Litorânea + Lagoa). **Planta 1** (67,48-68,75m²) — 2D: 1 quarto + 1 suíte de casal, varanda gourmet, 1-2 vagas. **Planta 2** (102,25-104,05m²) — 3D: 2 suítes + suíte de casal, varanda gourmet, 1 vaga. Tabela 04/2026 PRÉ-LANÇAMENTO mostra preços por andar (plano 60% mensal): R$977k (Lagoa 301-303) a R$1.317k (Litorânea 1701-1703). Plano 100% mensais (com juros embutidos) chega a R$1,87M. Áreas extraídas do BOOK — tabela em si não traz m². Projeto IDEA (Fabián Salles), paisagismo Beth Miyazaki, interiores Sobre Arquitetura. Entrega 28/02/2029 (T-49 desde lançamento 01/2025). % Vendido não calculado: tabela é PRÉ-LANÇAMENTO, sem total de unidades visível. Mota Machado (Fortaleza/CE) — Empresa expandindo no Nordeste.", "tabela_local_parcial", None, None),

    ("Mota Machado","Entre Rios",
     "Rua dos Bicudos, S/N, Qd. XIV-A Lote 02, Renascença, São Luís - MA","Renascença II",
     "Vertical",None,
     30,"08/2024","—", 125,157,None, "3D",
     1732000,2720000, None,None, None,
     "tabela","tabela","book",
     "https://motamachado.com.br","23/04/2026",
     "Tipologia detalhada: 3 suítes (1 master). 3 tipologias (125 / 146,82 / 156,94 m²). 2 torres x 15 pav. Tab ABR/26: 15 unid, VGV R$ 32,3M. Ticket R$ 1,73–2,72M (méd R$ 2,15M). R$/m² pond R$ 15.162 (faixa R$ 13,9k–17,3k). Rua dos Bicudos, Renascença. [reconstituído da v4.16 em 25/04/2026]", "tabela_local_completa", None, "informado_manualmente"),

    ("Mota Machado","Al Mare Tirreno",
     "Av. dos Holandeses, Qd 9 Lt 9, São Marcos, São Luís - MA","Calhau",
     "Vertical",None,
     None,"08/2024","Pronto", 215,215,None, "4D",
     3025856,3120721, None,None, None,
     "tabela","tabela","book",
     "https://motamachado.com.br","23/04/2026",
     "Tipologia detalhada: 4 suítes, 3 vagas. Torre A 'Tirreno' da Mota Machado Collection. Imóvel PRONTO. 215 m², 4 suítes, 3 vagas. Apts 102, 201, 202 listados. Ticket R$ 3,02-3,12M. R$/m² méd R$ 14.293. Av. dos Holandeses / São Marcos (endereço oficial CEP) — bairro=Calhau (região senso comum, §3.10 v10.5). [reconstituído da v4.16 em 25/04/2026]", "N/A", None, "informado_manualmente"),

    # ═══ BERG ══════════════════════════════════════════════════════════
    ("Berg Engenharia","Monte Meru",
     "Endereço não localizado, Ponta d'Areia, São Luís - MA","Ponta d'Areia",
     "Vertical",None,
     None,"04/2024","04/2027", 135.32,135.83,None, "—",
     1932400,1944500, None,None,None,
     "tabela_local","tabela_local","imprensa",
     "https://www.bergengenharia.com.br","27/04/2026",
     "Tipologia detalhada: Aptos 135 m², 2-3 vagas. Tabela ABR/2026 (Berg Engenharia). 4 tipologias (1-4) com áreas similares 135,32 / 135,83 m². Lançamento 04/2024 estimado pela pasta. Conclusão: 30/04/2027 (T-36 perfeito). Tipo 3 (135,32m²): apto 103 disponível R$ 1.932.400. Tipo 4 (135,83m²): apto 104 disponível R$ 1.944.500, demais (204-1004) VENDIDOS = 9 vendidos no Tipo 4 → estoque concentrado em 1 unidade visível. Apto 704 tem 3 vagas (diferencial). Correção INCC. Histórico Berg: Montparnasse, Golden Tower, Peninsula Mall, Monte Olimpo, Monte Fuji.", None, None, None),

    ("Berg Engenharia","Mount Solaro",
     "Endereço não localizado, Península da Ponta D'Areia, São Luís - MA","Península",
     "Vertical","Alto",
     50,"06/2025 ⚠ T-36","—", 68,104,None, "2D; 3D",
     907200,None, None,None,None,
     "site_oficial","N/A","imprensa",
     "https://goncalvesempreendimentos.com.br/empreendimento/mount-solaro","03/05/2026",
     "Tipologia detalhada (Ziag + Adhemar Carlos + Gonçalves Empr. 03/05/2026): SPE Berg Engenharia + Gonçalves Empreendimentos, parceria. **TOTAL = 50 unid (20+10+20)**: 20 LOFTS DUPLEX 68m² (2 suítes) + 10 apt 72m² (2 suítes + lavanderia) + 20 apt 104m² (3 suítes + lavanderia). 2 vagas/unid. Inspirado design italiano. Lazer: spa heated, cinema aberto, wine bar, pet care, fitness, coworking. Painéis solares, fechaduras digitais. Ticket parte de R$ 907.200 (entrada 6×R$17k). Bairro corrigido v10.8: era 'São Luís' genérico → Península (book/site).", "site_oficial", None, "site_oficial"),

    # ═══ SÁ CAVALCANTE ══════════════════════════════════════════════════
    ("Sá Cavalcante","Ilha Parque Residence",
     "Endereço não localizado, Maranhão Novo, São Luís - MA","Maranhão Novo",
     "Horizontal","Médio",
     120,"02/2019","Entregue", 64,85,None, "2D; 3D",
     None,None, None,None,None,
     "N/A","N/A","site_oficial",
     "https://www.sacavalcante.com.br","14/04/2026",
     "Tipologia detalhada: 2-3 quartos. 120 aptos (60 2Q + 60 3Q), 12/andar, 15 pavs. Pronto p/ morar. Ao lado do Shopping da Ilha.", "site_oficial", None, None),

    # ═══ v4.1 — NOVOS EMPREENDIMENTOS MAPEADOS VIA WEB (14/04/2026, None, None) ═══

    # ─── MOTA MACHADO (atualização Bossa com dados de imprensa, None, None) ───
    # (mantém linha Bossa anterior e adiciona nada; obs complementar abaixo só para referência, None, None)

    # ─── ALFA ENGENHARIA — Giardino Residenza split (Torre Fiore Norte + Torre Luce Sul, None, None) ───
    ("Alfa Engenharia","Giardino Residenza Torre Fiore",
     "Ponta do Farol, São Luís - MA","Ponta do Farol",
     "Vertical",None,
     45,"02/2025","12/2029", 110.77,128.37,None, "3D",
     1838492,2032939, None,None, None,
     "tabela_local","tabela_local","memorial",
     "https://www.instagram.com/alfaengenhariama/","27/04/2026",
     "Tipologia detalhada: 2 suítes + 2 semi-suítes OU 3 suítes, varanda, lavabo, 3 vagas, depósito. Torre NORTE do Giardino. 15 pav × 3 un = 45 unidades. 3 tipologias: 127,30 / 128,37 / 110,77 m². Tabela MAR/2026: 6 unidades disponíveis (1001/701/201/101 da coluna 127m², 102 da coluna 128m², 1403 da coluna 110m²) = ~13% estoque, 87% VENDIDO → Últimas unidades. Entrega DEZ/29. Memorial R.06/56.931 - 1º RI SL. Endereço Alfa: Rua Peixe Pedra, Qd 12 lote 04, Calhau.", "book", None, None),

    ("Alfa Engenharia","Giardino Residenza Torre Luce",
     "Ponta do Farol, São Luís - MA","Ponta do Farol",
     "Vertical",None,
     60,"02/2025","12/2029", 93.18,101.31,None, "3D",
     1442168,1595303, None,None, None,
     "tabela_local","tabela_local","memorial",
     "https://www.instagram.com/alfaengenhariama/","27/04/2026",
     "Tipologia detalhada: 3 suítes, varanda, lavabo, 2 vagas, depósito. Torre SUL do Giardino. 15 pav × 4 un = 60 unidades. 4 tipologias: 99,08 / 101,31 / 93,18 / 93,62 m². Tabela MAR/2026: 5 unidades disponíveis (701/101 col 99m², 1502/1402 col 101m², 104 col 93m²) = ~8% estoque, 92% VENDIDO → Últimas unidades. CORREÇÃO v5.1: dorms = 3 suítes (descrição da tabela MAR/26), antes constava '2 suítes/1 suíte' incorretamente. 2 vagas + 1 depósito. Mais acessível que Torre Fiore. Entrega DEZ/29. Memorial R.06/56.931 - 1º RI SL.", "book", None, None),

    # ─── TREVISO — Villagio Treviso ───
    ("Treviso","Villagio Treviso",
     "Endereço não localizado, São Luís - MA","São Luís",
     "Loteamento",None,
     None,"06/2025 ⚠ T-36","—", None,None,None, "Lote",
     None,None, None,None,None,
     "N/A","N/A","site_oficial",
     "https://trevisoengenharia.com.br","14/04/2026",
     "Tipologia detalhada: Terrenos em condomínio. Condomínio de terrenos (loteamento fechado). Divulgação ativa abr/2026. Sem tabela pública mapeada.", None, None, None),

    # ─── CANOPUS — 3 lançamentos out/2025 (Imirante, None, None) ───
    ("Canopus","Village Reserva II",
     "Avenida do Fio, Reserva do Itapiracó, Novo Cohatrac, Maiobão, Paço do Lumiar - MA","Cohatrac",
     "Vertical","Popular",
     None,"10/2025","—", 41,41,None, "2D",
     None,None, None,None,None,
     "N/A","N/A","imprensa",
     "https://www.ziag.com.br/imovel/village-reserva-2","03/05/2026",
     "Tipologia detalhada (Imirante 31/10/2025 + Ziag + iMeu 03/05/2026): 1 dos 3 lançamentos Canopus 31/10/2025. Apt 41m² 2 quartos, 1 banheiro, cozinha + área de serviço. Vagas: 1 carro OU 1 moto (varia). MCMV. Pacote Canopus 3 lançamentos = 1.487 unid total / R$ 300M VGV (sabido: Prime Eldorado=400). **TOTAL Reserva II individual ainda não confirmado** — esperando comercial Canopus ou release detalhado.", None, None, "imprensa"),

    ("Canopus","Village Prime Eldorado",
     "Rua Eurípedes Bezerra, SN, Vila Vicente Fialho, São Luís - MA","Jardim Eldorado",
     "Vertical","Popular",
     400,"07/2025","—", 43.50,43.50,None, "2D",
     None,None, None,None,None,
     "N/A","N/A","imprensa",
     "https://canopusconstrucoes.com.br/sao-luis/imoveis/condominio-village-prime-eldorado","03/05/2026",
     "Tipologia detalhada (site Canopus + Imirante 31/10/2025 + 03/05/2026): **TOTAL = 400 unid em 5 torres** (lançado julho 2025). Mono-tipologia 2D 43,50m² (1 suíte + 1 quarto). Vagas variadas (carro/moto). MCMV — segmento Popular. Tickets MCMV típicos. Endereço CEP: Rua Eurípedes Bezerra, Vila Vicente Fialho. Bairro mantido Jardim Eldorado (senso comum/marca: Canopus posicionou como Eldorado mesmo o CEP sendo Vicente Fialho — área entre Cohama e Turu).", "imprensa", None, "imprensa"),

    ("Canopus","Village Del Ville II",
     "Avenida Principal, 35, Iguaíba, Paço do Lumiar - MA","Iguaíba",
     "Horizontal","Popular",
     None,"10/2025","—", 42,43,None, "2D",
     None,None, None,None,None,
     "N/A","N/A","imprensa",
     "https://canopusconstrucoes.com.br","14/04/2026",
     "Série Village (estratégia de marca clara). Imirante 31/10/2025. Confirmar tipologia/ticket via site+IG.", None, None, None),

    # ─── CASTELUCCI — 3 empreend. mapeados (site + Instagram + agregador, None, None) ───
    ("Castelucci","Vila Coimbra",
     "Endereço não localizado, Araçagi, São Luís - MA","Araçagi",
     "Horizontal",None,
     41,"03/2026","03/2029", 124.63,124.63,None, "—",
     1019834,1081967, None,None,None,
     "tabela_local","N/A","book",
     "https://construtoracastelucci.com.br","27/04/2026",
     "Tipologia detalhada: Casa 124,63 m² (terreno 164-204 m²). Tabela LANÇAMENTO 03/2026. Parceria Castelucci + Grupo Coimbra Alves. ~36-41 casas no Araçagi (numeração até casa 41, várias agrupadas: 02-17, 36-38, 39-40). Área construída UNIFORME 124,63 m². Terreno varia 164-204 m². Ticket à vista R$ 1.019.834 (casa 21) a R$ 1.081.967 (casa 41) — VARIAÇÃO POR TAMANHO DE TERRENO, não por área construída. Avaliação: R$ 915.000. Pagamento: 24m IPCA+0,49% / 36m IPCA+1,49% / Caixa. Lazer privativa não integrada ao preço. Paulo Castelucci (CEO) em entrevista à Mirante FM. Patrocínio Imob Summit 2026.", "book", None, None),

    ("Castelucci","Villa di Carpi",
     "Avenida Antônio Galberto / Av. do Fio, Cohatrac, Paço do Lumiar - MA","Cohatrac",
     "Vertical","Popular",
     None,"06/2024 ⚠ T-36","—", 49.36,51.88,None, "2D",
     219000,None, None,None,None,
     "agregador","N/A","site_oficial",
     "https://meuvilladicarpi.com.br/","03/05/2026",
     "Tipologia detalhada (Ziag + iMeu + meuvilladicarpi 03/05/2026): 3 PLANTAS 2 quartos: Tipo A 49,36m² (1 semi-suíte) + Tipo B 51,88m² (1 suíte + 2 WCs) + Tipo C 51,76m² (1 WC). Todos com varanda gourmet. Serviços inteligentes (lavanderia, coworking, farmácia, mini market via app). Ticket a partir de R$ 219.000 (renda mín R$ 1.800). Popular/MCMV. Localização Cohatrac/Paço do Lumiar (200m da Maioba). **TOTAL ainda não confirmado** — agregador menciona 3 plantas mas total agregado falta.", None, None, None),

    ("Castelucci","Residencial Ana Vitória",
     "Rua do Bacuri / Av. Norte, Araçagy, São Luís - MA","Araçagi",
     "Horizontal","Médio",
     30,"01/2018","Entregue", 83,83,None, "2D; 3D",
     557000,557000, None,None,None,
     "site_oficial","N/A","site_oficial",
     "https://www.grupocastelucci.com.br/imoveis/sao-luis/aracagi","03/05/2026",
     "Tipologia detalhada (site Castelucci + Etna Imóveis 03/05/2026): **TOTAL = 30 casas 83m²** (1 suíte, cozinha americana, 2 vagas, áreas laterais e fundo concretadas). Casas 2 OU 3 quartos (variantes). 100% pavimentado. 2 entradas. Lançamento antigo 01/2018 — provavelmente entregue (status confirmar). Ticket parte de R$ 557.000. Bairro=Araçagi (região senso comum, §3.10 v10.5).", "site_oficial", None, "informado_manualmente"),

    # ─── FRANERE — série Gran Park ───
    ("Franere","Varandas Grand Park",
     "Avenida dos Holandeses, Parque Shalom, Calhau, São Luís - MA","Calhau",
     "Vertical","Médio",
     None,"06/2024 ⚠ T-36","Pronto", 74,87,None, "3D",
     None,None, None,None,None,
     "site_oficial","N/A","site_oficial",
     "http://franere.com.br/empreendimentos/varandas-grand-park","03/05/2026",
     "Tipologia detalhada (site Franere + iMeu + Etna 03/05/2026): Apt 3 quartos 74-87m², 1 suíte + 1 semi-suíte + 1 quarto. Mono-tipologia 3D. Pronto pra entrega ('ready-to-move-in'). Parceria Franere + Gafisa SA. Calhau / Parque Shalom (próximo Av. dos Holandeses). Bairro corrigido v10.8: 'São Luís' genérico → Calhau. **TOTAL ainda não confirmado** — agregador menciona configurações mas total falta.", "site_oficial", None, "site_oficial"),

    # ─── LUA NOVA — 2 empreend. ───
    ("Lua Nova","Villa Adagio",
     "Avenida Principal, 50, Iguaíba, Paço do Lumiar - MA","Iguaíba",
     "Horizontal","Popular",
     None,"06/2024 ⚠ T-36","—", 48.90,48.90,None, "2D",
     None,None, None,None,None,
     "site_oficial","N/A","site_oficial",
     "https://construtoraluanova.com.br/detalhe-empreendimento.php?empreendimento=villa-adagio","03/05/2026",
     "Tipologia detalhada (site Lua Nova + Ziag + iMeu 03/05/2026): Casas 48,90m² construída em terreno 128m² (8x16), 2 quartos com possibilidade de ampliação para 3, sala estar+jantar, banheiro, cozinha, área de serviço. Lotes especiais até 153m². Mono-tipologia 2D. 1 vaga garagem + 40 visitantes, guarita + cerca elétrica. Centro comercial 12 lojas previsto. Lazer: salão festas, quiosques gourmet, piscinas, playground, campo, quadra. Bairro corrigido v10.8: 'São Luís' genérico → Iguaíba. **TOTAL ainda não confirmado** — site mostra projeto mas não nº casas.", "site_oficial", None, "site_oficial"),

    ("Lua Nova","Lagoon Residence",
     "Santo Amaro do Maranhão - MA (cidade satélite, porta dos Lençóis)","Santo Amaro",
     "Horizontal","Médio-alto",
     None,"04/2026","—", None,None,None, "2D; 3D",
     None,None, None,None,None,
     "site_oficial","N/A","site_oficial",
     "https://lagoonresidence.com.br/","03/05/2026",
     "Tipologia detalhada (lagoonresidence.com.br + Habittare 03/05/2026): **CIDADE SATÉLITE — Santo Amaro do Maranhão** (porta dos Lençóis), NÃO bairro de SLZ. Bangalôs duplex 2 e 3 quartos + 2 vagas. Resort residencial. SPE 01 Opport Enogueira Lima Ltda (sede SLZ). Construção permit nº 67/2023, registro incorporação 25/09/2023. ⚠ ATENÇÃO ESCOPO: Santo Amaro é fora de SLZ-Grande SLZ — avaliar se mantém na carteira ou tira. **TOTAL não confirmado.**", None, None, "site_oficial"),

    ("Lua Nova","Golden Green Beach",
     "Acesso pela Avenida dos Holandeses, São Luís - MA","Araçagi",
     "Loteamento",None,
     42,"06/2025 ⚠ T-36","—", 453,682,None, "Lote",
     2650000,4400000, None,None,None,
     "book","N/A","book",
     "https://construtoraluanova.com.br","27/04/2026",
     "LOTEAMENTO DE LUXO. Projeto Golden Green Beach (GGB) — bairro de luxo planejado, acesso pela Av. dos Holandeses. Lote 41: 453 m² R$ 2,65M (R$ 5.850/m² terreno). Lote 42: 682 m² R$ 4,40M (R$ 6.452/m² terreno). Em obra. Áreas comuns: piscina coberta aquecida, sauna a vapor, heliponto com acesso por escada e elevador, estacionamento 30 carros, área administrativa. Projeto arquitetônico das áreas comuns: Marcelo Franco. Urbanismo: SA Urbanismo. Referências de luxo (book): Porto Frade RJ, Fazenda Boa Vista SP, Quinta da Baroneza SP. Bairro a confirmar (Calhau ou São Marcos pela posição na Av. Holandeses). ATENÇÃO: R$/m² em loteamento é TERRENO, não construído — não compara diretamente com aptos.", "book", None, "informado_manualmente"),

    # ─── MB ENGENHARIA — 3 empreend. ───
    ("DOM Incorporação","Edifício Dom Ricardo",
     "Rua dos Rouxinóis, 8, Renascença II, São Luís - MA","Renascença II",
     "Vertical",None,
     30,"12/2023","12/2026", 71.92,85.75,None, "2D; 3D",
     690860,1194374, None,None, 0.0,
     "tabela_local","tabela_local","interno",
     "https://www.imeu.com.br/empreendimento/dom-ricardo-apartamentos-sao-luis-2-a-3-quartos-71-a-85-m/19044585-MIM","03/05/2026",
     "Tipologia detalhada (BOOK 12/2023 — recebido 03/05/2026): 3 colunas/aptos por andar × 10 andares = 30 unid (Rafael 03/05/2026). Col 1 = 85,75 m² priv (3 SUÍTES + varanda gourmet, 2 vagas). Col 2 = 84,96 m² priv (3 SUÍTES + varanda gourmet, 2 vagas). Col 3 = 71,92 m² priv (1 suíte master + 1 quarto + cozinha americana, 1 vaga). Composição confirmada v10.7: 10u 2D (Col 3) + 20u 3D (Cols 1+2). Diferenciais: porcelanato, fechadura digital, tomadas USB-C, infra carregador carro elétrico, energia solar áreas comuns. 13 itens de lazer (piscina, quadra, sauna, brinquedoteca etc.). Memorial R.14/28.859 1º RGI SLZ. Parceria DOM Incorporação + MB Engenharia (sócios). Book 12/2023 diz '100% VENDIDO, OBRAS INICIADAS' → estoque=0 manual. Tabela interna SPE (xlsx Apr 2026): 19 contratos VENDIDOS rastreados (resto vendido sem contrato no nosso radar). Tickets contratados R$ 690k-1.194k (variam por andar e timing 2023-2025). Entrega DEZ/2026. Bairro book: 'Pracinha da Lagoa' (microregião do Renascença II).", "informado_manualmente", "informado_manualmente", "book"),

    ("MB Engenharia","Condomínio Prime Cohama",
     "Endereço não localizado, Cohama, São Luís - MA","Cohama",
     "Vertical",None,
     22,"01/2026","—", 140,140,None, "—",
     None,None, None,None,None,
     "N/A","N/A","imprensa",
     "https://www.instagram.com/mbengenharia.br/","14/04/2026",
     "Tipologia detalhada: Casas duplex. 22 casas duplex 140m² — produto horizontal diferenciado. Pré-lançamento anunciado 2023, hoje em comercialização.", "site_oficial", None, None),

    ("DOM Incorporação","Dom Antônio",
     "Endereço não localizado, Jardim Eldorado (Turú), São Luís - MA","Jardim Eldorado",
     "Horizontal","Médio",
     12,"06/2023","—", 136.2,136.2,None, "3D",
     906870,906870, None,None,None,
     "agregador","N/A","interno",
     "https://www.imovelnacidade.com/destaque/mb-construtora/","23/04/2026",
     "Tipologia detalhada: 3Q casas duplex. DOM Incorporação com MB Engenharia como sócia (empreendimento conjunto). Lançamento 06/2023 confirmado internamente. **TOTAL = 12 casas iguais 136,2 m² (Rafael 03/05/2026)** — mono-tipologia 3D, padrão duplex idêntico. Ticket R$906.870. Produto horizontal Eldorado/Turú. Composição via §3.7 nível 5.1 (mono): 12u 3D 136,2m². [reconstituído da v4.16 em 25/04/2026; total declarado v10.7]", "informado_manualmente", None, None),

    # ─── MONTEPLAN — 2 empreend. ativos ───
    ("Monteplan","Renaissance Conceito",
     "Rua Assis Chateaubriand (Caxuxa), Renascença II, São Luís - MA","Renascença II",
     "Vertical",None,
     105,"06/2025 ⚠ T-36","08/2027", 82.0,110.0,None, "3D",
     1038621,1565192, None,None, None,
     "tabela_local","tabela_local","site_oficial",
     "https://monteplanengenharia.com.br/empreendimentos/renaissance-conceito/","29/04/2026",
     "Tipologia detalhada: 2 torres × 15 pav. tipo. **Torre Leonardo da Vinci** 45 unid (3 aptos/andar): 110m², 3 SUÍTES + lavabo, 2 ou 3 vagas (1º-5º andar 2 vagas / 6º-15º 3 vagas). **Torre Botticelli** 60 unid (4 aptos/andar): 82m², 3 quartos (2 suítes, sendo 1 reversível), 2 vagas. Total 105 unidades. Tabela 04/2026 lista 22 unidades LIVRES (15 Botticelli + 7 Leonardo) — assumindo que tabela só lista LIVRES, estimativa 79% vendido (margem: pode haver reservadas/contratadas não mostradas). Tickets R$ 1.038k (BO 101, menor) a R$ 1.565k (LE 1401, maior). Conclusão obra AGO/2027. Construtora Monteplan. Versão tabela 1.04.", "tabela_local_parcial", None, None),

    ("Monteplan","Edifício Sanpaolo",
     "Rua Boa Esperança, 125, Cohama, São Luís - MA","Cohama",
     "Vertical",None,
     None,"12/2022","12/2025", 54.0,59.0,None, "2D; 3D",
     610000,610000, None,None, None,
     "tabela_local","tabela_local","site_oficial",
     "https://monteplanengenharia.com.br/empreendimentos/edificio-sanpaolo/","29/04/2026",
     "Tipologia detalhada: 2 plantas. **Colunas 1,2,7,8** com 59m² — 3 quartos (1 suíte), 2 vagas. **Colunas 3,4,5,6** com 54m² — 2 quartos (2 suítes, sendo 1 reversível), 1 vaga. Tabela 04/2026 lista APENAS 1 unidade LIVRE (apto 204-205, R$ 610.000 — par de unidades unidas, situação L-L). Estimativa ≥98% vendido. Confirma 'todas as unidades vendidas' (Facebook out/2025) — restou só 1 unid. dupla. Endereço completo: Rua Boa Esperança, 125, Cohama (ao lado da Igreja Batista). Conclusão obra DEZ/2025.", "N/A", None, None),

    ("Monteplan","Residencial Novo Anil",
     "Rua Estevão Braga, Cohab Anil IV, São Luís - MA","Cohab Anil IV",
     "Vertical","Médio",
     None,"01/2022","Pronto", 53.94,53.94,None, "—",
     324142,324142, None,None, None,
     "tabela","tabela","memorial",
     "https://monteplanengenharia.com.br","23/04/2026",
     "OBRA CONCLUÍDA (Monteplan, Cohab Anil IV). 32 unid listadas, todas ~R$ 324.142 (área 53,94 m²). R$/m² uniforme R$ 6.009. Padrão popular. SFH 60%. VGV residual listado R$ 10,4M. [reconstituído da v4.16 em 25/04/2026]", None, None, None),

    # ─── SÁ CAVALCANTE — Reserva Península (novo, None, None) ───
    ("Sá Cavalcante","Reserva Península",
     "Endereço não localizado, Península da Ponta D'Areia, São Luís - MA","Península",
     "Vertical","Luxo",
     None,"09/2025","—", 127.14,171.36,None, "4D",
     None,None, None,None,None,
     "site_oficial","N/A","site_oficial",
     "https://apto.vc/br/ma/sao-luis/ponta-dareia/reserva-peninsula","03/05/2026",
     "Tipologia detalhada (Apto.vc + Triunfo + Adhemar Carlos 03/05/2026): Apt 4 quartos (2-4 suítes), 127,14m² a 171,36m² priv. Mono-tipologia 4D (varia em # suítes). 2-3 vagas. 1.900m² lazer (piscinas com deck, beach tennis, soccer, festas, jogos, pet, gym, pilates/yoga, gourmet, sauna, spa, churrasqueira, kids, coworking). Lançamento estande 'Casa Sal' (out/2025 = 10/2025; release diz 09/2025 — ajustado para 09/2025). Bairro Península (alto padrão). **TOTAL ainda não confirmado** — aguarda book/release detalhado.", None, None, "site_oficial"),

    # ═══ HIALI ═════════════════════════════════════════════════════════
    ("Hiali","Le Noir",
     "Rua Osires, 05, Renascença II, São Luís - MA","Renascença II",
     "Vertical",None,
     25,"04/2025","12/2027", 49.74,62.62,None, "Studio; 1D; 2D",
     710000,870000, None,None, None,
     "tabela","tabela","memorial",
     "","23/04/2026",
     "Tipologia detalhada: Studios e 1-2 dorm (compactos premium). Parceria Hiali + Silveira Inc. Compactos premium: 49,74 / 58,91 / 62,62 m². 5 pavimentos tipo × 5 aptos/andar = ~25 unidades. Entrega DEZ/2027. Ticket R$ 710-870k. R$/m² méd R$ 13.810. Memorial R.09/25.101 registrado 17/04/2025 no 1º RI São Luís. Foco em mercado jovem / investidor. [reconstituído da v4.16 em 25/04/2026]", "tabela_local_parcial", None, None),

    # ═══ DOM INCORPORAÇÃO (própria, None, None) ═════════════════════════════════════
    ("DOM Incorporação","Dom Lucas",
     "Tv. Boa Esperança, 101 - Cantinho do Céu, São Luís - MA, 65074-030","Turu",
     "Horizontal",None,
     46,"02/2026","12/2028", 100.35,100.35,None, "3D",
     835894,850937, None,None, None,
     "tabela_local","tabela_local","interno",
     "","03/05/2026",
     "Tipologia detalhada: Casa 3 dorm (1 suíte) + 2 vagas. Condomínio horizontal (sobrados). 1 ÚNICA tipologia: casa 100,35 m² construída (área usada para R$/m²). Terreno varia 136,00-145,79 m² conforme posição. **TOTAL = 46 unidades** (UH 1 a 46, tabela ABR/2026 lista todas com status). Lazer: campo society, piscina, deck, salão, gourmet, petplay, playground. Status ABR/2026: **9 disponíveis** (UH 2, 4, 9, 17, 19, 21, 23, 24, 30) + 1 RESERVADA + 36 VENDIDAS = ~80% vendido. Entrega DEZ/2028 (atualizado da tabela 04/2026 — antes 01/2029). Ticket à vista R$ 835.894-850.937 → R$/m² construção R$ 8.330-8.481. Construção: DOM Incorporação + Agrasty Construções LTDA. Vendas: André Leite Imóveis. CORREÇÃO v5.2: Área máx era 145,78 (terreno) — corrigida para 100,35 (construída). Convenção PADRAO §1: Tipo=Horizontal usa área construída. Extraído via visão multimodal (PDFs imagem) v10.2. NB: origem total = `tabela_local_parcial` mesmo a tabela listando TODAS as 46 (com status), pois C_RAW guarda apenas DISPONÍVEIS — convenção §3.7 (mesma do Renaissance Conceito v9.2).", "tabela_local_parcial", None, "informado_manualmente"),

    ("DOM Incorporação","Dom José",
     "FQV9+JJ Jardim Eldorado, São Luís - MA","Turu",
     "Horizontal",None,
     22,"06/2024","07/2027", 154.64,154.64,None, "4D",
     1403358,1420196, None,None, None,
     "tabela_local","tabela_local","interno",
     "","03/05/2026",
     "Tipologia detalhada: Casa 4+ dorm, alto padrão. Condomínio horizontal alto padrão. 1 ÚNICA tipologia: casa 154,64 m² construída. Terreno varia 178,49-180,96 m² conforme posição. **TOTAL = 22 unidades** (UH 1 a 22, confirmado por implantação numerada do book + tabela). Status ABR/2026: **3 disponíveis** (UH 1, 8, 10) + 19 VENDIDAS = ~86% vendido. Tickets à vista R$ 1.403.358-1.420.196 → R$/m² construção R$ 9.075-9.184. Entrega JUL/2027 (atualizado da tabela 04/2026 — antes 06/2027). Construção: DOM Incorporação + LP Barros Const. e Serviços ME. Vendas: André Leite Imóveis. Bairro CEP/book: 'Jardim Eldorado'; bairro=Turu (região senso comum, §3.10 v10.5 — Rafael 03/05/2026). Lazer: piscina adulto+infantil, playground, espaço gourmet, churrasqueira. CORREÇÃO v5.2: Área máx era 180,98 (terreno) — corrigida para 154,64 (construída). Convenção PADRAO §1: Tipo=Horizontal usa área construída. Extraído via visão multimodal v10.2. NB: origem total = `tabela_local_parcial` mesmo a tabela listando TODAS as 22 (implantação 01-22), pois C_RAW guarda apenas DISPONÍVEIS — convenção §3.7.", "tabela_local_parcial", None, "informado_manualmente"),
]

# ═══════════════════════════════════════════════════════════════
# VALIDAÇÃO §0.1 do PADRAO v2.0: Mês lançamento DEVE ser MM/AAAA (com ⚠ T-36
# opcional) ou "—". Qualquer outro formato (AAAA puro, ~AAAA) é REJEITADO.
# ═══════════════════════════════════════════════════════════════
import re as _re_validate
_RGX_MES = _re_validate.compile(r'^(\d{2}/\d{4}( ⚠ T-36)?|—)$')
_problemas = []
for _row in E_RAW:
    _inc, _emp, _mes = _row[0], _row[1], _row[7]  # idx 7 = Mês lançamento (após remoção de Status na v6.0)
    if _mes is None:
        _problemas.append(f"  • {_inc} | {_emp}: Mês lançamento é None — usar \"—\" se faltam dados")
    elif not _RGX_MES.match(str(_mes)):
        _problemas.append(f"  • {_inc} | {_emp}: Mês lançamento {_mes!r} fora do padrão MM/AAAA")
if _problemas:
    raise ValueError(
        "❌ VALIDAÇÃO PADRAO v2.0 §1 (col 9 — Mês lançamento) FALHOU:\n" +
        "\n".join(_problemas) +
        "\n\nFormato exigido: MM/AAAA, ou MM/AAAA ⚠ T-36, ou \"—\" (sem dados)."
    )

# ═══════════════════════════════════════════════════════════════
# DATASET — 13 INCORPORADORAS (serão auto-calculadas a partir de E)
# ═══════════════════════════════════════════════════════════════
# Campos editáveis manualmente por incorporadora (os calculáveis ficam None):
# (incorporadora, site, instagram, posicionamento)
# v4.2: removidas colunas RI e Capital Aberto (irrelevantes para o mercado SLZ).
I_META = {
    "Mota Machado":     ("https://motamachado.com.br","@motamachado",
                         "Cearense em expansão no NE. Lançou Bossa em abr/2026 no Calhau (4 suítes, 191-195m², vista mar). Reserva São Marcos em obras. VGV 2025 R$350M."),
    "Berg Engenharia":  ("https://www.bergengenharia.com.br","@bergengenharia",
                         "28 anos em SL. Monte Meru (Península, 135m²) em comercialização. Mount Solaro em parceria com Gonçalves Empreend. (SPE, LI solicitada)."),
    "Alfa Engenharia":  ("https://www.alfaengenharia.com.br","@alfaengenhariama",
                         "Tech-forward (1º Housi do MA em LIV). Novo lançamento Giardino Residenza (Ponta do Farol). Marca no IG: @alfaengenhariama."),
    "Lua Nova":         ("https://construtoraluanova.com.br","@construtoraluanova",
                         "Desde 1985 (40+ anos). Villa Adagio + Lagoon Residence (Santo Amaro). Perfil médio-padrão. Mapeada via web."),
    "Delman":           ("https://www.delman.com.br","@delmanincorporadora",
                         "Referência de luxo em Ponta d'Areia; 7 empreend. mapeados (ticket R$530k-R$5,8M). 2026 traz forte movimento no Calhau: Landscape (lançamento) + The View (pré-lançamento ABR/2026, 36-101m², 17º andar premium ~R$18k/m²)."),
    "Treviso":          ("https://trevisoengenharia.com.br","@treviso.engenharia",
                         "Vernazza (120 un. Ponta d'Areia) + Villagio Treviso (loteamento). Ed. Biadene Oaice é sede."),
    "Ergus":            ("https://ergusengenharia.com.br","@ergusengenharia",
                         "25 anos em 2026. Multi-produto (Zion, Nexus, Lead Office, On Residence, Open Design). Proposta sustentável. Projeto Nexus 10.000m² em Renascença."),
    "Monteplan":        ("https://monteplanengenharia.com.br","@monteplanengenharia",
                         "Renaissance Conceito (Renascença II, alto padrão) + Sanpaolo (esgotado). Portfólio residencial e comercial."),
    "Franere":          ("https://franere.com.br","@franereoficial_",
                         "'Maior construtora do Maranhão' (self-proclamada). 40 anos. Série Gran Park / Varandas Grand Park."),
    "Canopus":          ("https://canopusconstrucoes.com.br","@canopusconstrucoes",
                         "Endereço: Av. Cel. Colares Moreira, 1, J. Renascença. 3 lançamentos out/2025: Village Reserva II, Prime Eldorado, Del Ville II. Atua também em Imperatriz."),
    "Niágara":          ("https://niagaraempreendimentos.com.br","@niagaraimoveis",
                         "ORO com ampla faixa 80-160m², parcelamento 48m pós-assinatura. Reserva dos Vinhais (Vinhais 2Q 48m²) + Reserva dos Buritis."),
    "MB Engenharia":    ("","@mbengenharia.br",
                         "Sócia da DOM Incorporação em Dom Antônio + Edifício Dom Ricardo (reclassificados como DOM em v4.18). Próprios: Prime Cohama (22 casas duplex 140m²) + Fernando de Noronha 2ª etapa."),
    "Sá Cavalcante":    ("https://www.sacavalcante.com.br","@sacavalcantema",
                         "Grupo desde 1974 (shoppings + incorporação). Ilha Parque entregue + NOVO Reserva Península (out/2025, Ponta d'Areia)."),
    "Castelucci":       ("https://construtoracastelucci.com.br","@construtoracastelucci",
                         "10+ anos. CEO Paulo Castelucci (Mirante FM). Vila Coimbra (parceria Coimbra Alves, Araçagi alto padrão) + Villa di Carpi (Cohatrac) + Ana Vitória (Araçagy). Patrocinador Imob Summit."),
    "Hiali":            ("","@hialiconstrucoes",
                         "Parceria com Silveira Inc. Le Noir (Renascença II): compactos premium 49-62m², ticket R$ 710-870k. Memorial 04/2025. Posicionamento jovem/investidor. Mapeada via tabela arquivada."),
    "DOM Incorporação": ("https://domincorporacao.com.br","@domincorporacao",
                         "EU. Portfólio próprio: Dom Lucas (horizontal Cantinho do Céu, 100m²+terreno, R$ 835-851k), Dom José (horizontal Eldorado alto padrão, casa 154m², R$ 1,4M). Em parceria com MB Engenharia: Dom Antônio (Turú) + Edifício Dom Ricardo (Jd. Renascença). Tracked aqui para benchmarking interno."),
}

# ═══════════════════════════════════════════════════════════════
# C_RAW — Composição por tipologia (v8.0+)
# ═══════════════════════════════════════════════════════════════
# 1 entry por (empreendimento, tipologia) extraída de tabelas locais.
# Schema (10 colunas): Incorporadora, Empreendimento, Tipologia, Nº Unidades,
# Área min, Área max, Ticket min, Ticket max, R$/m² médio, Origem.
#
# Lote 1 (entregue v8.0): 15 linhas / 322 unidades / 8 empreend. processados.
# Heurística tipologia × área (SLZ-padrão): <40 Studio, 40-55 1D, 55-75 2D,
# 75-95 3D, >95 4D. Em casos especiais (mono-tipologia declarada, áreas em
# fronteira), classificação manual prevalece.
C_RAW = [
    # The View (Delman) — 93 unid: tabela 28/04/2026 v3
    ("Delman", "The View", "Studio", 28, 36.05, 36.45, 539969, 640469, 16763, "tabela_local"),
    ("Delman", "The View", "1D", 26, 42.45, 54.06, 630255, 959721, 16006, "tabela_local"),
    ("Delman", "The View", "2D", 15, 68.91, 74.92, 1062210, 1355300, 16577, "tabela_local"),
    ("Delman", "The View", "3D", 24, 80.72, 85.87, 1354011, 1504011, 16637, "tabela_local"),
    # Landscape (Delman) — 51 unid: tabela 04/2026
    ("Delman", "Landscape", "3D", 17, 88.07, 88.07, 1206282, 1400282, 14755, "tabela_local"),
    ("Delman", "Landscape", "4D", 34, 103.60, 143.64, 1428516, 2328766, 14716, "tabela_local"),
    # Studio Design 7 Península (Delman) — 31 unid: tabela 04/2026
    ("Delman", "Studio Design 7 Península", "1D", 13, 42.40, 43.50, 710160, 740727, 16813, "tabela_local"),
    ("Delman", "Studio Design 7 Península", "2D", 17, 61.08, 65.80, 1004013, 1146600, 16264, "tabela_local"),
    ("Delman", "Studio Design 7 Península", "3D", 1, 88.78, 88.78, 1553650, 1553650, 17500, "tabela_local"),
    # Wave Residence (Delman) — 5 unid: penthouses 293m² Ponta do Farol
    ("Delman", "Wave Residence", "4D", 5, 293.69, 293.69, 5581808, 5828289, 19360, "tabela_local"),
    # Edifício Bossa (Mota Machado) — 36 unid: tabela 04/2026
    ("Mota Machado", "Edifício Bossa", "4D", 36, 191.02, 196.04, 2850507, 3708342, 16663, "tabela_local"),
    # Altos do São Francisco (Treviso) — 26 unid: imóvel pronto, ago/2025
    ("Treviso", "Altos do São Francisco", "2D", 26, 57.93, 67.15, 495809, 761677, 10024, "tabela_local"),
    # Renaissance Conceito (Monteplan) — 44 unid: 2 torres Botticelli 82m² (3D) + Leonardo 110m² (4D)
    ("Monteplan", "Renaissance Conceito", "3D", 15, 82.00, 82.00, 1038621, 1177759, 13686, "tabela_local"),  # v9.2: corrigido 30→15 (parser SFH+FDC duplicava)
    ("Monteplan", "Renaissance Conceito", "4D", 7, 110.00, 110.00, 1359410, 1565192, 13168, "tabela_local"),  # v9.2: corrigido 14→7 (parser duplicava)
    # Vila Coimbra (Castelucci) — 36 casas horizontais Araçagy, área construída uniforme
    ("Castelucci", "Vila Coimbra", "4D", 36, 124.63, 124.63, 1019834, 1081967, 8367, "tabela_local"),

    # ─── Lote 2 (v8.1) ──────────────────────────────────────────
    # Vernazza Torre Norte (Treviso) — 37 unid 130-130,5m² Renascença
    ("Treviso", "Vernazza Torre Norte", "4D", 37, 130.00, 130.49, 1784821, 2191902, 15219, "tabela_local"),
    # Vernazza Torre Sul (Treviso) — 26 unid 88-90m²
    ("Treviso", "Vernazza Torre Sul", "3D", 26, 87.98, 90.10, 1277584, 1586363, 15599, "tabela_local"),
    # Quartier 22 (Delman) — 1 unid Ponta d'Areia, 165m² (apto 601)
    ("Delman", "Quartier 22", "3D", 1, 165.00, 165.00, 3000000, 3000000, 18182, "tabela_local"),
    # Sky Residence (Delman) — 1 unid Ponta d'Areia, 247m² (apto 200)
    ("Delman", "Sky Residence", "4D", 1, 246.69, 246.69, 4700000, 4700000, 19052, "tabela_local"),
    # Azimuth (Delman) — 1 unid Ponta d'Areia, 197m² (apto 901)
    ("Delman", "Azimuth", "3D", 1, 196.62, 196.62, 3600000, 3600000, 18309, "tabela_local"),
    # Al Mare Tirreno (Mota Machado) — 1 unid 215m² Av. dos Holandeses Q9 L9
    ("Mota Machado", "Al Mare Tirreno", "4D", 1, 215.00, 215.00, 3025856, 3025856, 14074, "tabela_local"),
    # Entre Rios (Mota Machado) — 30 unid (Torre Douro + Tejo) 146,82m² Renascença
    ("Mota Machado", "Entre Rios", "3D", 30, 146.82, 146.82, 1732638, 2719860, 14679, "tabela_local"),
    # Reserva São Marcos (Mota Machado) — 2 plantas, áreas do BOOK
    # Planta 1 (2D) 67-69m² | Planta 2 (3D) 102-104m². Tickets agregados (~6 grupos).
    ("Mota Machado", "Reserva São Marcos", "2D", 3, 67.48, 68.75, 977382, 1316965, 17535, "tabela_local"),
    ("Mota Machado", "Reserva São Marcos", "4D", 3, 102.25, 104.05, 1467380, 1869461, 16590, "tabela_local"),
    # Le Noir (Hiali) — 4 unid Renascença II, áreas 49-62m² compactos
    ("Hiali", "Le Noir", "1D", 1, 49.74, 49.74, 710000, 710000, 14274, "tabela_local"),
    ("Hiali", "Le Noir", "2D", 3, 58.91, 62.62, 800000, 870000, 13682, "tabela_local"),
    # ORO Ponta d'Areia (Niágara) — 96 unid: 88 padrão 80m² + 8 cobertura duplex 160m²
    ("Niágara", "ORO Ponta d'Areia", "3D", 88, 80.32, 80.32, 1003326, 1537396, 15952, "tabela_local"),
    ("Niágara", "ORO Ponta d'Areia", "4D", 8, 160.65, 160.65, 2275000, 2575000, 15095, "tabela_local"),
    # Edifício Sanpaolo (Monteplan) — 1 unid dupla restante (apto 204-205, R$610k 108m²)
    # Tratada como 2 unidades 54m² R$305k cada. Posicionamento Cohama, ~99% vendido.
    ("Monteplan", "Edifício Sanpaolo", "1D", 2, 54.00, 54.00, 305000, 305000, 5648, "tabela_local"),

    # ─── Lote 3 (v8.2) — Zion via visão multimodal ───────────────
    # Zion Ponta d'Areia (Ergus) — 10 unid disponíveis 04/2026, todas 148,55m² 4D.
    # Tabela em PDF imagem extraída via pdftoppm + Claude visão. R$/m² médio 15.500.
    ("Ergus", "Zion Ponta d'Areia", "4D", 10, 148.55, 148.55, 2170378, 2556972, 15500, "tabela_local_imagem"),

    # ─── Lote 4 (v10.1) — 4 empreend. com tabela texto ─────────────
    # Monte Meru (Berg Engenharia) — 2 unid 3D, ~135m²
    ("Berg Engenharia", "Monte Meru", "3D", 2, 135.32, 135.83, 1932400, 1944500, 14298, "tabela_local"),
    # Residencial Novo Anil (Monteplan) — 30 unid 2D 53,94m² uniforme (8 blocos A1-B4)
    ("Monteplan", "Residencial Novo Anil", "2D", 30, 53.94, 53.94, 324142, 324143, 6009, "tabela_local"),
    # Giardino Fiore (Alfa) — 6 unid 3D, 110-128m², 3 colunas
    ("Alfa Engenharia", "Giardino Residenza Torre Fiore", "3D", 6, 110.77, 128.37, 1838492, 2032938, 15373, "tabela_local"),
    # Giardino Luce (Alfa) — 5 unid 3D, 93-101m², 4 colunas
    ("Alfa Engenharia", "Giardino Residenza Torre Luce", "3D", 5, 93.62, 101.31, 1442168, 1595303, 15299, "tabela_local"),

    # ─── Lote 5 (v10.2) — DOM Incorporação via visão multimodal ─────
    # Dom Lucas (DOM) — 9 unid disponíveis 04/2026, todas 100,35 m² casa (3D)
    # Tabela ABR/2026 em PDF imagem — extraída via pdftoppm + visão Claude.
    # 46 unid totais (UH 1-46), 9 disp + 1 reserv + 36 vend → ~80% vendido.
    ("DOM Incorporação", "Dom Lucas", "3D", 9, 100.35, 100.35, 835894, 850937, 8406, "tabela_local_imagem"),
    # Dom José (DOM) — 3 unid disponíveis 04/2026, todas 154,64 m² casa (4D)
    # 22 unid totais (UH 1-22 confirmado por implantação numerada), 3 disp + 19 vend → ~86% vendido.
    ("DOM Incorporação", "Dom José", "4D", 3, 154.64, 154.64, 1403358, 1420196, 9130, "tabela_local_imagem"),

    # ─── DOM Ricardo (Lote 6 v10.7) — book DOM 12/2023 + total declarado Rafael 03/05/2026 ───
    # 30 unid total = 10 andares × 3 col por andar. Col 1 (85,75m² 3D) + Col 2 (84,96m² 3D) + Col 3 (71,92m² 2D)
    # Book diz "100% VENDIDO" → disp = 0 em ambas tipologias. Para origem `book`, Σ disp não é base de Total tip.
    # NOTA: c[3] aqui é "disp", mas pra origem `book` o cálculo de Total tip usa lookup direto (ver compute_total_per_tipologia)
    ("DOM Incorporação", "Edifício Dom Ricardo", "2D", 0, 71.92, 71.92, 690860, 690860, 9606, "book"),
    ("DOM Incorporação", "Edifício Dom Ricardo", "3D", 0, 84.96, 85.75, 1194374, 1194374, 13989, "book"),

    # ─── Dom Antônio será aplicado pelo nível 5.1 mono automaticamente (12u 3D 136,2m²) ───

    # ─── Mount Solaro (Lote 7 v10.8) — book Berg+Gonçalves via site oficial 03/05/2026 ───
    # 50 unid total = 20 lofts 68m² (2D) + 10 apt 72m² (2D) + 20 apt 104m² (3D)
    # Pré-lançamento (06/2025 T-36) → todos disponíveis (estoque 100%)
    ("Berg Engenharia", "Mount Solaro", "2D", 30, 68.00, 72.00, 907200, 1100000, 13750, "site_oficial"),
    ("Berg Engenharia", "Mount Solaro", "3D", 20, 104.00, 104.00, 1500000, 1700000, 15384, "site_oficial"),
]

# ═══════════════════════════════════════════════════════════════
# v10.6 — VIRADA §3.7 v2 (PADRAO v6.2)
#   Funções: consolidação multi-torre, estimativa nível 5, total por tipologia
# ═══════════════════════════════════════════════════════════════

def consolidate_multi_torre(E_RAW, C_RAW):
    """§3.7.D regra A: torres da mesma marca/lançamento viram entry única.

    Consolida os 2 pares confirmados em audit (v10.6):
      - Treviso: Vernazza Norte (120) + Sul (60) → "Vernazza Residenza" (180)
      - Alfa: Giardino Fiore (45) + Luce (60) → "Giardino Residenza" (105)
    """
    pairs = [
        # (incorporadora, [nomes_torres], nome_consolidado)
        ("Treviso",
         ["Vernazza Torre Norte", "Vernazza Torre Sul"],
         "Vernazza Residenza"),
        ("Alfa Engenharia",
         ["Giardino Residenza Torre Fiore", "Giardino Residenza Torre Luce"],
         "Giardino Residenza"),
    ]
    log = []

    for inc, torres, novo in pairs:
        idx_torres = [i for i, e in enumerate(E_RAW) if e[0] == inc and e[1] in torres]
        if len(idx_torres) < 2:
            log.append(f"  [skip] {inc}/{novo}: encontrou apenas {len(idx_torres)} torre(s)")
            continue
        torres_entries = [E_RAW[i] for i in idx_torres]

        # Sanity: bairro/tipo/segmento iguais
        bairros = set(e[3] for e in torres_entries)
        tipos = set(e[4] for e in torres_entries)
        if len(bairros) > 1 or len(tipos) > 1:
            log.append(f"  [skip] {inc}/{novo}: bairro/tipo divergente entre torres ({bairros}/{tipos})")
            continue

        # Construir entry consolidada como lista mutável (E_RAW tem 27 col)
        base = list(torres_entries[0])
        base[1] = novo  # nome consolidado

        # Total = soma
        totais = [e[6] for e in torres_entries if e[6]]
        base[6] = sum(totais) if totais else None

        # Mês lançamento = mais antigo (formato MM/YYYY)
        def lanc_key(s):
            if not s or s == '—': return (9999, 12)
            import re as _r
            m = _r.match(r'^(\d{1,2})/(\d{4})$', str(s))
            if m: return (int(m.group(2)), int(m.group(1)))
            return (9999, 12)
        lanc = min(torres_entries, key=lambda e: lanc_key(e[7]))[7]
        base[7] = lanc

        # Mês entrega = mais tardio
        def ent_key(s):
            if not s or s in ('—', 'Pronto'): return (0, 0)
            import re as _r
            m = _r.match(r'^(\d{1,2})/(\d{4})$', str(s))
            if m: return (int(m.group(2)), int(m.group(1)))
            return (0, 0)
        ent = max(torres_entries, key=lambda e: ent_key(e[8]))[8]
        base[8] = ent

        # Áreas/tickets: min/max combinado
        def safe_min(vals): vs = [v for v in vals if v is not None]; return min(vs) if vs else None
        def safe_max(vals): vs = [v for v in vals if v is not None]; return max(vs) if vs else None
        base[9] = safe_min([e[9] for e in torres_entries])   # area_min
        base[10] = safe_max([e[10] for e in torres_entries]) # area_max
        base[13] = safe_min([e[13] for e in torres_entries]) # ticket_min
        base[14] = safe_max([e[14] for e in torres_entries]) # ticket_max
        # R$/m² recalculado depois pelo script
        base[15] = None
        base[16] = None  # VGV recalculado

        # Tipologia: união
        tips = set()
        for e in torres_entries:
            if e[12]:
                for t in str(e[12]).split(';'):
                    t = t.strip()
                    if t: tips.add(t)
        TIP_ORDER = ['Studio', '1D', '2D', '3D', '4D', 'Lote']
        tips_sorted = sorted(tips, key=lambda t: TIP_ORDER.index(t) if t in TIP_ORDER else 99)
        base[12] = '; '.join(tips_sorted) if tips_sorted else (torres_entries[0][12] or None)

        # Observações: registrar consolidação
        obs_old = base[23] or ''
        base[23] = (f"[v10.6 consolidado §3.7.D-A — torres originais: {' + '.join(torres)} "
                    f"(totais {' + '.join(str(e[6]) for e in torres_entries)} = {base[6]})] " + obs_old)

        # Origem total: revisar — se ambas torres tinham origem completa, mantém; senão, marca como inferido
        origens_total = set(e[24] for e in torres_entries)
        if 'tabela_local_completa' in origens_total and len(origens_total) == 1:
            base[24] = 'tabela_local_completa'
        elif 'memorial' in origens_total:
            base[24] = 'memorial'
        else:
            base[24] = list(origens_total)[0] if origens_total else None

        # Substituir entries de torres pela consolidada
        # Remove em ordem reversa pra não bagunçar índices
        for i in sorted(idx_torres, reverse=True):
            del E_RAW[i]
        E_RAW.append(tuple(base))

        # Consolidar C_RAW: trocar nome da torre pelo consolidado, agregar mesmas tipologias
        crows_torres = [c for c in C_RAW if c[0] == inc and c[1] in torres]
        crows_outras = [c for c in C_RAW if not (c[0] == inc and c[1] in torres)]
        agg_by_tip = {}
        for c in crows_torres:
            tip = c[2]
            if tip not in agg_by_tip:
                agg_by_tip[tip] = list(c)
                agg_by_tip[tip][1] = novo  # rename
                continue
            cur = agg_by_tip[tip]
            cur[3] = (cur[3] or 0) + (c[3] or 0)  # disp
            cur[4] = safe_min([cur[4], c[4]])     # area_min
            cur[5] = safe_max([cur[5], c[5]])     # area_max
            cur[6] = safe_min([cur[6], c[6]])     # ticket_min
            cur[7] = safe_max([cur[7], c[7]])     # ticket_max
            # R$/m² médio: recalcular se possível
            if cur[6] and cur[7] and cur[4] and cur[5]:
                cur[8] = round(((cur[6] + cur[7]) / 2) / ((cur[4] + cur[5]) / 2))
        C_RAW.clear()
        C_RAW.extend(crows_outras)
        for tip in sorted(agg_by_tip.keys()):
            C_RAW.append(tuple(agg_by_tip[tip]))

        log.append(f"  [ok] {inc}/{novo}: {len(idx_torres)} torres → 1 entry, total={base[6]}, "
                   f"{len(crows_torres)} C_RAW → {len(agg_by_tip)} tipologias agregadas")

    if log:
        print("§3.7.D — Consolidação multi-torre (v10.6):")
        for l in log: print(l)
    return E_RAW, C_RAW


def compute_total_per_tipologia(E_RAW, C_RAW):
    """§3.7 v6.2: calcula Total por tipologia em runtime, retorna dict.

    Regras (já documentadas em PADRAO §3.7):
      - Mono-tipologia em E_RAW: Total tipologia = E_RAW.Total
      - Multi-tip + origem `tabela_local_completa`: Σ disp já = Total → Total tip = disp
      - Multi-tip + origem `tabela_local_parcial`: pro-rata pelo % do disp_emp,
        origem ganha sufixo `_pro_rata`
      - Origens estimativa_*: Total tip = disp (já é o total por construção)
      - Total empreend. ausente → Total tip = None
    """
    # Lookup empreend. → entry
    emp_lookup = {(e[0], e[1]): e for e in E_RAW}
    # Soma disp por empreend.
    disp_by_emp = {}
    for c in C_RAW:
        key = (c[0], c[1])
        disp_by_emp[key] = disp_by_emp.get(key, 0) + (c[3] or 0)

    # Conta n entries C_RAW por empreend. (chave: mono-em-C_RAW vs multi-em-C_RAW)
    n_tips_c_raw = {}
    for c in C_RAW:
        n_tips_c_raw[(c[0], c[1])] = n_tips_c_raw.get((c[0], c[1]), 0) + 1

    # Override manual de Total tipologia para casos especiais onde o book declara totais
    # mas C_RAW.disp armazena 0 (100% vendido). Chave: (inc, emp, tip) → total_tip.
    # NOVO v10.7: usado quando origem=`book` e disp=0 (book diz 100% vendido) — declara totais explícitos.
    BOOK_TOTAL_OVERRIDE = {
        ('DOM Incorporação', 'Edifício Dom Ricardo', '2D'): 10,
        ('DOM Incorporação', 'Edifício Dom Ricardo', '3D'): 20,
    }

    result = {}  # (inc, emp, tip) → (total_tip, origem_revisada)
    # Pré-cálculo p/ pro-rata: se ∆ entre soma e total é não-zero, ajustar último entry pra fechar
    pending_pro_rata = {}  # (inc, emp) → list of (key, raw_value)

    for c in C_RAW:
        inc, emp, tip, disp = c[0], c[1], c[2], c[3]
        origem = c[9]
        emp_entry = emp_lookup.get((inc, emp))
        if not emp_entry:
            result[(inc, emp, tip)] = (None, origem)
            continue
        total_emp = emp_entry[6]

        # Override book_total_explicito (v10.7): origem `book` com Total tipologia declarado manualmente
        if (inc, emp, tip) in BOOK_TOTAL_OVERRIDE:
            result[(inc, emp, tip)] = (BOOK_TOTAL_OVERRIDE[(inc, emp, tip)], origem)
            continue

        # Origens estimativa_*: total tip = disp (estimativa já é total por construção)
        if origem and origem.startswith('estimativa_distribuição'):
            result[(inc, emp, tip)] = (disp, origem)
            continue

        if not total_emp:
            result[(inc, emp, tip)] = (None, origem)
            continue

        # Mono-tipologia em C_RAW (única entry pro empreend.) → toda composição é dessa tipologia
        if n_tips_c_raw.get((inc, emp), 0) == 1:
            result[(inc, emp, tip)] = (total_emp, origem)
            continue

        # Multi-tip + origem completa: Σ disp já = Total (por construção)
        if origem == 'tabela_local_completa':
            result[(inc, emp, tip)] = (disp, origem)
            continue

        # Multi-tip + origem parcial (incl. tabela_local, tabela_local_imagem, tabela_local_parcial):
        # pro-rata pelo % do disp_emp. Ajuste de off-by-one feito num passo final.
        soma_disp = disp_by_emp.get((inc, emp), 0)
        if soma_disp > 0:
            total_tip = round((disp or 0) * total_emp / soma_disp)
            ori_marcada = (origem + '_pro_rata') if (origem and not origem.endswith('_pro_rata')) else (origem or 'pro_rata')
            result[(inc, emp, tip)] = (total_tip, ori_marcada)
            pending_pro_rata.setdefault((inc, emp), []).append((inc, emp, tip))
        else:
            result[(inc, emp, tip)] = (None, origem)

    # Ajuste de arredondamento: pra cada empreend. com pro-rata, força Σ = Total
    for (inc, emp), keys in pending_pro_rata.items():
        total_emp = emp_lookup[(inc, emp)][6]
        if not total_emp: continue
        soma = sum(result[k][0] for k in keys if result[k][0] is not None)
        diff = total_emp - soma
        if diff != 0 and keys:
            # Ajusta no entry com maior Total tip
            keys_sorted = sorted(keys, key=lambda k: -(result[k][0] or 0))
            k_maior = keys_sorted[0]
            tt, ori = result[k_maior]
            result[k_maior] = (tt + diff, ori)

    return result


def apply_estimativa_distribuicao(E_RAW, C_RAW):
    """§3.7.A.1 nível 5: aplica sub-regras 5.1-5.4 nos empreend. com Total mas sem C_RAW."""
    from statistics import median
    from collections import defaultdict as _dd

    # Medianas de área por tipologia (calculadas runtime de C_RAW existente, fontes fortes apenas)
    areas_por_tip = _dd(list)
    for c in C_RAW:
        origem = c[9] or ''
        if origem.startswith('estimativa_distribuição'):
            continue
        a_min, a_max = c[4], c[5]
        if a_min and a_max:
            areas_por_tip[c[2]].append((a_min + a_max) / 2)
    median_area = {t: median(vs) for t, vs in areas_por_tip.items() if vs}

    emps_com_comp = set((c[0], c[1]) for c in C_RAW)
    new_entries = []
    log = []
    bloqueados = []

    TIP_ORDER = ['Studio', '1D', '2D', '3D', '4D', 'Lote']

    for entry in E_RAW:
        inc, emp = entry[0], entry[1]
        if (inc, emp) in emps_com_comp: continue
        total = entry[6]
        if not total:
            bloqueados.append((inc, emp, entry[12] or '—'))
            continue
        tip_decl = (entry[12] or '—').strip()
        a_min = entry[9]; a_max = entry[10]
        tip_list = [t.strip() for t in str(tip_decl).split(';') if t.strip() and t.strip() != '—']

        if not tip_list:
            # 5.4: sem tipologia
            new_entries.append((inc, emp, '—', total, None, None, None, None, None,
                                'estimativa_distribuição_sem_tipologia'))
            log.append(f"  [5.4] {inc} | {emp}: 1 entry '—' Total={total}")
        elif len(tip_list) == 1:
            # 5.1: mono-tipologia
            t = tip_list[0]
            if a_min and a_max:
                am, ax = a_min, a_max
            else:
                med = median_area.get(t)
                am = ax = med
            new_entries.append((inc, emp, t, total, am, ax, None, None, None,
                                'estimativa_distribuição_mono'))
            log.append(f"  [5.1] {inc} | {emp}: 1 entry {t} Total={total} área={am}-{ax}")
        else:
            # 5.2 ou 5.3
            tip_sorted = sorted(tip_list, key=lambda t: TIP_ORDER.index(t) if t in TIP_ORDER else 99)
            n = len(tip_sorted)
            base = total // n
            sobra = total % n
            unids = [base + (1 if i < sobra else 0) for i in range(n)]

            tem_area = a_min is not None and a_max is not None
            if tem_area and n >= 2:
                # 5.2 com área: menor → tip menor, maior → tip maior, intermediárias = mediana
                area_per_tip = {}
                for i, t in enumerate(tip_sorted):
                    if i == 0: area_per_tip[t] = a_min
                    elif i == n - 1: area_per_tip[t] = a_max
                    else: area_per_tip[t] = median_area.get(t, (a_min + a_max) / 2)
                origem = 'estimativa_distribuição_multi_com_area'
            else:
                # 5.3 sem área
                area_per_tip = {t: median_area.get(t) for t in tip_sorted}
                origem = 'estimativa_distribuição_multi_sem_area'

            for t, u in zip(tip_sorted, unids):
                a = area_per_tip.get(t)
                new_entries.append((inc, emp, t, u, a, a, None, None, None, origem))
            log.append(f"  [{'5.2' if tem_area else '5.3'}] {inc} | {emp}: {n} tipologias "
                       f"({'+'.join(f'{u}{t}' for u,t in zip(unids,tip_sorted))}) origem={origem}")

    C_RAW.extend(new_entries)
    if log:
        print(f"§3.7.A.1 — Estimativa nível 5 aplicada ({len(new_entries)} entries em "
              f"{len(set((e[0],e[1]) for e in new_entries))} empreend.):")
        for l in log: print(l)
    if bloqueados:
        print(f"§3.7 — BLOQUEADOS sem Total ({len(bloqueados)} empreend., entram em pendencias_TOTAL.md):")
        for inc, emp, tip in bloqueados:
            print(f"  - {inc} | {emp} | tip={tip}")
    return new_entries, bloqueados


# Aplicar (ordem importa: consolidação primeiro, depois estimativa)
E_RAW = list(E_RAW)  # tornar mutável
C_RAW = list(C_RAW)
E_RAW, C_RAW = consolidate_multi_torre(E_RAW, C_RAW)
EST_NEW, BLOQUEADOS = apply_estimativa_distribuicao(E_RAW, C_RAW)
TOTAL_TIP_DICT = compute_total_per_tipologia(E_RAW, C_RAW)

# ═══════════════════════════════════════════════════════════════
# CÁLCULOS AUTOMÁTICOS (ver §3 do PADRAO.md)
# ═══════════════════════════════════════════════════════════════
def calc_preco_m2(tmin,tmax,amin,amax):
    """§3.1 — padrão: ticket_médio / área_média"""
    if None in (tmin,tmax,amin,amax): return None
    tm = (tmin+tmax)/2
    am = (amin+amax)/2
    if am==0: return None
    return tm/am

def calc_vgv(tmin,tmax,unidades):
    """§3.2"""
    if None in (tmin,tmax,unidades): return None
    return ((tmin+tmax)/2) * unidades

def calc_area_media(amin,amax):
    if None in (amin,amax): return None
    return (amin+amax)/2

def extract_year(mes_str):
    """Extrai ano de string tipo '04/2025', '~2024', '~2025 ⚠ T-36', '2024 ♦'"""
    if not mes_str or mes_str=="—": return None
    import re
    m=re.search(r'(\d{4})',mes_str)
    return int(m.group(1)) if m else None

def parse_launch_date(mes_str):
    """Converte string de mês de lançamento para tupla (ano, mês) para ordenação.
    PADRAO v2.0 §1: formato esperado é SEMPRE MM/AAAA (com flag opcional ⚠ T-36).
    Retorna (0,0) para vazios E para formatos inválidos (AAAA puro, ~AAAA) —
    todos vão para o fim em ordem decrescente, sinalizando que falta dado preciso."""
    if not mes_str or mes_str=="—":
        return (0, 0)
    import re
    # Aceita SOMENTE MM/AAAA (com ou sem ⚠ T-36)
    m = re.match(r'^(\d{1,2})/(\d{4})( ⚠ T-36)?$', mes_str)
    if m:
        return (int(m.group(2)), int(m.group(1)))
    # Formato inválido — vai pro fim (faltam dados)
    return (0, 0)

# Processa cada empreendimento: preenche campos calculados
# ═══════════════════════════════════════════════════════════════
# v9.4 §3.8 — CÁLCULO AUTOMÁTICO DE % VENDIDO (estoque)
# ═══════════════════════════════════════════════════════════════
# Regra (PADRAO §3.8):
#   estoque_calc = Σ disponíveis em C_RAW / total_unidades
#   (E_RAW armazena estoque; xlsx mostra 1-estoque como "% Vendido")
# Casos:
#   1) origem total = tabela_local_completa AND soma C_RAW = total → estoque=1.0 (nada vendido). Origem: tabela_local_completa_zero
#   2) total + soma C_RAW conhecidos AND não-Niágara → calcular. Origem: calculado_automatico
#   3) Niágara → não-determinável (formato agrupa, sem inferência segura)
#   4) Manual existente → manter, validar contra calc (WARN se >5% diff)
#   5) Sem dado base → estoque=None, origem=N/A (vira lista de busca)

warnings_38 = []
E_RAW_v94 = []
for entry in E_RAW:
    entry_l = list(entry)  # tupla → lista mutável
    inc, emp = entry_l[0], entry_l[1]
    total = entry_l[6]
    estoque_atual = entry_l[17]
    origem_total = entry_l[24]
    soma_comp = sum(c[3] for c in C_RAW if c[0] == inc and c[1] == emp)

    if estoque_atual is not None:
        # Já preenchido manualmente — manter, marcar origem
        entry_l[25] = 'informado_manualmente'
        # Validação: se temos dados pra calcular, comparar
        if total and total > 0 and soma_comp > 0:
            estoque_calc = soma_comp / total
            diff = abs(estoque_atual - estoque_calc)
            if diff > 0.05:  # 5%
                warnings_38.append(
                    f"  WARN {inc} | {emp}: estoque manual {estoque_atual:.2f} ({(1-estoque_atual)*100:.0f}% vendido) vs calc {estoque_calc:.2f} ({(1-estoque_calc)*100:.0f}% vendido)"
                )
    elif inc == 'Niágara':
        entry_l[25] = 'nao_determinavel'  # tabela agrupa, não permite inferir
    elif origem_total == 'tabela_local_completa' and total and soma_comp == total:
        # Pré-lançamento: tudo disponível, 0 vendidas
        entry_l[17] = 1.0
        entry_l[25] = 'tabela_local_completa_zero'
    elif total and total > 0 and soma_comp > 0:
        # Caso comum: calcular
        entry_l[17] = soma_comp / total
        entry_l[25] = 'calculado_automatico'
    else:
        # Sem dado base
        entry_l[25] = 'N/A'

    E_RAW_v94.append(tuple(entry_l))

E_RAW = E_RAW_v94

# Resumo de % Vendido após cálculo
n_calc = sum(1 for e in E_RAW if len(e) > 25 and e[25] == 'calculado_automatico')
n_manual = sum(1 for e in E_RAW if len(e) > 25 and e[25] == 'informado_manualmente')
n_zero = sum(1 for e in E_RAW if len(e) > 25 and e[25] == 'tabela_local_completa_zero')
n_nao_det = sum(1 for e in E_RAW if len(e) > 25 and e[25] == 'nao_determinavel')
n_na = sum(1 for e in E_RAW if len(e) > 25 and e[25] == 'N/A')
print(f"§3.8 % Vendido: {n_calc} calculados | {n_manual} manuais | {n_zero} zero(completa) | {n_nao_det} não-det | {n_na} N/A (lista de busca)")

if warnings_38:
    print(f"\n⚠ §3.8 — {len(warnings_38)} divergência(s) manual vs calc > 5%:")
    for w in warnings_38:
        print(w)

E_PROCESSED = []
for row in E_RAW:
    row = list(row)
    # calcular área média (idx 11 — após remoção do Status na v6.0)
    if row[11] is None:
        row[11] = calc_area_media(row[9],row[10])
    # calcular preço médio R$/m² (idx 15)
    if row[15] is None:
        row[15] = calc_preco_m2(row[13],row[14],row[9],row[10])
    # calcular VGV (idx 16)
    if row[16] is None:
        row[16] = calc_vgv(row[13],row[14],row[6])
    # auto-classificar segmento se não definido (idx 5 — não shifta, é antes do Status removido)
    if row[5] is None and row[15] is not None:
        row[5] = classificar_segmento_por_m2(row[15])
    elif row[5] is None:
        row[5] = "—"
    # (v6.0) reclassificar_status removido junto com a coluna Status
    E_PROCESSED.append(tuple(row))

# ═══════════════════════════════════════════════════════════════
# FUNÇÕES DE ESTILO
# ═══════════════════════════════════════════════════════════════
def fill(c): return PatternFill("solid",fgColor=c)
def font(color=DOM_GRAY_DARK,size=10,bold=False,italic=False):
    return Font(name="Calibri",color=color,size=size,bold=bold,italic=italic)
def border_thin(c=DOM_GRAY_MID):
    s=Side(style="thin",color=c); return Border(left=s,right=s,top=s,bottom=s)
def center(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
def left():   return Alignment(horizontal="left",vertical="center",wrap_text=True)

def apply_header_row(ws,row,headers):
    ws.row_dimensions[row].height=42
    for i,h in enumerate(headers):
        c=ws.cell(row=row,column=1+i,value=h)
        c.font=font(DOM_WHITE,9,bold=True); c.fill=fill(DOM_GRAY_DARK); c.alignment=center()
        c.border=Border(bottom=Side(style="medium",color=DOM_GOLD),
                        left=Side(style="thin",color=DOM_GRAY_MID),
                        right=Side(style="thin",color=DOM_GRAY_MID))

def insert_logo(ws,path,cell="A1",height_px=55):
    if not os.path.exists(path): return
    img=XLImage(path); r=img.width/img.height
    img.height=height_px; img.width=int(height_px*r); ws.add_image(img,cell)

def set_column_widths(ws,widths):
    for i,w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1+i)].width=w

# ═══════════════════════════════════════════════════════════════
# CONSTRUÇÃO DA PLANILHA
# ═══════════════════════════════════════════════════════════════
wb = Workbook()

# ── ABA 1: EMPREENDIMENTOS ─────────────────────────────────────
ws1 = wb.active; ws1.title = "Empreendimentos"
N_COLS_E = 27  # v10.0: +1 (Origem Bairro)
ws1.row_dimensions[1].height=22; ws1.row_dimensions[2].height=28; ws1.row_dimensions[3].height=18
for r in (1,2,3):
    for c in range(1,N_COLS_E+1):
        ws1.cell(row=r,column=c).fill=fill(DOM_BLACK)
insert_logo(ws1,LOGO_TRANSP,"A1",55)
ws1.merge_cells(start_row=2,start_column=4,end_row=2,end_column=N_COLS_E)
t=ws1.cell(row=2,column=4,value="INTELIGÊNCIA DE MERCADO — Panorama de Empreendimentos")
t.font=font(DOM_WHITE,16,bold=True); t.alignment=Alignment(horizontal="left",vertical="center")
ws1.merge_cells(start_row=3,start_column=4,end_row=3,end_column=N_COLS_E)
s=ws1.cell(row=3,column=4,
    value=f"São Luís / MA  •  v{VERSION}  •  {DATE_STR}  •  Fase 1 (padrão congelado — ver _PADRAO_FASE_1/PADRAO.md)")
s.font=font(DOM_GOLD,10,italic=True); s.alignment=Alignment(horizontal="left",vertical="center")

headers_e = [
    "Incorporadora","Empreendimento","Endereço","Bairro","Origem Bairro","Tipo","Segmento",
    "Nº unid.","Origem total unid.","Mês lançamento","Mês entrega",
    "Área mín (m²)","Área máx (m²)","Tipologia média (m²)","Tipologia",
    "Ticket mín (R$)","Ticket máx (R$)","R$/m²","VGV (R$)","% Vendido","Origem % Vendido",
    "Orig. preços","Orig. estoque","Orig. lançamento",
    "Link fonte principal","Data verif.","Observações"
]
apply_header_row(ws1,5,headers_e)

formats_e = [None]*N_COLS_E
# v9.0: +1 col 8 (Origem total unid)
# v9.4: +1 col 20 (Origem % Vendido)
# v10.0: +1 col 5 (Origem Bairro) — todos índices ≥5 deslocaram +1
formats_e[11] = formats_e[12] = formats_e[13] = '0.0'  # Áreas
formats_e[15] = formats_e[16] = 'R$ #,##0'  # Tickets
formats_e[17] = 'R$ #,##0'  # R$/m²
formats_e[18] = 'R$ #,##0'  # VGV
formats_e[19] = '0.0%'  # % Vendido

# §7 PADRAO.md: ordenar por Mês Lançamento DESCENDENTE (mais recente 1º),
# depois Incorporadora A-Z, depois Empreendimento A-Z
empreend_sorted = sorted(
    E_PROCESSED,
    key=lambda r: (
        -parse_launch_date(r[7])[0],  # ano desc
        -parse_launch_date(r[7])[1],  # mês desc
        r[0],                          # incorporadora asc
        r[1]                           # empreendimento asc
    )
)

# v9.0: reordenar row_data: idx 24 (Origem total unid.) vai para pos 7 da xlsx
# Schema E_RAW (Python): [0..6, 7=Mês_lanc, ..., 23=Obs, 24=OrigTotal]
# Schema xlsx desejado:  [0..6, 7=OrigTotal, 8=Mês_lanc, ..., 24=Obs]
def reorder_for_xlsx(row):
    # v10.0: schema E_RAW (Python) = 27 cols
    # Layout visual da xlsx:
    #   [0..3, 26=OrigBairro, 4..6, 24=OrigTotal, 7..17, 25=OrigPctVend, 18..23]
    return (list(row[:4]) + [row[26]] + list(row[4:7]) + [row[24]] +
            list(row[7:18]) + [row[25]] + list(row[18:24]))

for i, row_data in enumerate(empreend_sorted):
    row_idx = 6+i
    ws1.row_dimensions[row_idx].height = 52
    row_fill = DOM_WHITE if row_idx%2==0 else DOM_GRAY_LIGHT
    row_values = reorder_for_xlsx(row_data)
    # v9.0: %Vendido agora é pos 18 (era 17) por causa do shift
    if isinstance(row_values[19], (int, float)):
        row_values[19] = 1 - row_values[19]  # v10.0: estoque agora em pos 19
    for j, v in enumerate(row_values):
        cel = ws1.cell(row=row_idx, column=1+j, value=v)
        cel.font = font(DOM_GRAY_DARK, 9)
        cel.fill = fill(row_fill)
        cel.alignment = left() if j in (2, 14, 26) else center()  # v10.0: Tipologia=14, Observações=26
        cel.border = border_thin()
        if formats_e[j]:
            cel.number_format = formats_e[j]
    # Sem coloração condicional na coluna % Vendido (Rafael pediu p/ remover)
    # Destaque da incorporadora
    ws1.cell(row=row_idx, column=1).font = font(DOM_GRAY_DARK, 9, bold=True)

total_row_e = 6+len(empreend_sorted)

widths_e = [15,22,30,14, 16, 11, 11, 7,18,14,11, 10,10,11,20, 13,13,11,14,10, 16,
            14,14,18, 28,10,50]  # v10.0: +16 na pos 4 (Origem Bairro)
set_column_widths(ws1, widths_e)
ws1.freeze_panes = "C6"
ws1.auto_filter.ref = f"A5:{get_column_letter(N_COLS_E)}{total_row_e-1}"

# Legenda
ws1.merge_cells(start_row=total_row_e, start_column=1, end_row=total_row_e, end_column=N_COLS_E)
leg = ws1.cell(row=total_row_e, column=1,
    value="ESTOQUE — 🟢 ≤15% (últimas unidades) | 🟡 15-40% (em absorção) | 🔴 >40% (estoque amplo).    "
          "STATUS = comercial (não físico).    "
          "SEGMENTO = classificado pelo R$/m² calculado (ver §4.2 do PADRAO.md).    "
          "⚠ T-36 = lançamento estimado por Entrega−42 meses. Substituir assim que tiver fonte.")
leg.font = font(DOM_GRAY_DARK,9,italic=True); leg.fill = fill(DOM_GRAY_LIGHT)
leg.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
ws1.row_dimensions[total_row_e].height = 50

# Rodapé
ws1.merge_cells(start_row=total_row_e+1, start_column=1, end_row=total_row_e+1, end_column=N_COLS_E)
ft = ws1.cell(row=total_row_e+1, column=1,
    value=f"DOM Incorporação  •  Inteligência de Mercado  •  Planilha Mestre v{VERSION} (Fase 1)")
ft.font = Font(name="Calibri",color=DOM_GRAY_MID,size=8,italic=True)
ft.alignment = Alignment(horizontal="right",vertical="center")

# ── ABA 2: INCORPORADORAS ──────────────────────────────────────
ws2 = wb.create_sheet("Incorporadoras")
N_COLS_I = 15  # v4.2: removidas RI e Cap. aberto
ws2.row_dimensions[1].height=22; ws2.row_dimensions[2].height=28; ws2.row_dimensions[3].height=18
for r in (1,2,3):
    for c in range(1,N_COLS_I+1):
        ws2.cell(row=r,column=c).fill=fill(DOM_BLACK)
insert_logo(ws2,LOGO_TRANSP,"A1",55)
ws2.merge_cells(start_row=2,start_column=3,end_row=2,end_column=N_COLS_I)
t2=ws2.cell(row=2,column=3,value="INTELIGÊNCIA DE MERCADO — Panorama por Incorporadora")
t2.font=font(DOM_WHITE,16,bold=True); t2.alignment=Alignment(horizontal="left",vertical="center")
ws2.merge_cells(start_row=3,start_column=3,end_row=3,end_column=N_COLS_I)
s2=ws2.cell(row=3,column=3,
    value=f"14 incorporadoras monitoradas  •  v{VERSION}  •  {DATE_STR}")
s2.font=font(DOM_GOLD,10,italic=True); s2.alignment=Alignment(horizontal="left",vertical="center")

headers_i = [
    "Incorporadora","Nº empreend.","VGV total (R$)",
    "VGV lançado 2024","VGV lançado 2025","VGV lançado 2026",
    "Segmentos","Bairros","Ticket médio carteira","R$/m² médio carteira",
    "% carteira com arquivo","Site oficial","Instagram",
    "Posicionamento","Data verif."
]
apply_header_row(ws2,5,headers_i)

# Auto-calcula agregados por incorporadora
I_ROWS = []
for inc_name in INCORPORADORAS:
    meta = I_META.get(inc_name, ("","",""))
    emps = [r for r in E_PROCESSED if r[0]==inc_name]
    n = len(emps)
    vgv_list = [r[16] for r in emps if r[16] is not None]
    vgv_total = sum(vgv_list) if vgv_list else None
    vgv_2024=vgv_2025=vgv_2026=0
    for r in emps:
        if r[16] is None: continue
        y = extract_year(r[7])
        if y==2024: vgv_2024 += r[16]
        elif y==2025: vgv_2025 += r[16]
        elif y==2026: vgv_2026 += r[16]
    segs = sorted(set(r[5] for r in emps if r[5] and r[5]!="—"))
    bairros = sorted(set(r[3] for r in emps if r[3]))
    tickets = []
    for r in emps:
        if r[13] is not None and r[14] is not None:
            tickets.append((r[13]+r[14])/2)
    ticket_med = sum(tickets)/len(tickets) if tickets else None
    precos_m2 = [r[15] for r in emps if r[15] is not None]
    preco_m2_med = sum(precos_m2)/len(precos_m2) if precos_m2 else None
    # % com arquivo: heurística — se tem tabela_local em preços OU estoque
    com_arquivo = sum(1 for r in emps if r[18]=="tabela_local" or r[19]=="tabela_local")
    pct_arquivo = (com_arquivo / n) if n else 0
    I_ROWS.append((
        inc_name, n,
        vgv_total or 0,
        vgv_2024 or 0, vgv_2025 or 0, vgv_2026 or 0,
        ", ".join(segs) if segs else "—",
        ", ".join(bairros) if bairros else "—",
        ticket_med or 0,
        preco_m2_med or 0,
        pct_arquivo,
        meta[0], meta[1], meta[2],  # site, IG, posicionamento (v4.2)
        DATE_STR
    ))

# Ordena: com empreend. primeiro (desc por VGV), depois SEM material
I_ROWS_SORTED = sorted(I_ROWS, key=lambda r:(-(r[1]>0), -(r[2] or 0), r[0]))

formats_i = [None]*N_COLS_I
formats_i[2] = formats_i[3] = formats_i[4] = formats_i[5] = 'R$ #,##0'
formats_i[8] = formats_i[9] = 'R$ #,##0'
formats_i[10] = '0%'

for i, row_data in enumerate(I_ROWS_SORTED):
    row_idx = 6+i
    ws2.row_dimensions[row_idx].height = 38
    empty_row = row_data[1]==0
    row_fill = DOM_WHITE if row_idx%2==0 else DOM_GRAY_LIGHT
    for j, v in enumerate(row_data):
        c = ws2.cell(row=row_idx, column=1+j, value=v)
        c.font = font(DOM_GRAY_DARK,9)
        c.fill = fill(row_fill)
        c.alignment = left() if j in (6,7,11,13) else center()  # v4.2: posicionamento agora é idx 13
        c.border = border_thin()
        if formats_i[j]:
            c.number_format = formats_i[j]
        if empty_row:
            c.fill = fill(DOM_GOLD_LIGHT)
            c.font = font(DOM_GOLD_DARK,9,italic=True)
    # Destaque nome
    ws2.cell(row=row_idx, column=1).font = font(
        DOM_GOLD_DARK if empty_row else DOM_GRAY_DARK, 9, bold=True,
        italic=empty_row)

total_row_i = 6+len(I_ROWS_SORTED)

widths_i = [18, 10, 15, 15, 15, 15, 22, 30, 15, 14, 13, 30, 22, 48, 10]
set_column_widths(ws2, widths_i)
ws2.freeze_panes = "B6"
ws2.auto_filter.ref = f"A5:{get_column_letter(N_COLS_I)}{total_row_i-1}"

# Legenda aba 2
ws2.merge_cells(start_row=total_row_i, start_column=1, end_row=total_row_i, end_column=N_COLS_I)
leg2 = ws2.cell(row=total_row_i, column=1,
    value="Linhas DOURADAS = incorporadoras SEM material mapeado (0 empreend.) — prioridade de pesquisa web + contato local.    "
          "VGV total = soma do VGV estimado de todos os empreend.    VGV por ano = subconjunto por ano de lançamento.")
leg2.font = font(DOM_GRAY_DARK,9,italic=True); leg2.fill = fill(DOM_GRAY_LIGHT)
leg2.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
ws2.row_dimensions[total_row_i].height = 42

ws2.merge_cells(start_row=total_row_i+1, start_column=1, end_row=total_row_i+1, end_column=N_COLS_I)
ft2 = ws2.cell(row=total_row_i+1, column=1,
    value=f"DOM Incorporação  •  Inteligência de Mercado  •  Planilha Mestre v{VERSION} (Fase 1)")
ft2.font = Font(name="Calibri",color=DOM_GRAY_MID,size=8,italic=True)
ft2.alignment = Alignment(horizontal="right",vertical="center")

# ═══════════════════════════════════════════════════════════════
# ABA 3 — COMPOSIÇÃO (v8.0+) — 1 linha por (empreendimento, tipologia)
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Composição")
N_COLS_C = 11  # v6.2: +Total tipologia
HEADERS_C = ["Incorporadora", "Empreendimento", "Tipologia",
             "Total tipologia", "Disponíveis",  # v6.2: schema 10→11 col
             "Área mín (m²)", "Área máx (m²)",
             "Ticket mín (R$)", "Ticket máx (R$)",
             "R$/m² médio", "Origem"]

insert_logo(ws3, LOGO_TRANSP, "A1", 55)
ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N_COLS_C)
title_c = ws3.cell(row=2, column=1, value="Composição por Tipologia")
title_c.font = Font(name="Calibri", color=DOM_GOLD_DARK, size=14, bold=True)
title_c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[2].height = 22

ws3.merge_cells(start_row=3, start_column=1, end_row=3, end_column=N_COLS_C)
sub_c = ws3.cell(row=3, column=1,
    value=f"1 linha por (empreendimento, tipologia)  •  Visão analítica derivada de tabelas locais  •  v{VERSION}")
sub_c.font = Font(name="Calibri", color=DOM_GRAY_MID, size=10, italic=True)
sub_c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[3].height = 18

for j, h in enumerate(HEADERS_C):
    c = ws3.cell(row=5, column=1+j, value=h)
    c.font = font(DOM_WHITE, 10, bold=True)
    c.fill = fill(DOM_GRAY_DARK)
    c.alignment = center()
    c.border = border_thin()
ws3.row_dimensions[5].height = 28

TIPO_ORDER_C = ["Studio", "1D", "2D", "3D", "4D", "Lote"]
C_RAW_SORTED = sorted(C_RAW, key=lambda r: (r[0], r[1], TIPO_ORDER_C.index(r[2]) if r[2] in TIPO_ORDER_C else 99))

formats_c = [None]*N_COLS_C
formats_c[3] = formats_c[4] = '0'  # Total tipologia + Disponíveis
formats_c[5] = formats_c[6] = '0.00" m²"'
formats_c[7] = formats_c[8] = formats_c[9] = 'R$ #,##0'

# v6.2: rebuild C_RAW_SORTED com 11 col: insere Total tipologia entre Tipologia (idx 2) e Disponíveis (idx 3 antigo)
def expand_c_row_v62(row):
    inc, emp, tip = row[0], row[1], row[2]
    disp = row[3]
    rest = row[4:]  # area_min, area_max, ticket_min, ticket_max, rsm2, origem
    total_tip, origem_revisada = TOTAL_TIP_DICT.get((inc, emp, tip), (None, row[9]))
    return (inc, emp, tip, total_tip, disp) + rest[:-1] + (origem_revisada,)

C_RAW_EXPANDED = [expand_c_row_v62(r) for r in C_RAW_SORTED]

for i, row_data in enumerate(C_RAW_EXPANDED):
    row_idx = 6+i
    ws3.row_dimensions[row_idx].height = 22
    row_fill = DOM_WHITE if row_idx%2==0 else DOM_GRAY_LIGHT
    for j, v in enumerate(row_data):
        c = ws3.cell(row=row_idx, column=1+j, value=v)
        c.font = font(DOM_GRAY_DARK, 10)
        c.fill = fill(row_fill)
        c.alignment = center() if j not in (0,1,10) else left()
        c.border = border_thin()
        if formats_c[j]:
            c.number_format = formats_c[j]
    ws3.cell(row=row_idx, column=2).font = font(DOM_GRAY_DARK, 10, bold=True)
    ws3.cell(row=row_idx, column=3).font = font(DOM_GOLD_DARK, 10, bold=True)
    ws3.cell(row=row_idx, column=4).font = font(DOM_GOLD_DARK, 10, bold=True)  # Total tipologia destaque
    # Marca visualmente entries de origem estimativa_*
    origem_val = row_data[10] or ''
    if origem_val.startswith('estimativa_distribuição') or origem_val.endswith('_pro_rata'):
        for j in range(N_COLS_C):
            ws3.cell(row=row_idx, column=1+j).font = font(DOM_GRAY_MID, 10, italic=True)

total_row_c = 6 + len(C_RAW_SORTED)
widths_c = [18, 28, 12, 11, 13, 13, 16, 16, 14, 18]
set_column_widths(ws3, widths_c)
ws3.freeze_panes = "C6"
ws3.auto_filter.ref = f"A5:{get_column_letter(N_COLS_C)}{total_row_c-1}"

ws3.merge_cells(start_row=total_row_c, start_column=1, end_row=total_row_c, end_column=N_COLS_C)
leg3 = ws3.cell(row=total_row_c, column=1,
    value=f"Aba alimentada de tabelas locais arquivadas. Heurística tipologia x área (<40 Studio, 40-55 1D, 55-75 2D, 75-95 3D, >95 4D). "
          f"Empreendimentos sem entry aqui = ainda sem tabela detalhada extraível (roadmap: Lote 2 e 3). v{VERSION}.")
leg3.font = font(DOM_GRAY_DARK, 9, italic=True); leg3.fill = fill(DOM_GRAY_LIGHT)
leg3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws3.row_dimensions[total_row_c].height = 36

ws3.merge_cells(start_row=total_row_c+1, start_column=1, end_row=total_row_c+1, end_column=N_COLS_C)
ft3 = ws3.cell(row=total_row_c+1, column=1,
    value=f"DOM Incorporação  •  Inteligência de Mercado  •  Composição v{VERSION} (Fase 1)")
ft3.font = Font(name="Calibri", color=DOM_GRAY_MID, size=8, italic=True)
ft3.alignment = Alignment(horizontal="right", vertical="center")

# ═══════════════════════════════════════════════════════════════
# VALIDAÇÕES AUTOMÁTICAS v9.3 (PADRAO §3.6 + §3.7)
# ═══════════════════════════════════════════════════════════════

# §3.6 — soma C_RAW vs total quando origem=tabela_local_completa (threshold 5%)
THRESHOLD_PCT = 5.0
warnings_36 = []
for entry in E_RAW:
    inc, emp = entry[0], entry[1]
    total = entry[6]
    origem_total = entry[24] if len(entry) > 24 else None
    if origem_total != 'tabela_local_completa' or total is None:
        continue
    soma_comp = sum(c[3] for c in C_RAW if c[0] == inc and c[1] == emp)
    if soma_comp == 0:
        continue
    diff_pct = abs(total - soma_comp) / total * 100
    if diff_pct > THRESHOLD_PCT:
        warnings_36.append(
            f"  WARN {inc} | {emp}: total={total} mas Σ C_RAW={soma_comp} ({diff_pct:.1f}% diff)"
        )

# §3.7.C.1 — Anti-duplicação: (inc, emp, tipologia) único em C_RAW
from collections import Counter
keys_comp = Counter((c[0], c[1], c[2]) for c in C_RAW)
errors_dup = [f"  ERROR duplicado: {inc} | {emp} | {tipo} aparece {n}x"
              for (inc, emp, tipo), n in keys_comp.items() if n > 1]

# §3.7.C.2 — Heurística vs Tipologia declarada
# Se empreend é mono-tipologia E_RAW (Tipologia tem 1 valor sem ;), comparar com C_RAW
warnings_heur = []
for entry in E_RAW:
    inc, emp = entry[0], entry[1]
    tipologia_decl = entry[12] if entry[12] else ""
    # Mono se não tem ; e não é "—"
    if ';' not in tipologia_decl and tipologia_decl not in ('', '—', 'Lote'):
        # Achar entries C_RAW desse empreend
        comp_tipos = set(c[2] for c in C_RAW if c[0] == inc and c[1] == emp)
        if comp_tipos and tipologia_decl not in comp_tipos:
            warnings_heur.append(
                f"  WARN {inc} | {emp}: E_RAW.Tipologia='{tipologia_decl}' mas C_RAW tem {sorted(comp_tipos)}"
            )

# §3.7.C.3 — Cobertura: empreend com tabela arquivada mas 0 entries C_RAW
import re as _re_cov
warnings_cov = []
emps_com_comp = set((c[0], c[1]) for c in C_RAW)
script_dir = pathlib.Path(__file__).resolve().parent.parent.parent
if script_dir.exists():
    for inc_dir in script_dir.iterdir():
        if not inc_dir.is_dir() or not _re_cov.match(r'^\d+_', inc_dir.name):
            continue
        for emp_dir in inc_dir.iterdir():
            if not emp_dir.is_dir() or emp_dir.name.startswith('_'):
                continue
            tab_dir = emp_dir / 'TABELA'
            if tab_dir.exists() and any(tab_dir.glob('*.pdf')):
                # Tem tabela. Verificar se algum empreend de E_RAW desta inc_dir está no C_RAW
                # Heurística: nome empreend = pasta sem sufixo _MMAAAA
                emp_name = _re_cov.sub(r'_\d{6}$', '', emp_dir.name).replace('_', ' ')
                # Procurar match em E_RAW
                for entry in E_RAW:
                    e_inc, e_emp = entry[0], entry[1]
                    e_emp_norm = e_emp.lower().replace('í', 'i').replace('ã', 'a').replace('ç', 'c').replace('é', 'e').replace('ó', 'o').replace('ô', 'o').replace('á', 'a').replace("'", '').replace(' ', '')
                    p_emp_norm = emp_name.lower().replace("'", '').replace(' ', '')
                    if e_emp_norm.startswith(p_emp_norm) or p_emp_norm.startswith(e_emp_norm):
                        if (e_inc, e_emp) not in emps_com_comp:
                            warnings_cov.append(f"  WARN {e_inc} | {e_emp}: tem tabela em /{emp_dir.name}/ mas zero entries em C_RAW")
                        break

# Imprimir resumo
def _print_section(title, items, prefix='⚠'):
    if items:
        print(f"\n{prefix} {title}: {len(items)}")
        for w in items:
            print(w)

_print_section("VALIDAÇÃO §3.6 — soma C_RAW vs total (>5%)", warnings_36)
_print_section("VALIDAÇÃO §3.7.C.1 — duplicação em C_RAW", errors_dup, prefix='✗ ERROR')
_print_section("VALIDAÇÃO §3.7.C.2 — heurística vs Tipologia declarada", warnings_heur)
_print_section("VALIDAÇÃO §3.7.C.3 — cobertura (tabela arquivada sem C_RAW)", warnings_cov)

# §3.7.C.4 — INVARIANTE v6.2: Σ Total tipologia = E_RAW.Total para todo empreend. com Total apurado
warnings_374_forte = []  # fontes 1-4 não fechando (precisa buscar mais)
warnings_374_estimativa = []  # estimativas não fechando (auto-ajustadas, info)
ok_374 = []  # fechados ✓
sum_total_tip_emp = {}
for (inc, emp, tip), (total_tip, origem_rev) in TOTAL_TIP_DICT.items():
    if total_tip is None: continue
    sum_total_tip_emp[(inc, emp)] = sum_total_tip_emp.get((inc, emp), 0) + total_tip

for entry in E_RAW:
    inc, emp = entry[0], entry[1]
    total_emp = entry[6]
    if total_emp is None: continue
    soma = sum_total_tip_emp.get((inc, emp))
    if soma is None: continue
    diff = total_emp - soma
    if diff == 0:
        ok_374.append((inc, emp))
    else:
        # Verificar origem dominante das entries
        origens_emp = set()
        for (i_, e_, t_), (tt, ori) in TOTAL_TIP_DICT.items():
            if i_ == inc and e_ == emp:
                origens_emp.add(ori or '')
        is_estimativa = all(o.startswith('estimativa_distribuição') for o in origens_emp if o)
        msg = f"  {inc} | {emp}: Total={total_emp} mas Σ Total tipologia={soma} (∆={diff:+d}) origens={origens_emp}"
        if is_estimativa:
            warnings_374_estimativa.append(msg + " — estimativa será reconciliada")
        else:
            warnings_374_forte.append(msg + " — buscar mais Composição (Total não muda)")

# Reconciliação automática de estimativas: ajusta Total tipologia majoritária pra fechar
for entry in E_RAW:
    inc, emp = entry[0], entry[1]
    total_emp = entry[6]
    if total_emp is None: continue
    keys_emp = [(k, v) for k, v in TOTAL_TIP_DICT.items() if k[0] == inc and k[1] == emp]
    if not keys_emp: continue
    # Aplica só se TODAS as entries são estimativa_*
    all_est = all((v[1] or '').startswith('estimativa_distribuição') for _, v in keys_emp)
    if not all_est: continue
    soma = sum(v[0] for _, v in keys_emp if v[0] is not None)
    if soma == total_emp: continue
    diff = total_emp - soma
    # Pegar a tipologia majoritária (maior Total tip)
    keys_emp_valid = [(k, v) for k, v in keys_emp if v[0] is not None]
    if not keys_emp_valid: continue
    keys_emp_valid.sort(key=lambda x: -x[1][0])
    k_maior, (tt_maior, ori_maior) = keys_emp_valid[0]
    TOTAL_TIP_DICT[k_maior] = (tt_maior + diff, ori_maior)

_print_section("VALIDAÇÃO §3.7.C.4 — invariante v6.2: fontes fortes (1-4) não fechando com Total",
               warnings_374_forte, prefix='⚠ WARN')
if warnings_374_estimativa:
    print(f"\nℹ §3.7.C.4 — estimativas reconciliadas automaticamente: {len(warnings_374_estimativa)}")
    for w in warnings_374_estimativa:
        print(w)
print(f"\n✓ §3.7.C.4 — invariante Σ=Total fechada exato: {len(ok_374)}/{len(set(sum_total_tip_emp.keys()))} empreend.")

# §3.9 — Validação Mês de Lançamento: estimativa_T-36 desatualizada (v9.5+)
warnings_39 = []
from datetime import datetime as _dt
hoje = _dt.now()
for entry in E_RAW:
    inc, emp = entry[0], entry[1]
    origem_lanc = entry[20] if len(entry) > 20 else None
    data_verif_str = entry[22] if len(entry) > 22 else None
    if origem_lanc == 'estimativa_T-36' and data_verif_str:
        try:
            data_verif = _dt.strptime(data_verif_str, '%d/%m/%Y')
            dias = (hoje - data_verif).days
            if dias > 180:
                warnings_39.append(
                    f"  WARN {inc} | {emp}: origem=estimativa_T-36 há {dias} dias (verif {data_verif_str}) — buscar fonte real"
                )
        except Exception:
            pass

_print_section("VALIDAÇÃO §3.9 — Mês Lançamento (estimativa_T-36 > 180d)", warnings_39)

# §3.10 — Validação Bairro/Região (v10.0+)
warnings_310 = []
for entry in E_RAW:
    inc, emp = entry[0], entry[1]
    bairro = entry[3]
    origem_bairro = entry[26] if len(entry) > 26 else None
    if bairro and origem_bairro is None:
        warnings_310.append(f"  WARN {inc} | {emp}: Bairro='{bairro}' mas Origem Bairro=None — preencher por §3.10")
    if bairro in ('São Luís', 'Não identificado'):
        warnings_310.append(f"  WARN {inc} | {emp}: Bairro='{bairro}' é genérico — refinar com book/site")

_print_section("VALIDAÇÃO §3.10 — Bairro/Região", warnings_310)

if not (warnings_36 or errors_dup or warnings_heur or warnings_cov or warnings_39 or warnings_310):
    print("✓ Validações §3.6 + §3.7 + §3.9 + §3.10: todas passaram")

# ═══════════════════════════════════════════════════════════════
# SALVAR — usa a pasta NFD (a real do usuário, com .DS_Store) para
# evitar criar pasta fantasma NFC por causa do Unicode do nome.
# ═══════════════════════════════════════════════════════════════
# v6.5: DST_BASE derivado do próprio caminho do script (auto-suficiente)
# /<...>/01.Inteligência Mercado/00_ESTUDO_CONSOLIDADO/_PADRAO_FASE_1/gerar_planilha.py
#  → parent = _PADRAO_FASE_1
#  → parent.parent = 00_ESTUDO_CONSOLIDADO
SCRIPT_PARENT = pathlib.Path(__file__).resolve().parent.parent
DST_BASE = str(SCRIPT_PARENT)
# Verificar NFD (memória feedback_unicode_nfd_paths.md): garantir que a forma do
# path resolvido pelo pathlib bate com a forma real do filesystem.
if not os.path.exists(DST_BASE):
    raise FileNotFoundError(f"Pasta resolvida não existe: {DST_BASE}")
OUT = os.path.join(DST_BASE, f"Planilha_Mestre_Panorama_v{VERSION}.xlsx")
wb.save(OUT)

print(f"✓ Salvo: {OUT}")
print(f"  Empreendimentos: {len(E_PROCESSED)}")
print(f"  Incorporadoras:  {len(I_ROWS)} (ativas: {sum(1 for r in I_ROWS if r[1]>0)}, sem material: {sum(1 for r in I_ROWS if r[1]==0)})")
print(f"  Composição:      {len(C_RAW)} linhas / {sum(r[3] for r in C_RAW)} unidades extraídas")
print(f"  VGV total mapeado: R$ {sum(r[16] for r in E_PROCESSED if r[16]):,.0f}")
print(f"  Preço médio calculado para: {sum(1 for r in E_PROCESSED if r[15])} de {len(E_PROCESSED)} empreend.")
