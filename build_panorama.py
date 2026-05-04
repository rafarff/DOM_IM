#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_panorama.py
─────────────────────────────────────────────────────────────────────────────
Gera o dashboard HTML "Panorama de Lançamentos — Grande São Luís" a partir
da Planilha Mestre de Inteligência de Mercado da DOM Incorporação.

Fluxo:
  1. Localiza a versão mais recente de `Planilha_Mestre_Panorama_v*.xlsx`
  2. Lê a aba "Empreendimentos"
  3. Filtra lançamentos 2025/2026 (ou em comercialização nesse ciclo)
  4. Geocodifica cada empreendimento por bairro (coordenadas aproximadas)
  5. Gera o arquivo `Panorama_Lancamentos_2025_2026.html` na pasta raiz

Uso:
    python3 build_panorama.py
    python3 build_panorama.py --all   # Inclui todos (não filtra por ciclo)

Autor: DOM Incorporação · Inteligência de Mercado
─────────────────────────────────────────────────────────────────────────────
"""
from __future__ import annotations
import argparse
import json
import math
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl não instalado. Rode: pip install openpyxl --break-system-packages")


# ─── Caminhos ────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_HTML = SCRIPT_DIR / "index.html"  # nome padrão do GitHub Pages

# v8.1.0 (04/05/2026): classificação Panorama em 3 tabelas (decisão Rafael 04/05).
#   Tabela A — Empreendimentos com tabela de vendas (acesso a tickets unidade-a-unidade)
#   Tabela B — Breve Lançamentos mapeados (sem tabela: prática de mercado nessa fase)
#   Tabela C — Demais (lançados sem tabela, info parcial)
# Rafael mantém manualmente a lista de Breve Lançamentos abaixo.
BREVE_LANCAMENTO_NAMES = {
    "Nexus Renascença",   # Ergus, lanc. 04/2026 — site sem ficha técnica
    "Bay View",           # Alfa Engenharia, lanc. 05/2026 — Rafael 04/05/2026
    "Villa Terrari",      # DOM Incorporação, lanc. 07/2026 — interno DOM
    "Dom Manuel",         # DOM Incorporação, lanc. 08/2026 — interno DOM
    "Dom Guilherme",      # DOM Incorporação, lanc. 12/2026 — interno DOM
    "Dom Rafael",         # DOM Incorporação, lanc. 10/2026 — interno DOM
    "Dom Roberto",        # DOM Incorporação, lanc. 11/2026 — interno DOM
}

def fase_comercial(empreendimento: str, orig_precos: str) -> str:
    """
    Classifica fase comercial em 3 buckets para as Tabelas A/B/C do Panorama.
    - lancado_com_tabela: temos a tabela de vendas (orig_precos in tabela|tabela_local)
    - breve_lancamento  : empreendimento listado em BREVE_LANCAMENTO_NAMES (manual)
    - lancado_sem_tabela: demais
    """
    if empreendimento in BREVE_LANCAMENTO_NAMES:
        return "breve_lancamento"
    # v11.8: aceita informado_manualmente como A (caso Rafael conhece o empreend. com certeza
    # mesmo sem tabela formal — ex: empreend. próprios DOM, ou casos tipo Dom Antônio 100% vendido)
    if orig_precos in ("tabela", "tabela_local", "informado_manualmente"):
        return "lancado_com_tabela"
    return "lancado_sem_tabela"


def discover_search_dirs() -> list[Path]:
    """
    Descobre dinamicamente as pastas candidatas onde a Planilha Mestre pode estar.
    Resolve problema de normalização Unicode (NFD vs NFC) no macOS procurando
    irmãos da pasta do script pelo padrão "01.*" + subpasta "00_ESTUDO_CONSOLIDADO".
    """
    dirs: list[Path] = []
    parent = SCRIPT_DIR.parent
    # 1. Procura pasta irmã "01.*" (Inteligência Mercado numerada)
    for sibling in parent.iterdir():
        if sibling.is_dir() and sibling.name.startswith("01."):
            estudo = sibling / "00_ESTUDO_CONSOLIDADO"
            if estudo.exists():
                dirs.append(estudo)
            dirs.append(sibling)
    # 2. Subpasta no próprio diretório do script
    local_estudo = SCRIPT_DIR / "00_ESTUDO_CONSOLIDADO"
    if local_estudo.exists():
        dirs.append(local_estudo)
    # 3. Diretório do script como último recurso
    dirs.append(SCRIPT_DIR)
    return dirs


# ─── Coordenadas aproximadas de bairros de São Luís/MA ───────────────────────
COORDS_BAIRRO = {
    # São Luís / MA — coordenadas no CENTRO geográfico de cada bairro (em terra firme).
    # Metodologia: afastado no mínimo ~400m da linha de costa ou rio,
    # priorizando áreas densamente urbanizadas.
    "Calhau":            (-2.5030, -44.2740),
    "Renascença":        (-2.5085, -44.2945),
    "Renascença II":     (-2.5110, -44.2965),
    "Jardim Renascença": (-2.5090, -44.2830),
    "Ponta d'Areia":     (-2.4985, -44.3000),
    "Ponta D'Areia":     (-2.4985, -44.3000),
    "Ponta do Farol":    (-2.4925, -44.2960),
    "São Marcos":        (-2.5030, -44.2855),
    "Cohama":            (-2.5330, -44.2760),
    "Cohab Anil IV":     (-2.5535, -44.2655),
    "Anil":              (-2.5485, -44.2645),
    "Cohab Anil":        (-2.5535, -44.2655),
    "Jardim Eldorado":   (-2.5590, -44.2485),
    "Turú":              (-2.5620, -44.2515),
    "Turu":              (-2.5620, -44.2515),  # alias sem acento (senso comum, §3.10)
    "Cantinho do Céu":   (-2.5650, -44.2660),
    "Araçagi":           (-2.4920, -44.2200),
    "Araçagy":           (-2.4920, -44.2200),
    "São Francisco":     (-2.5285, -44.3000),
    "Maranhão Novo":     (-2.5585, -44.2810),
    "Santo Amaro":       (-2.5635, -44.2685),
    "Cohatrac":          (-2.5685, -44.2280),
    "Iguaíba":           (-2.5730, -44.3580),
    "São Luís":          (-2.5310, -44.3068),
}

# Overrides de endereço-específico quando temos endereço preciso
COORDS_ENDERECO = {
    "Dom Lucas":            (-2.5665, -44.2675),
    "Dom José":             (-2.5594, -44.2492),
    "Edifício Bossa":       (-2.5002, -44.2788),
    "Al Mare Tirreno":      (-2.5018, -44.2870),
    "Le Noir":              (-2.5115, -44.2960),
    "Entre Rios":           (-2.5090, -44.2940),
    "Residencial Novo Anil":(-2.5540, -44.2660),
    "Edifício Sanpaolo":    (-2.5335, -44.2755),
}


# ─── Paleta DOM (cores por incorporadora no mapa) ────────────────────────────
INC_COLORS = {
    "Mota Machado":   "#B87333",  # bronze
    "Treviso":        "#8B2E2E",
    "Delman":         "#1B4584",
    "Canopus":        "#5D6E3C",
    "Berg Engenharia":"#6A4C93",
    "Niágara":        "#0F7B6C",
    "Castelucci":     "#A0522D",
    "Alfa Engenharia":"#2F4858",
    "Monteplan":      "#D17A22",
    "Sá Cavalcante":  "#7A2E7C",
    "Ergus":          "#3E6B93",
    "Lua Nova":       "#BA5A31",
    "Franere":        "#4D4D4D",
    "MB Engenharia":  "#8C8C8C",
    "DOM Incorporação":"#C9A84C",  # DOURADO DOM (marca da casa — destaque)
    "Hiali":          "#8C3B4A",  # vinho/bordô
}


# ─── Utilidades ──────────────────────────────────────────────────────────────
def find_latest_planilha() -> Path:
    """Busca a Planilha_Mestre_Panorama_v*.xlsx mais recente."""
    search_dirs = discover_search_dirs()
    candidates: list[Path] = []
    for d in search_dirs:
        if d.exists():
            for p in d.glob("Planilha_Mestre_Panorama_v*.xlsx"):
                candidates.append(p)
    if not candidates:
        sys.exit(
            "❌ Não encontrei nenhuma Planilha_Mestre_Panorama_v*.xlsx\n"
            f"   Busquei em: {[str(d) for d in search_dirs]}"
        )
    # Ordena por versão (extraindo X.Y.Z do nome)
    def version_key(p: Path) -> tuple:
        m = re.search(r"v(\d+)(?:\.(\d+))?(?:\.(\d+))?", p.stem)
        if not m:
            return (0, 0, 0, p.stat().st_mtime)
        parts = tuple(int(g) if g else 0 for g in m.groups())
        return parts + (p.stat().st_mtime,)
    candidates.sort(key=version_key, reverse=True)
    return candidates[0]


def parse_lancamento_sort(lancamento: str) -> int:
    """
    Converte string de lançamento em inteiro AAAAMM para ordenação cronológica.
    PADRAO v2.0 §1: formato esperado é SEMPRE MM/AAAA (com flag opcional ⚠ T-36).
    Exemplos:
        '04/2026'            → 202604
        '04/2025 ⚠ T-36'     → 202504
        '10/2025'            → 202510
        '—' / vazio          → 0 (faltam dados — vai para o fim da lista)
        'AAAA' puro          → 0 (formato inválido — força correção, vai pro fim)
    """
    if not lancamento or lancamento == "—":
        return 0
    s = str(lancamento).strip()
    # Aceita SOMENTE MM/AAAA (com ou sem ⚠ T-36)
    m = re.match(r"^(\d{1,2})/(\d{4})( ⚠ T-36)?$", s)
    if m:
        month, year = int(m.group(1)), int(m.group(2))
        return year * 100 + max(1, min(12, month))
    # Formato inválido (AAAA puro, ~AAAA, etc.) → empurrar pro fim
    # Sinaliza visualmente que falta dado preciso
    return 0


def tem_endereco_completo(endereco: str) -> bool:
    """
    Heurística para decidir se o endereço identifica uma localização precisa
    (rua + número/quadra) ou é apenas o bairro.
    Critério (PADRAO §1 col 3): pin no mapa SÓ se houver endereço completo.
    """
    if not endereco:
        return False
    e = str(endereco).strip()
    if e.startswith("Endereço não localizado"):
        return False
    if re.search(r'\b(Rua|Av\.?|Avenida|Travessa|Tv\.?|Praça|Pç\.?|Estrada|Rod\.|Rodovia)\s+', e):
        return True
    if re.search(r'\b[A-Z0-9]{4,8}\+[A-Z0-9]{2,3}\b', e):
        return True
    return False


def geocode_bairro(bairro: str) -> tuple[float, float] | None:
    """
    Retorna (lat, lng) aproximados para o bairro, ou None se o bairro
    não puder ser identificado.

    Regras:
      - "São Luís" (placeholder genérico de bairro desconhecido) → None
      - Bairros vazios/nulos → None
      - Bairros mapeados → coordenadas conhecidas
      - Bairros não mapeados → tenta match parcial; se falhar → None
    """
    if not bairro:
        return None
    b = bairro.strip()
    # "São Luís" genérico = bairro não identificado
    if b.lower() in ("são luís", "sao luis", "são luís - ma", "—", "-"):
        return None
    # Busca exata
    if b in COORDS_BAIRRO:
        c = COORDS_BAIRRO[b]
        # Ignora se o match for o default "São Luís"
        return c if b != "São Luís" else None
    # Busca case-insensitive / aproximada
    for key, coord in COORDS_BAIRRO.items():
        if key == "São Luís":
            continue  # nunca usar como match aproximado
        if key.lower() == b.lower():
            return coord
        if key.lower() in b.lower() or b.lower() in key.lower():
            return coord
    return None


# ─── Parse de data de lançamento ─────────────────────────────────────────────
def parse_lancamento(raw: str, orig_col: str = "") -> tuple[str, str]:
    """
    Normaliza a string de "Mês lançamento" e deriva a Origem da informação.

    Retorna (data_formatada, origem):
      - '04/2026'           → ('04/2026', 'Tabela')
      - '10/2025'           → ('10/2025', 'Tabela')
      - '2026'              → ('2026',    'Book')
      - '~2026'             → ('2026',    'Estimado')
      - '~2025'             → ('2025',    'Estimado')
      - '04/2025 ⚠ T-36'    → ('04/2025', 'T-36')
      - '—' ou vazio        → ('—',      '—')

    `orig_col` = valor da coluna "Orig. lançamento" (imprensa/book/tabela),
    usado como fonte primária quando disponível.
    """
    if not raw or str(raw).strip() in ("—", "-", ""):
        # Data vazia — respeita origem informada (ex: 'pendente', 'estimado-fraco')
        if orig_col and str(orig_col).strip() not in ("", "—", "N/A"):
            origem = str(orig_col).strip().title()
            return ("—", origem)
        return ("—", "—")
    s = str(raw).strip()

    # Identificar origem a partir de marcadores no texto
    is_approx = "~" in s
    is_t36 = "T-36" in s or "t-36" in s

    # Extrair data normalizada
    m_full = re.search(r"(\d{1,2})/(\d{4})", s)
    if m_full:
        mo, yr = int(m_full.group(1)), int(m_full.group(2))
        data_fmt = f"{mo:02d}/{yr}"
    else:
        m_year = re.search(r"(\d{4})", s)
        data_fmt = m_year.group(1) if m_year else s

    # Definir origem com a seguinte prioridade:
    # 1. T-36 (marcador explícito no texto — forma mais crua)
    # 2. Estimado (marcador ~)
    # 3. Valor da coluna "Orig. lançamento" da planilha
    # 4. Fallback: "Book" (se só tem ano) ou "Tabela" (se tem MM/AAAA)
    if is_t36:
        origem = "T-36"
    elif is_approx:
        origem = "Estimado"
    elif orig_col and str(orig_col).strip() not in ("", "—", "N/A"):
        oc = str(orig_col).strip().lower()
        mapping = {
            "tabela": "Tabela",
            "book": "Book",
            "imprensa": "Imprensa",
            "site": "Site",
        }
        origem = mapping.get(oc, str(orig_col).strip().title())
    elif m_full:
        origem = "Tabela"
    else:
        origem = "Book"
    return (data_fmt, origem)


def should_include(row: dict, include_all: bool = False) -> bool:
    """
    Filtro removido na v6.0 (decisão do Rafael 27/04/2026).
    
    Antes da v6.0 havia distinção entre "ativos no ciclo" (Panorama) e "todos"
    (Dados Completos). Após análise, decidimos que o Panorama deve mostrar
    TODOS os empreendimentos mapeados — sem filtro de Status nem de data de
    lançamento. As 3 abas do HTML mostram o mesmo universo de 45 entries,
    diferenciando-se apenas pela visualização (mapa, pendências, tabela full).
    """
    return True


# ─── Leitura da planilha ─────────────────────────────────────────────────────
def read_planilha(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Empreendimentos"]
    # Cabeçalho está na linha 5
    headers = [c.value for c in ws[5]]
    rows: list[dict] = []
    for r in range(6, ws.max_row + 1):
        row = {headers[i]: c.value for i, c in enumerate(ws[r])}
        if row.get("Incorporadora") and row.get("Empreendimento"):
            rows.append(row)
    return rows


def read_composicao(path: Path) -> list[dict]:
    """v7.0+: lê aba Composição com 12 colunas (1 linha por empreend × tipologia × planta).

    Schema v7.0 (xlsx v11.0+): adiciona colunas 'Planta' e 'Área (m²)' (única, não mais range).
    Backward-compat: lê v6.2 (11 col com Área mín/máx) se for xlsx legacy.

    Output dict (estável):
      - tipologia: agrupador
      - planta: label da planta (vazio = identifica por área)
      - area: valor único (calculado da média min/max se vier de v6.2)
      - total_planta: NOVO v7.0 — quantidade total de unidades dessa planta
      - unidades: Total tipologia (mantido pra compat — análise de oferta)
      - disponiveis: estoque (análise de absorção)
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Composição" not in wb.sheetnames:
        return []
    ws = wb["Composição"]
    headers = [c.value for c in ws[5]]
    rows: list[dict] = []
    for r in range(6, ws.max_row + 1):
        row = {headers[i]: c.value for i, c in enumerate(ws[r]) if i < len(headers)}
        if not (row.get("Incorporadora") and row.get("Empreendimento") and row.get("Tipologia")):
            continue
        # v7.0: schema 12 col tem "Planta" e "Área (m²)" únicos
        # v6.2 fallback: tem "Área mín/máx" e não tem "Planta"
        is_v70 = "Planta" in headers and "Área (m²)" in headers
        if is_v70:
            area = row.get("Área (m²)")
            planta = row.get("Planta") or ""
            total_planta = row.get("Total planta")
            area_min = area
            area_max = area
        else:
            area_min = row.get("Área mín (m²)")
            area_max = row.get("Área máx (m²)")
            area = (area_min + area_max) / 2 if (area_min and area_max) else (area_min or area_max)
            planta = ""
            total_planta = None
        total_tip = row.get("Total tipologia") or row.get("Nº Unidades")
        disp = row.get("Disponíveis") if "Disponíveis" in row else row.get("Nº Unidades")
        rows.append({
            "incorporadora": row.get("Incorporadora"),
            "empreendimento": row.get("Empreendimento"),
            "tipologia": row.get("Tipologia"),
            "planta": planta,                  # v7.0
            "area": area,                       # v7.0 (valor único)
            "total_planta": total_planta,      # v7.0
            "unidades": total_tip,              # mantido — Total tipologia (análise de oferta)
            "disponiveis": disp,                # mantido — estoque (análise de absorção)
            "area_min": area_min,               # back-compat
            "area_max": area_max,               # back-compat
            "ticket_min": row.get("Ticket mín (R$)"),
            "ticket_max": row.get("Ticket máx (R$)"),
            "rsm2": row.get("R$/m² médio"),
            "origem": row.get("Origem"),
        })
    return rows


def enrich(rows: list[dict], include_all: bool = False) -> list[dict]:
    """
    Retorna TODAS as linhas enriquecidas, com flag `is_active` indicando
    quais são ativas no ciclo atual (atendem a `should_include`).
    A UI usa `is_active` para filtrar a aba Panorama, mas a aba Dados
    Completos mostra todos.
    """
    bairro_count: dict = defaultdict(int)
    enriched: list[dict] = []
    for r in rows:
        is_active = should_include(r, include_all)
        bairro_raw = r.get("Bairro")
        bairro = str(bairro_raw).strip() if bairro_raw else ""
        emp_name = str(r.get("Empreendimento") or "").strip()
        endereco = str(r.get("Endereço") or "").strip()
        # Política v6.2: pin no mapa SÓ se endereço completo (rua + nº/quadra OU Plus Code).
        endereco_ok = tem_endereco_completo(endereco)
        if emp_name in COORDS_ENDERECO:
            coord = COORDS_ENDERECO[emp_name]
        else:
            coord = geocode_bairro(bairro)

        lat_j: float | None = None
        lng_j: float | None = None
        on_map = False
        if endereco_ok and coord is not None:
            lat, lng = coord
            idx = bairro_count[bairro]
            r_off = 0.0008 * idx
            theta = idx * 2.4
            lat_j = round(lat + r_off * math.cos(theta), 5)
            lng_j = round(lng + r_off * math.sin(theta), 5)
            bairro_count[bairro] += 1
            on_map = True

        lancamento_raw = str(r.get("Mês lançamento") or "—")
        orig_col = r.get("Orig. lançamento") or ""
        data_fmt, origem = parse_lancamento(lancamento_raw, orig_col)

        # Label do bairro para a UI: se não for identificado, explicitamos
        bairro_label = bairro if on_map else (bairro if bairro and bairro.lower() not in ("são luís", "sao luis") else "Não identificado")

        enriched.append({
            "incorporadora":  r.get("Incorporadora"),
            "empreendimento": r.get("Empreendimento"),
            "endereco":       r.get("Endereço") or "",
            "bairro":         bairro_label,
            "tipo":           r.get("Tipo") or "—",
            "segmento":       r.get("Segmento") or "—",
            "unidades":       r.get("Nº unid."),
            "orig_total":     r.get("Origem total unid.") or "N/A",  # v9.0
            "lancamento":     data_fmt,
            "lancamento_origem": origem,
            "lancamento_raw": lancamento_raw,
            "lancamento_sort": parse_lancamento_sort(lancamento_raw),
            "entrega":        str(r.get("Mês entrega") or "—"),
            "area_min":       r.get("Área mín (m²)"),
            "area_max":       r.get("Área máx (m²)"),
            "area_med":       r.get("Tipologia média (m²)"),
            "dorms":          r.get("Tipologia") or "—",  # v6.1: header xlsx renomeado de "Tipologia (dorms)" para "Tipologia"
            "ticket_min":     r.get("Ticket mín (R$)"),
            "ticket_max":     r.get("Ticket máx (R$)"),
            "rsm2":           r.get("R$/m²"),
            "vgv":            r.get("VGV (R$)"),
            "vendido":        r.get("% Vendido"),
            "orig_pct_vendido": r.get("Origem % Vendido") or "N/A",  # v9.4: PADRAO §3.8
            "orig_precos":    r.get("Orig. preços") or "—",
            "orig_estoque":   r.get("Orig. estoque") or "—",
            "orig_lancamento": r.get("Orig. lançamento") or "—",
            "link":           r.get("Link fonte principal") or "",
            "data_verif":     str(r.get("Data verif.") or "—"),
            "obs":            r.get("Observações") or "",
            "lat":            lat_j,
            "lng":            lng_j,
            "on_map":         on_map,
            "is_active":      is_active,
            "fase_comercial": fase_comercial(
                str(r.get("Empreendimento") or "").strip(),
                str(r.get("Orig. preços") or "—")
            ),  # v8.1.0
        })
    # Ordena cronologicamente (mais recente primeiro)
    enriched.sort(key=lambda e: e["lancamento_sort"], reverse=True)
    return enriched


# ─── Template HTML (mantém identidade visual DOM) ────────────────────────────
HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<meta http-equiv="Pragma" content="no-cache">
<meta http-equiv="Expires" content="0">
<title>DOM Incorporação — Panorama de Lançamentos Grande São Luís</title>
<style>
  :root {
    --dom-black: #000000; --dom-gray-dark: #4D4D4D; --dom-gray-mid: #8C8C8C;
    --dom-gray-light: #F2F2F2; --dom-white: #FFFFFF;
    --dom-gold: #C9A84C; --dom-gold-light: #E8D5A3; --dom-gold-dark: #8B6914;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'Calibri', 'Arial', sans-serif; background: var(--dom-gray-light); color: var(--dom-gray-dark); line-height: 1.5; }
  .hero { background: var(--dom-gray-dark); color: var(--dom-white); padding: 32px 40px 28px; border-bottom: 4px solid var(--dom-gold); }
  .hero-inner { max-width: 1400px; margin: 0 auto; display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 16px; }
  .hero-brand { display: flex; align-items: center; gap: 20px; }
  .hero-logo { height: 90px; width: auto; }  /* v6.6: 60 → 90 */
  .hero h1 { font-size: 24px; font-weight: 700; letter-spacing: 2px; color: var(--dom-white); }
  .hero h1 .gold { color: var(--dom-gold); }
  .hero .subtitle { font-size: 13px; color: var(--dom-gold); margin-top: 4px; letter-spacing: 1px; text-transform: uppercase; }
  .hero .meta { font-size: 12px; color: var(--dom-gray-mid); text-align: right; }
  .container { max-width: 1400px; margin: 0 auto; padding: 24px 40px; }
  .kpis { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; margin-bottom: 24px; }
  .kpi { background: var(--dom-white); border-left: 4px solid var(--dom-gold); padding: 18px 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }
  .kpi .label { font-size: 11px; color: var(--dom-gray-mid); text-transform: uppercase; letter-spacing: 1.2px; }
  .kpi .value { font-size: 26px; font-weight: 700; color: var(--dom-black); margin-top: 6px; }
  .kpi .unit { font-size: 13px; color: var(--dom-gray-dark); margin-left: 4px; font-weight: 400; }
  .kpi.highlight { background: var(--dom-gray-dark); color: var(--dom-white); border-left-color: var(--dom-gold); }
  .kpi.highlight .value { color: var(--dom-gold); } .kpi.highlight .label { color: var(--dom-gold-light); } .kpi.highlight .unit { color: var(--dom-white); }
  .filters { background: var(--dom-white); padding: 16px 20px; margin-bottom: 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); display: flex; flex-wrap: wrap; gap: 14px; align-items: center; }
  .filters label { font-size: 11px; color: var(--dom-gray-mid); text-transform: uppercase; letter-spacing: 0.8px; display: block; margin-bottom: 4px; }
  .filters select, .filters input { padding: 8px 12px; border: 1px solid var(--dom-gray-mid); font-family: inherit; font-size: 13px; color: var(--dom-gray-dark); background: var(--dom-white); min-width: 160px; }
  .filters select:focus, .filters input:focus { outline: 2px solid var(--dom-gold); border-color: var(--dom-gold); }
  .filter-group { display: flex; flex-direction: column; }
  .reset-btn { padding: 10px 18px; background: var(--dom-black); color: var(--dom-gold); border: none; cursor: pointer; font-family: inherit; font-size: 12px; letter-spacing: 1px; text-transform: uppercase; font-weight: 700; margin-top: 15px; transition: all 0.2s; }
  .reset-btn:hover { background: var(--dom-gold); color: var(--dom-black); }
  .results-count { margin-left: auto; font-size: 12px; color: var(--dom-gray-dark); margin-top: 15px; }
  .results-count strong { color: var(--dom-gold-dark); font-size: 14px; }

  .table-wrap { background: var(--dom-white); box-shadow: 0 1px 3px rgba(0,0,0,0.08); overflow-x: auto; }
  .table-header { background: var(--dom-black); color: var(--dom-white); padding: 14px 20px; font-weight: 700; font-size: 14px; letter-spacing: 1.5px; text-transform: uppercase; border-bottom: 3px solid var(--dom-gold); }
  table { width: 100%; border-collapse: collapse; font-size: 12.5px; }
  thead tr { background: var(--dom-gray-dark); color: var(--dom-white); }
  thead th { padding: 12px 10px; text-align: left; font-weight: 700; font-size: 11px; letter-spacing: 0.8px; text-transform: uppercase; border-bottom: 2px solid var(--dom-gold); cursor: pointer; user-select: none; position: sticky; top: 0; }
  thead th:hover { background: var(--dom-black); }
  thead th.sorted::after { content: ' ▼'; color: var(--dom-gold); font-size: 9px; }
  thead th.sorted.asc::after { content: ' ▲'; }
  tbody tr { border-bottom: 1px solid var(--dom-gray-light); transition: background 0.15s; }
  tbody tr:nth-child(even) { background: var(--dom-gray-light); }
  tbody tr:hover { background: var(--dom-gold-light); cursor: pointer; }
  tbody td { padding: 10px; vertical-align: middle; }
  .chip { display: inline-block; padding: 3px 10px; border-radius: 20px; font-size: 10.5px; font-weight: 700; letter-spacing: 0.5px; text-transform: uppercase; }
  .chip.seg-luxo { background: var(--dom-black); color: var(--dom-gold); }
  .chip.seg-alto { background: var(--dom-gold); color: var(--dom-black); }
  .chip.seg-medioalto { background: var(--dom-gold-light); color: var(--dom-gold-dark); }
  .chip.seg-medio { background: var(--dom-gray-mid); color: var(--dom-white); }
  .chip.seg-popular { background: #B8DBC5; color: #1F5138; }
  .chip.seg-other { background: var(--dom-gray-light); color: var(--dom-gray-dark); }
  .chip.st-lancamento { background: #FFE4A8; color: #7A5500; }
  .chip.st-pre { background: #E8D5A3; color: var(--dom-gold-dark); }
  .chip.st-comerc { background: #C3DAF9; color: #1B4584; }
  .chip.st-ultimas { background: #F7C8C8; color: #8C2525; }
  /* Origem da data — formatação minimalista com graduação de confiança */
  .origem { display: inline-block; font-size: 10.5px; padding: 1px 7px 1px 8px; border-radius: 3px; color: var(--dom-gray-dark); border-left: 3px solid var(--dom-gray-mid); background: #FAFAF5; font-variant: small-caps; letter-spacing: 0.3px; }
  .origem.strong { border-left-color: var(--dom-black); color: var(--dom-black); font-weight: 600; }
  .origem.medium { border-left-color: var(--dom-gold); color: var(--dom-gold-dark); }
  .origem.weak   { border-left-color: var(--dom-gray-mid); color: var(--dom-gray-mid); font-style: italic; }
  /* Tooltip clicável (ℹ) — info adicional sobre fonte do dado */
  .info-icon { display: inline-block; width: 14px; height: 14px; line-height: 13px; text-align: center; border-radius: 50%; background: var(--dom-gold-light); color: var(--dom-gold-dark); font-size: 10px; font-style: normal; font-weight: 700; cursor: help; margin-left: 4px; vertical-align: middle; transition: all 0.15s; }
  .info-icon:hover { background: var(--dom-gold); color: var(--dom-black); }
  .origem.pending{ border-left-color: #B54B3A; color: #B54B3A; font-weight: 600; }
  .chip.tp-vertical   { background: #3B4371; color: #FFFFFF; }
  .chip.tp-horizontal { background: #5D7A3C; color: #FFFFFF; }
  .chip.tp-other      { background: var(--dom-gray-light); color: var(--dom-gray-dark); }
  .table-intro { background: var(--dom-white); padding: 14px 20px; font-size: 12px; color: var(--dom-gray-dark); border-left: 3px solid var(--dom-gold); margin-top: 18px; margin-bottom: 0; }
  .table-intro strong { color: var(--dom-gold-dark); }
  .table-intro.incomplete { border-left-color: #B54B3A; }
  .table-intro.incomplete strong { color: #B54B3A; }

  .inc-name { font-weight: 700; color: var(--dom-black); }
  .emp-name { font-weight: 700; color: var(--dom-gold-dark); }
  .price { font-variant-numeric: tabular-nums; color: var(--dom-black); font-weight: 600; }
  .dim { color: var(--dom-gray-mid); font-size: 11.5px; }
  .legend { display: flex; gap: 20px; flex-wrap: wrap; background: var(--dom-white); padding: 12px 20px; margin-bottom: 12px; font-size: 11.5px; border-left: 3px solid var(--dom-gold); }
  .legend-item { display: flex; align-items: center; gap: 6px; }
  .legend-dot { width: 12px; height: 12px; border-radius: 50%; border: 2px solid var(--dom-white); box-shadow: 0 0 0 1px var(--dom-gray-mid); }
  .footer { text-align: center; padding: 20px; font-size: 11px; color: var(--dom-gray-mid); border-top: 1px solid var(--dom-gray-mid); margin-top: 32px; }
  .footer strong { color: var(--dom-gold-dark); }
  /* Tab navigation */
  .tabs { background: var(--dom-gray-dark); border-bottom: 3px solid var(--dom-gold); }
  .tabs-inner { max-width: 1400px; margin: 0 auto; padding: 0 40px; display: flex; gap: 2px; }
  .tab-btn { background: transparent; color: var(--dom-gray-mid); border: none; padding: 16px 28px; font-family: inherit; font-size: 12px; font-weight: 700; letter-spacing: 1.5px; text-transform: uppercase; cursor: pointer; border-bottom: 3px solid transparent; margin-bottom: -3px; transition: all 0.2s; }
  .tab-btn:hover { color: var(--dom-gold-light); }
  .tab-btn.active { color: var(--dom-gold); border-bottom-color: var(--dom-gold); }
  .tab-btn .count { font-size: 10px; opacity: 0.7; margin-left: 6px; font-weight: 400; }
  .tab-panel { display: none; }
  .tab-panel.active { display: block; }
  /* Compact table for Dados Completos */
  .tbl-compact table { font-size: 11.5px; }
  .tbl-compact thead th { padding: 10px 8px; font-size: 10px; white-space: nowrap; }
  .tbl-compact tbody td { padding: 8px; white-space: nowrap; }
  .tbl-compact tbody td.wrap { white-space: normal; min-width: 180px; }
  .tbl-compact { overflow-x: auto; max-height: 75vh; }
  /* Dashboard cards */
  .dash-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(380px, 1fr)); gap: 18px; margin-bottom: 24px; }
  .dash-card { background: var(--dom-white); border-left: 4px solid var(--dom-gold); padding: 18px 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }
  .dash-card h3 { font-size: 12px; font-weight: 700; color: var(--dom-gray-mid); text-transform: uppercase; letter-spacing: 1.2px; margin-bottom: 14px; }
  .dash-card .chart-wrap { position: relative; height: 260px; }
  .dash-card-wide { grid-column: 1 / -1; }
  .matrix-tbl { width: 100%; border-collapse: collapse; font-size: 12px; }
  .matrix-tbl th, .matrix-tbl td { padding: 10px 8px; text-align: center; border: 1px solid var(--dom-gray-light); }
  .matrix-tbl th { background: var(--dom-gray-dark); color: var(--dom-white); font-size: 10px; text-transform: uppercase; letter-spacing: 1px; }
  .matrix-tbl td.row-label { text-align: left; font-weight: 600; color: var(--dom-gray-dark); background: var(--dom-gray-light); }
  .matrix-tbl td .cell-val { font-weight: 700; font-size: 14px; }
  /* v7.0 dashboard redesenhado */
  .dash-filters-bar { display: flex; flex-wrap: wrap; gap: 12px; align-items: flex-end; padding: 14px 16px; background: var(--dom-white); border-left: 4px solid var(--dom-gold); margin-bottom: 18px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }
  .dash-section { margin-bottom: 28px; }
  .dash-section-h { font-size: 16px; font-weight: 700; color: var(--dom-black); border-bottom: 2px solid var(--dom-gold); padding-bottom: 6px; margin-bottom: 14px; letter-spacing: 0.5px; }
  .dash-kpis-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr)); gap: 12px; }
  .kpi-gold { border-left-color: var(--dom-gold-dark) !important; }
  .kpi-label { font-size: 11px; color: var(--dom-gray-mid); text-transform: uppercase; letter-spacing: 1px; margin-bottom: 6px; }
  .kpi-value { font-size: 22px; font-weight: 700; color: var(--dom-black); }
  .dash-section-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 14px; margin-bottom: 14px; }
  .dash-grid-half { display: grid; grid-template-columns: 1fr 1fr; gap: 14px; }
  @media (max-width: 900px) { .dash-section-grid, .dash-grid-half { grid-template-columns: 1fr; } }
  .dash-card-mini { background: var(--dom-white); border-left: 3px solid var(--dom-gold); padding: 14px 16px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }
  .dash-card-mini h3 { font-size: 11px; font-weight: 700; color: var(--dom-gray-mid); text-transform: uppercase; letter-spacing: 1px; margin-bottom: 10px; }
  .dash-card-mini .chart-wrap { position: relative; height: 280px; }
  .dash-table-wrap { overflow-x: auto; background: var(--dom-white); border-left: 3px solid var(--dom-gold); padding: 14px 16px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); margin-top: 14px; }
  .dash-table { width: 100%; font-size: 12px; border-collapse: collapse; }
  .dash-table th { background: var(--dom-gray-light); color: var(--dom-gray-dark); font-size: 10px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.5px; padding: 8px 10px; text-align: left; border-bottom: 2px solid var(--dom-gold); white-space: nowrap; }
  .dash-table td { padding: 7px 10px; border-bottom: 1px solid var(--dom-gray-light); }
  .dash-table td.num { text-align: right; font-variant-numeric: tabular-nums; }
  .dash-table tbody tr:hover { background: var(--dom-gray-light); }
  .dash-note { font-size: 11px; color: var(--dom-gray-mid); font-style: italic; margin-top: 8px; padding: 0 4px; }
  .heatmap-host { overflow-x: auto; max-height: 500px; overflow-y: auto; }
  .heatmap-tbl { width: 100%; border-collapse: collapse; font-size: 11px; }
  .heatmap-tbl th, .heatmap-tbl td { padding: 8px 6px; text-align: center; border: 1px solid var(--dom-gray-light); }
  .heatmap-tbl th { background: var(--dom-gray-light); color: var(--dom-gray-dark); font-weight: 700; text-transform: uppercase; font-size: 10px; letter-spacing: 0.5px; position: sticky; top: 0; }
  .heatmap-tbl td.row-label { background: var(--dom-gray-light); font-weight: 600; text-align: left; padding-left: 10px; white-space: nowrap; }
  .heatmap-tbl td .cell-val { font-weight: 600; color: var(--dom-black); font-size: 13px; }
  .insights { font-size: 13px; line-height: 1.7; color: var(--dom-gray-dark); }
  .insights strong { color: var(--dom-black); }
  .insights .insight { padding: 8px 0; border-bottom: 1px solid var(--dom-gray-light); }
  .insights .insight:last-child { border-bottom: none; }
  @media (max-width: 768px) {
    .hero { padding: 20px; } .hero h1 { font-size: 20px; } .container { padding: 16px; }
    table { font-size: 11px; } thead th, tbody td { padding: 8px 6px; }
    .tabs-inner { padding: 0 16px; } .tab-btn { padding: 12px 16px; font-size: 11px; }
  }
</style>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
</head>
<body>
<!-- Pseudo-senha (não é segurança forte; bloqueia usuários casuais) -->
<div id="pwd-gate" style="position:fixed;inset:0;background:#4D4D4D;z-index:99999;display:flex;align-items:center;justify-content:center;font-family:'Calibri','Arial',sans-serif;">
  <div style="background:#fff;padding:40px 48px;border-top:4px solid #C9A84C;box-shadow:0 4px 24px rgba(0,0,0,0.3);max-width:380px;text-align:center;">
    <div style="font-size:11px;letter-spacing:2px;color:#8C8C8C;text-transform:uppercase;margin-bottom:8px;">DOM Incorporação</div>
    <div style="font-size:18px;font-weight:600;color:#000;margin-bottom:24px;letter-spacing:0.5px;">Inteligência de Mercado</div>
    <input id="pwd-input" type="password" placeholder="Senha de acesso" autofocus
           style="width:100%;padding:12px 14px;border:1px solid #ccc;border-radius:4px;font-size:14px;font-family:inherit;outline:none;margin-bottom:12px;"
           onkeydown="if(event.key==='Enter')checkPwd()" />
    <button onclick="checkPwd()" style="width:100%;padding:12px;background:#C9A84C;border:none;color:#000;font-size:13px;font-weight:600;letter-spacing:1px;text-transform:uppercase;cursor:pointer;border-radius:4px;font-family:inherit;">Entrar</button>
    <div id="pwd-err" style="color:#8C2525;font-size:12px;margin-top:12px;display:none;">Senha incorreta</div>
  </div>
</div>
<script>
const PWD_HASH = "81da67e9aed4bda371aaaeae1b81d014d7e70a355bd917fce8c2e718745e4848"; // senha: DOM2026 (troque o hash com sha256 da nova senha)
async function sha256Hex(s) {
  const buf = new TextEncoder().encode(s);
  const hash = await crypto.subtle.digest("SHA-256", buf);
  return Array.from(new Uint8Array(hash)).map(b => b.toString(16).padStart(2, "0")).join("");
}
async function checkPwd() {
  const inp = document.getElementById("pwd-input").value;
  const h = await sha256Hex(inp);
  if (h === PWD_HASH) {
    sessionStorage.setItem("dom-auth", "1");
    document.getElementById("pwd-gate").style.display = "none";
  } else {
    document.getElementById("pwd-err").style.display = "block";
  }
}
// Pula gate se já autenticou nessa sessão
if (sessionStorage.getItem("dom-auth") === "1") {
  document.getElementById("pwd-gate").style.display = "none";
}
</script>

<header class="hero">
  <div class="hero-inner">
    <div class="hero-brand">
      <img class="hero-logo" src="__LOGO_B64__" alt="DOM Incorporação" />
      <div>
        <h1>INTELIGÊNCIA DE MERCADO</h1>
        <div class="subtitle">Panorama de Lançamentos — Grande São Luís</div>
      </div>
    </div>
    <div class="meta">
      Atualizado em __DATA_UPDATE__<br>
      <span style="color: var(--dom-gold);">__VERSAO__ · Fase 2 — Dashboard Interativo</span>
    </div>
  </div>
</header>
<nav class="tabs">
  <div class="tabs-inner">
    <button class="tab-btn active" data-tab="panorama">Panorama <span class="count" id="tab-count-panorama">—</span></button>
    <button class="tab-btn" data-tab="dashboard">Dashboard</button>
    <button class="tab-btn" data-tab="dados">Dados Completos <span class="count" id="tab-count-dados">—</span></button>
  </div>
</nav>
<main class="container">
  <!-- TAB: PANORAMA -->
  <div class="tab-panel active" id="tab-panorama">
  <section class="kpis" id="kpis">
    <div class="kpi highlight"><div class="label">Empreendimentos</div><div class="value" id="kpi-count">—</div></div>
    <div class="kpi"><div class="label">Incorporadoras</div><div class="value" id="kpi-inc">—</div></div>
    <div class="kpi"><div class="label">Ticket médio</div><div class="value" id="kpi-ticket">—<span class="unit">R$</span></div></div>
    <div class="kpi"><div class="label">R$/m² médio</div><div class="value" id="kpi-rsm2">—<span class="unit">R$</span></div></div>
    <div class="kpi"><div class="label">VGV somado</div><div class="value" id="kpi-vgv">—<span class="unit">R$</span></div></div>
  </section>
  <section class="filters">
    <div class="filter-group"><label>Incorporadora</label><select id="f-inc"><option value="">Todas</option></select></div>
    <div class="filter-group"><label>Bairro</label><select id="f-bairro"><option value="">Todos</option></select></div>
    <div class="filter-group"><label>Tipo</label><select id="f-tipo"><option value="">Todos</option></select></div>
    <div class="filter-group"><label>Segmento</label><select id="f-seg"><option value="">Todos</option></select></div>
    <div class="filter-group"><label>Ano lançamento</label><select id="f-ano"><option value="">Todos</option></select></div>
    <div class="filter-group"><label>Busca</label><input type="text" id="f-search" placeholder="Nome do empreendimento..." /></div>
    <button class="reset-btn" onclick="resetFilters()">Limpar Filtros</button>
    <div class="results-count"><strong id="res-count">0</strong> de <span id="res-total">0</span> empreendimentos</div>
  </section>
  <section class="legend" id="legend"></section>

  <div class="table-intro">
    <strong>📊 Tabela A — Empreendimentos com tabela de vendas</strong> · acesso completo a tickets, áreas e estoque por unidade (<strong id="cnt-a">—</strong>)
  </div>
  <section class="table-wrap">
    <table id="tbl">
      <thead><tr>
        <th data-col="incorporadora">Incorporadora</th>
        <th data-col="empreendimento">Empreendimento</th>
        <th data-col="bairro">Bairro</th>
        <th data-col="tipo">Tipo</th>
        <th data-col="segmento">Segmento</th>
        <th data-col="lancamento_sort">Lançamento</th>
        <th data-col="dorms">Tipologia</th>
        <th data-col="unidades">Total Unid.</th>
        <th data-col="area_med">Área méd (m²)</th>
        <th data-col="ticket_med">Ticket (R$)</th>
        <th data-col="rsm2">R$/m²</th>
        <th data-col="vgv">VGV (R$)</th>
        <th data-col="vendido">% Vendido</th>
      </tr></thead>
      <tbody id="tbody"></tbody>
    </table>
  </section>

  <div class="table-intro" style="background:linear-gradient(90deg,#fff,#fff5e6);border-left-color:var(--dom-gold)">
    <strong>🚀 Tabela B — Breve Lançamentos mapeados</strong> · prática de mercado: tabela de vendas indisponível nessa fase (<strong id="cnt-b">—</strong>)
  </div>
  <section class="table-wrap">
    <table id="tbl-b">
      <thead><tr>
        <th data-col="incorporadora">Incorporadora</th>
        <th data-col="empreendimento">Empreendimento</th>
        <th data-col="bairro">Bairro</th>
        <th data-col="tipo">Tipo</th>
        <th data-col="segmento">Segmento</th>
        <th data-col="lancamento_sort">Lançamento</th>
        <th data-col="dorms">Tipologia</th>
        <th data-col="unidades">Total Unid.</th>
        <th data-col="area_med">Área méd (m²)</th>
        <th data-col="ticket_med">Ticket (R$)</th>
        <th data-col="rsm2">R$/m²</th>
        <th data-col="vgv">VGV (R$)</th>
        <th data-col="vendido">% Vendido</th>
      </tr></thead>
      <tbody id="tbody-b"></tbody>
    </table>
  </section>

  <div class="table-intro incomplete">
    <strong>🔎 Tabela C — Demais empreendimentos sem tabela de vendas</strong> · <strong id="cnt-c">—</strong> empreendimento(s) — <em>priorizar buscar tabela com corretor/site/imprensa</em>
  </div>
  <section class="table-wrap">
    <table id="tbl-c">
      <thead><tr>
        <th data-col="incorporadora">Incorporadora</th>
        <th data-col="empreendimento">Empreendimento</th>
        <th data-col="bairro">Bairro</th>
        <th data-col="tipo">Tipo</th>
        <th data-col="segmento">Segmento</th>
        <th data-col="lancamento_sort">Lançamento</th>
        <th data-col="dorms">Tipologia</th>
        <th data-col="unidades">Total Unid.</th>
        <th data-col="area_med">Área méd (m²)</th>
        <th data-col="ticket_med">Ticket (R$)</th>
        <th data-col="rsm2">R$/m²</th>
        <th data-col="vgv">VGV (R$)</th>
        <th data-col="vendido">% Vendido</th>
      </tr></thead>
      <tbody id="tbody-c"></tbody>
    </table>
  </section>
  </div><!-- /tab-panorama -->

  <!-- TAB: DASHBOARD v8.0.0 — análise por bairro enxuta + bubble oferta granularidade PLANTA (v7.0 PADRAO) -->
  <div class="tab-panel" id="tab-dashboard">

    <section class="dash-filters-bar">
      <div class="filter-group"><label>Período de lançamento</label>
        <select id="dash-f-periodo">
          <option value="12">Último ano</option>
          <option value="24" selected>Últimos 2 anos</option>
          <option value="36">Últimos 3 anos</option>
          <option value="0">Todos</option>
        </select>
      </div>
      <div class="filter-group"><label>Segmento</label><select id="dash-f-seg"></select></div>
      <div class="filter-group"><label>Tipo</label><select id="dash-f-tipo"></select></div>
      <div class="filter-group"><label>Bairro</label><select id="dash-f-bairro"></select></div>
      <div class="filter-group"><label>Incorporadora</label><select id="dash-f-inc"></select></div>
      <button class="reset-btn" onclick="resetDashFilters()">Limpar</button>
      <div class="results-count"><strong id="dash-res-count">0</strong> empreend. analisados</div>
    </section>

    <section class="dash-section">
      <h2 class="dash-section-h">Visão Geral</h2>
      <div class="kpis dash-kpis-grid" id="dash-kpis"></div>
    </section>

    <section class="dash-section">
      <h2 class="dash-section-h">Análise por Bairro</h2>

      <!-- 1. Posicionamento competitivo: bubble ticket × R$/m² (cada bolha = 1 bairro, tamanho = nº empreend.) -->
      <div class="dash-card-mini">
        <div class="chart-wrap" style="height:420px"><canvas id="ch-bairro-bubble"></canvas></div>
        <p class="dash-note">Posicionamento competitivo dos bairros. Eixo X = R$/m² médio · Eixo Y = ticket médio · tamanho da bolha = nº de empreendimentos no bairro. Bairros isolados (poucos players) tendem a indicar nichos premium ou vácuo de oferta.</p>
      </div>

      <!-- 2. Mapa de oferta: bubble área × unidades disponíveis (cor = bairro, 1 bolha por (bairro, tipologia)) -->
      <div class="dash-card-mini" style="margin-top:16px">
        <div class="chart-wrap" style="height:460px"><canvas id="ch-bairro-oferta"></canvas></div>
        <p class="dash-note">Mapa de oferta por bairro × tamanho do apartamento. Cada bolha = um cruzamento (bairro × tipologia). <strong>Eixo X</strong> = área média do apto (m²) · <strong>Eixo Y</strong> = unidades disponíveis · <strong>cor</strong> = bairro · <strong>tamanho</strong> = nº de empreendimentos competindo no cruzamento. Bairro com bolhas concentradas em uma só faixa de área = mono-oferta (oportunidade de diferenciação). Bolha alta + grande = oferta volumosa e pulverizada (mercado disputado). Bolha alta + pequena = oferta volumosa mas concentrada em poucos players.</p>
      </div>

      <!-- 3. Tabela resumo (último elemento) -->
      <div class="dash-table-wrap" style="margin-top:16px">
        <table class="dash-table">
          <thead><tr>
            <th>Bairro</th><th># Emp.</th><th>Total Unid.</th><th>Disponíveis</th>
            <th>% Absorção</th><th>Ticket méd.</th><th>R$/m² méd.</th>
            <th>VGV</th><th>Segmentos</th>
          </tr></thead>
          <tbody id="tbl-bairros-body"></tbody>
        </table>
      </div>
    </section>

    <section class="dash-section">
      <h2 class="dash-section-h">Análise por Tipologia</h2>
      <div class="dash-card-mini"><div class="chart-wrap"><canvas id="ch-tipo-rsm2"></canvas></div></div>
      <div class="dash-table-wrap">
        <table class="dash-table">
          <thead><tr>
            <th>Tipologia</th><th># Empreend.</th><th>Total Unid.</th>
            <th>R$/m² mín</th><th>R$/m² médio</th><th>R$/m² máx</th>
            <th>Área mín</th><th>Área média</th><th>Área máx</th>
            <th>Ticket médio</th>
          </tr></thead>
          <tbody id="tbl-tipologias-body"></tbody>
        </table>
      </div>
      <p class="dash-note">v8.0: dados precisos por tipologia (aba Composição da xlsx). Cobertura atual: empreendimentos com tabela detalhada arquivada. Roadmap (Lote 2/3): processar tabelas restantes para fechar cobertura.</p>
    </section>

    <section class="dash-section">
      <h2 class="dash-section-h">Análise por Incorporadora</h2>
      <div class="dash-section-grid">
        <div class="dash-card-mini"><div class="chart-wrap"><canvas id="ch-inc-vgv"></canvas></div></div>
        <div class="dash-card-mini"><div class="chart-wrap" style="height:340px"><canvas id="ch-inc-scatter"></canvas></div></div>
      </div>
      <div class="dash-table-wrap">
        <table class="dash-table">
          <thead><tr>
            <th>Incorporadora</th><th># Emp.</th><th>VGV</th>
            <th>Ticket méd.</th><th>R$/m² méd.</th><th>% Abs.</th>
            <th>Segmentos</th><th>Bairros</th>
          </tr></thead>
          <tbody id="tbl-incorps-body"></tbody>
        </table>
      </div>
    </section>

    <section class="dash-section">
      <h2 class="dash-section-h">Análise por Segmento</h2>
      <div class="dash-card-mini"><div class="chart-wrap"><canvas id="ch-seg-pie"></canvas></div></div>
      <div class="dash-table-wrap">
        <table class="dash-table">
          <thead><tr>
            <th>Segmento</th><th># Emp.</th><th>Total Unid.</th><th>Disponíveis</th>
            <th>% Abs.</th><th>Ticket méd.</th><th>R$/m² méd.</th><th>Área méd.</th>
            <th>VGV</th><th>Bairros</th>
          </tr></thead>
          <tbody id="tbl-segmentos-body"></tbody>
        </table>
      </div>
    </section>

    <!-- v7.1.0: seção Mapas de Calor consolidada dentro de Análise por Bairro -->

  </div><!-- /tab-dashboard -->


  <!-- TAB: DADOS COMPLETOS -->
  <div class="tab-panel" id="tab-dados">
    <section class="filters" style="margin-top:8px;">
      <div class="filter-group"><label>Busca</label><input type="text" id="fd-search" placeholder="Buscar em qualquer coluna..." style="min-width:280px"/></div>
      <div class="filter-group"><label>Incorporadora</label><select id="fd-inc"><option value="">Todas</option></select></div>
      
      <div class="results-count"><strong id="fd-res-count">0</strong> de <span id="fd-res-total">0</span> linhas</div>
    </section>
    <section class="table-wrap tbl-compact">
      <div class="table-header">Dados Completos · Reflete todas as colunas da Planilha Mestre</div>
      <table id="tbl-full">
        <thead><tr>
          <th data-col="incorporadora">Incorporadora</th>
          <th data-col="empreendimento">Empreendimento</th>
          <th data-col="endereco">Endereço</th>
          <th data-col="bairro">Bairro</th>
          <th data-col="tipo">Tipo</th>
          <th data-col="segmento">Segmento</th>
          <th data-col="unidades">Nº unid.</th>
          <th data-col="lancamento_sort">Lançamento</th>
          <th data-col="entrega">Entrega</th>
          <th data-col="area_min">Área mín</th>
          <th data-col="area_max">Área máx</th>
          <th data-col="area_med">Área méd</th>
          <th data-col="dorms">Tipologia</th>
          <th data-col="ticket_min">Ticket mín</th>
          <th data-col="ticket_max">Ticket máx</th>
          <th data-col="rsm2">R$/m²</th>
          <th data-col="vgv">VGV</th>
          <th data-col="vendido">% Vendido</th>
          <th data-col="orig_precos">Orig. preços</th>
          <th data-col="orig_estoque">Orig. estoque</th>
          <th data-col="orig_lancamento">Orig. lançamento</th>
          <th data-col="link">Fonte</th>
          <th data-col="data_verif">Data verif.</th>
          <th data-col="obs">Observações</th>
        </tr></thead>
        <tbody id="tbody-full"></tbody>
      </table>
    </section>
  </div><!-- /tab-dados -->

  <div class="footer">
    <strong>DOM Incorporação</strong> · Inteligência de Mercado · São Luís/MA<br>
    Fonte: books e tabelas oficiais das incorporadoras · Gerado de __PLANILHA_NAME__
  </div>
</main>
<script>
const ALL_DATA = __DATA_PLACEHOLDER__;
const COMP_DATA = __COMP_PLACEHOLDER__;  // v8.0: aba Composição (por tipologia)
const DATA = ALL_DATA.filter(e => e.is_active);  // Panorama = apenas ativos no ciclo
const INC_COLORS = __INC_COLORS_PLACEHOLDER__;
function getColor(inc) { return INC_COLORS[inc] || "#8C8C8C"; }
function formatBRL(v, compact=false) {
  if (v == null || v === '' || isNaN(v)) return '—';
  v = Number(v);
  if (compact) {
    if (v >= 1e9) return (v/1e9).toFixed(1).replace('.', ',') + ' bi';
    if (v >= 1e6) return (v/1e6).toFixed(1).replace('.', ',') + ' mi';
    if (v >= 1e3) return (v/1e3).toFixed(0) + ' mil';
  }
  return v.toLocaleString('pt-BR', {maximumFractionDigits: 0});
}
function formatArea(v) { if (v == null || isNaN(v)) return '—'; return Number(v).toLocaleString('pt-BR', {maximumFractionDigits: 1}); }
function segClass(seg) {
  if (!seg) return 'seg-other';
  const s = seg.toLowerCase();
  if (s.includes('luxo')) return 'seg-luxo';
  if (s.includes('médio-alto') || s.includes('medio-alto')) return 'seg-medioalto';
  if (s.includes('alto')) return 'seg-alto';
  if (s.includes('médio') || s.includes('medio')) return 'seg-medio';
  if (s.includes('popular')) return 'seg-popular';
  return 'seg-other';
}
function statusClass(st) {
  if (!st) return '';
  const s = st.toLowerCase();
  if (s.includes('pré')) return 'st-pre';
  if (s.includes('lançamento')) return 'st-lancamento';
  if (s.includes('comerc')) return 'st-comerc';
  if (s.includes('últimas')) return 'st-ultimas';
  return '';
}
function origemClass(o) {
  if (!o) return 'origem';
  const s = String(o).toLowerCase();
  if (s === 'memorial' || s === 'interno') return 'origem strong';
  if (s === 't-36' || s === 'book' || s === 'informado') return 'origem medium';
  if (s === 'pendente') return 'origem pending';
  return 'origem weak'; // imprensa, site, estimado-fraco, outros
}
function tipoClass(t) {
  if (!t) return 'tp-other';
  const s = String(t).toLowerCase();
  if (s.includes('horizontal')) return 'tp-horizontal';
  if (s.includes('vertical')) return 'tp-vertical';
  return 'tp-other';
}
function isComplete(e) {
  return e.area_med != null && e.ticket_min != null && e.ticket_max != null && e.rsm2 != null;
}
function populateFilters() {
  const incs = [...new Set(DATA.map(e => e.incorporadora))].sort();
  const bairros = [...new Set(DATA.map(e => e.bairro))].sort();
  const tipos = [...new Set(DATA.map(e => e.tipo).filter(t => t && t !== '—'))].sort();
  const segs = [...new Set(DATA.map(e => e.segmento).filter(s => s && s !== '—'))].sort();
  // Extrai anos do campo lancamento (formato MM/AAAA ou AAAA)
  const anos = [...new Set(DATA.map(e => {
    if (!e.lancamento || e.lancamento === '—') return null;
    const m = String(e.lancamento).match(/(\d{4})/);
    return m ? m[1] : null;
  }).filter(Boolean))].sort().reverse(); // mais recentes primeiro
  const fillSel = (id, arr) => {
    const sel = document.getElementById(id);
    arr.forEach(v => { const o = document.createElement('option'); o.value = v; o.textContent = v; sel.appendChild(o); });
  };
  fillSel('f-inc', incs); fillSel('f-bairro', bairros); fillSel('f-tipo', tipos); fillSel('f-seg', segs); fillSel('f-ano', anos);
}
function buildLegend() {
  const incs = [...new Set(DATA.map(e => e.incorporadora))].sort();
  document.getElementById('legend').innerHTML = incs.map(inc => `
    <div class="legend-item"><span class="legend-dot" style="background:${getColor(inc)}"></span><span>${inc} (${DATA.filter(e=>e.incorporadora===inc).length})</span></div>
  `).join('');
}
let sortCol = 'lancamento_sort';
let sortAsc = false;
function applyFilters() {
  const fi = document.getElementById('f-inc').value;
  const fb = document.getElementById('f-bairro').value;
  const ft = document.getElementById('f-tipo').value;
  const fs = document.getElementById('f-seg').value;
  const fa = document.getElementById('f-ano').value;
  const fq = document.getElementById('f-search').value.toLowerCase();
  const filt = DATA.filter(e => {
    if (fi && e.incorporadora !== fi) return false;
    if (fb && e.bairro !== fb) return false;
    if (ft && e.tipo !== ft) return false;
    if (fs && e.segmento !== fs) return false;
    if (fa) {
      const m = String(e.lancamento || '').match(/(\d{4})/);
      if (!m || m[1] !== fa) return false;
    }
    if (fq && !(e.empreendimento.toLowerCase().includes(fq) || e.incorporadora.toLowerCase().includes(fq))) return false;
    return true;
  });
  renderTable(filt); renderKPIs(filt);
  document.getElementById('res-count').textContent = filt.length;
  document.getElementById('res-total').textContent = DATA.length;
}
function renderKPIs(data) {
  document.getElementById('kpi-count').textContent = data.length;
  document.getElementById('kpi-inc').textContent = new Set(data.map(e => e.incorporadora)).size;
  const tickets = data.flatMap(e => { const arr = []; if (e.ticket_min) arr.push(e.ticket_min); if (e.ticket_max) arr.push(e.ticket_max); return arr; });
  const tkMed = tickets.length ? tickets.reduce((a,b)=>a+b,0)/tickets.length : 0;
  document.getElementById('kpi-ticket').innerHTML = (tkMed ? formatBRL(tkMed, true) : '—') + ' <span class="unit">R$</span>';
  const rsm2 = data.map(e => e.rsm2).filter(v => v);
  const rMed = rsm2.length ? rsm2.reduce((a,b)=>a+b,0)/rsm2.length : 0;
  document.getElementById('kpi-rsm2').innerHTML = (rMed ? formatBRL(rMed) : '—') + ' <span class="unit">R$/m²</span>';
  const vgv = data.map(e => e.vgv).filter(v => v).reduce((a,b)=>a+b, 0);
  document.getElementById('kpi-vgv').innerHTML = (vgv ? formatBRL(vgv, true) : '—') + ' <span class="unit">R$</span>';
}
function tipologiaDetail(obs) {
  // Extrai trecho "Tipologia detalhada: ..." das Observações para tooltip do ℹ
  if (!obs) return null;
  const m = String(obs).match(/Tipologia detalhada: (.+?)\.\s/);
  return m ? m[1] : null;
}

function renderTable(data) {
  const doSort = arr => [...arr].sort((a,b) => {
    let av = a[sortCol] ?? '';
    let bv = b[sortCol] ?? '';
    if (sortCol === 'ticket_med') { av = ((a.ticket_min||0) + (a.ticket_max||0))/2; bv = ((b.ticket_min||0) + (b.ticket_max||0))/2; }
    if (typeof av === 'number' && typeof bv === 'number') return sortAsc ? av-bv : bv-av;
    return sortAsc ? String(av).localeCompare(String(bv)) : String(bv).localeCompare(String(av));
  });

  // v8.1.0: 3 buckets pelo fase_comercial (Tabela A/B/C — mesmas colunas)
  const tabelaA = doSort(data.filter(e => e.fase_comercial === 'lancado_com_tabela'));
  const tabelaB = doSort(data.filter(e => e.fase_comercial === 'breve_lancamento'));
  const tabelaC = doSort(data.filter(e => e.fase_comercial === 'lancado_sem_tabela'));

  document.getElementById('cnt-a').textContent = tabelaA.length;
  document.getElementById('cnt-b').textContent = tabelaB.length;
  document.getElementById('cnt-c').textContent = tabelaC.length;

  // Renderizador único (mesmas colunas em A/B/C — só diferencia tooltip da tag de fase)
  function renderRow(e, fase) {
    const ticket = (e.ticket_min != null && e.ticket_max != null)
      ? `R$ ${formatBRL(e.ticket_min, true)}–${formatBRL(e.ticket_max, true)}`
      : '<span class="dim">—</span>';
    const area = (e.area_med != null) ? formatArea(e.area_med) : '<span class="dim">—</span>';
    const rsm2 = (e.rsm2 != null) ? `R$ ${formatBRL(e.rsm2)}` : '<span class="dim">—</span>';
    const vgv = (e.vgv != null && e.vgv > 0) ? `R$ ${formatBRL(e.vgv, true)}` : '<span class="dim">—</span>';
    const precoTip = (e.orig_precos && e.orig_precos !== '—' && e.orig_precos !== 'N/A')
      ? ` <span class="info-icon" title="Origem dos preços: ${e.orig_precos}">ℹ</span>` : '';
    return `
      <tr onclick="focusEmp('${e.empreendimento.replace(/'/g, "\\'")}')">
        <td><span class="inc-name" style="color:${getColor(e.incorporadora)}">●</span> <span class="inc-name">${e.incorporadora}</span></td>
        <td class="emp-name">${e.empreendimento}</td>
        <td>${e.bairro}${e.endereco && e.endereco !== '—' ? ` <span class="info-icon" title="Endereço: ${e.endereco.replace(/"/g, '&quot;')}">ℹ</span>` : ''}</td>
        <td><span class="chip ${tipoClass(e.tipo)}">${e.tipo}</span></td>
        <td><span class="chip ${segClass(e.segmento)}">${e.segmento}</span></td>
        <td class="${fase==='lancado_com_tabela'?'price':'dim'}" style="font-weight:${fase==='lancado_com_tabela'?'600':'400'}">${e.lancamento}${e.lancamento_origem && e.lancamento_origem !== '—' ? ` <span class="info-icon" title="Origem da data: ${e.lancamento_origem}">ℹ</span>` : ''}</td>
        <td class="dim" style="font-size:11px">${e.dorms || '—'}${(() => { const d = tipologiaDetail(e.obs); return d ? ` <span class="info-icon" title="Tipologia detalhada: ${d.replace(/"/g, '&quot;')}">ℹ</span>` : ''; })()}</td>
        <td class="price">${totalUnidCell(e)}</td>
        <td class="price">${area}</td>
        <td class="price">${ticket}${precoTip}</td>
        <td class="price">${rsm2}${precoTip}</td>
        <td class="price">${vgv}</td>
        <td class="price">${vendidoCell(e)}</td>
      </tr>
    `;
  }

  document.getElementById('tbody').innerHTML   = tabelaA.map(e => renderRow(e, 'lancado_com_tabela')).join('');
  document.getElementById('tbody-b').innerHTML = tabelaB.map(e => renderRow(e, 'breve_lancamento')).join('');
  document.getElementById('tbody-c').innerHTML = tabelaC.map(e => renderRow(e, 'lancado_sem_tabela')).join('');

  document.querySelectorAll('#tbl thead th, #tbl-b thead th, #tbl-c thead th').forEach(th => {
    th.classList.remove('sorted', 'asc');
    if (th.dataset.col === sortCol) { th.classList.add('sorted'); if (sortAsc) th.classList.add('asc'); }
  });
}
// v9.0: helper para célula Total Unidades com tooltip rico de origem
function totalUnidCell(e) {
  const total = e.unidades;
  const orig = e.orig_total || 'N/A';
  if (total == null) {
    return `<span class="dim">—</span> <span class="info-icon" title="Origem: ${orig}. Total não declarado no E_RAW.">ℹ</span>`;
  }
  // Calcular vendidas inferidas se temos % vendido
  let extra = '';
  if (e.vendido != null && total > 0) {
    const vendidas = Math.round(total * e.vendido);
    const dispon = total - vendidas;
    extra = ` Vendidas estimadas: ${vendidas}. Disponíveis estimadas: ${dispon}.`;
  }
  const titleAttr = `Origem: ${orig}.${extra}`.replace(/"/g, '&quot;');
  return `<strong>${total}</strong> <span class="info-icon" title="${titleAttr}">ℹ</span>`;
}

// v9.4: descrição amigável das origens §3.8 (PADRAO v5.3)
function descreveOrigemVendido(o) {
  const map = {
    'calculado_automatico': 'estoque = Σ disponíveis (Composição) / total. Cálculo direto pela fórmula §3.8.',
    'informado_manualmente': 'valor passado manualmente (Rafael/corretor/reunião).',
    'tabela_local_completa_zero': 'origem total=tabela_local_completa AND soma C_RAW=total → 0% vendido (pré-lançamento).',
    'nao_determinavel': 'tabela da incorporadora agrupa unidades por linha (caso Niágara). Fórmula §3.8 não aplicável.',
    'N/A': 'sem dado base (sem total OU sem composição). Vira lista de busca de info.'
  };
  return map[o] || ('origem: ' + o);
}

// v9.4: célula % Vendido — usa origem §3.8 (cálculo automático ou manual)
function vendidoCell(e) {
  const orig = e.orig_pct_vendido || 'N/A';
  if (e.vendido == null || isNaN(e.vendido)) {
    const tip = `Origem: ${orig}. ${descreveOrigemVendido(orig)}`.replace(/"/g, '&quot;');
    return `<span class="dim">—</span> <span class="info-icon" title="${tip}">ℹ</span>`;
  }
  const pct = Math.round(e.vendido * 100);
  const partes = [];
  partes.push(`Origem: ${orig}`);
  partes.push(`Método PADRAO §3.8: ${descreveOrigemVendido(orig)}`);
  if (e.data_verif && e.data_verif !== '—') partes.push(`Última verificação: ${e.data_verif}`);
  const titleAttr = partes.join(' · ').replace(/"/g, '&quot;');
  let style = '';
  if (pct >= 85) style = 'color:#8B6914;font-weight:600';
  else if (pct >= 60) style = 'color:#4D4D4D';
  else style = 'color:#8C8C8C';
  return `<span style="${style}">${pct}%</span> <span class="info-icon" title="${titleAttr}">ℹ</span>`;
}

function focusEmp(name) { /* mapa removido v6.4 — função stub para evitar quebrar onclick antigos */ }
function resetFilters() {
  ['f-inc','f-bairro','f-tipo','f-seg','f-ano','f-search'].forEach(id => document.getElementById(id).value = '');
  applyFilters();
}
document.querySelectorAll('.filters select, .filters input').forEach(el => { el.addEventListener('input', applyFilters); el.addEventListener('change', applyFilters); });
document.querySelectorAll('#tbl thead th, #tbl-b thead th, #tbl-c thead th').forEach(th => {
  th.addEventListener('click', () => {
    const col = th.dataset.col;
    if (sortCol === col) sortAsc = !sortAsc; else { sortCol = col; sortAsc = false; }
    applyFilters();
  });
});
// ─── TAB: DADOS COMPLETOS ─────────────────────────────────────────────────
let sortColFull = 'lancamento_sort';
let sortAscFull = false;

function populateFullFilters() {
  const incs = [...new Set(ALL_DATA.map(e => e.incorporadora))].filter(Boolean).sort();
  const fillSel = (id, arr) => {
    const sel = document.getElementById(id);
    arr.forEach(v => { const o = document.createElement('option'); o.value = v; o.textContent = v; sel.appendChild(o); });
  };
  fillSel('fd-inc', incs);
}

function cell(v, opts={}) {
  if (v == null || v === '' || v === '—') return '<span class="dim">—</span>';
  if (opts.currency) return 'R$ ' + formatBRL(v);
  if (opts.area) return formatArea(v);
  if (opts.pct && typeof v === 'number') return (v*100).toFixed(0) + '%';
  return String(v);
}

function applyFullFilters() {
  const fq = document.getElementById('fd-search').value.toLowerCase();
  const fi = document.getElementById('fd-inc').value;
  const filt = ALL_DATA.filter(e => {
    if (fi && e.incorporadora !== fi) return false;
    if (fq) {
      const blob = [e.incorporadora, e.empreendimento, e.endereco, e.bairro, e.segmento,
                    e.lancamento, e.entrega, e.dorms, e.orig_precos, e.orig_estoque, e.orig_lancamento,
                    e.data_verif, e.obs].filter(Boolean).join(' ').toLowerCase();
      if (!blob.includes(fq)) return false;
    }
    return true;
  });
  renderFullTable(filt);
  document.getElementById('fd-res-count').textContent = filt.length;
  document.getElementById('fd-res-total').textContent = ALL_DATA.length;
}

function renderFullTable(data) {
  const sorted = [...data].sort((a,b) => {
    let av = a[sortColFull] ?? '';
    let bv = b[sortColFull] ?? '';
    if (typeof av === 'number' && typeof bv === 'number') return sortAscFull ? av-bv : bv-av;
    if (av === null || av === undefined) av = '';
    if (bv === null || bv === undefined) bv = '';
    return sortAscFull ? String(av).localeCompare(String(bv)) : String(bv).localeCompare(String(av));
  });
  document.getElementById('tbody-full').innerHTML = sorted.map(e => `
    <tr>
      <td><span class="inc-name" style="color:${getColor(e.incorporadora)}">●</span> <strong>${e.incorporadora || '—'}</strong></td>
      <td class="emp-name">${e.empreendimento || '—'}</td>
      <td class="wrap dim">${e.endereco || '—'}</td>
      <td>${e.bairro || '—'}</td>
      <td>${e.tipo ? `<span class="chip ${tipoClass(e.tipo)}">${e.tipo}</span>` : '—'}</td>
      <td>${e.segmento ? `<span class="chip ${segClass(e.segmento)}">${e.segmento}</span>` : '—'}</td>
      <td class="price">${cell(e.unidades)}</td>
      <td class="price">${e.lancamento || '—'}</td>
      <td>${e.entrega || '—'}</td>
      <td class="price">${cell(e.area_min, {area:true})}</td>
      <td class="price">${cell(e.area_max, {area:true})}</td>
      <td class="price">${cell(e.area_med, {area:true})}</td>
      <td class="dim">${e.dorms || '—'}</td>
      <td class="price">${cell(e.ticket_min, {currency:true})}</td>
      <td class="price">${cell(e.ticket_max, {currency:true})}</td>
      <td class="price">${cell(e.rsm2, {currency:true})}</td>
      <td class="price">${cell(e.vgv, {currency:true})}</td>
      <td class="price">${cell(e.vendido, {pct:true})}</td>
      <td class="dim">${e.orig_precos || '—'}</td>
      <td class="dim">${e.orig_estoque || '—'}</td>
      <td class="dim">${e.orig_lancamento || '—'}</td>
      <td class="dim" style="text-align:center">${e.link ? `<a href="${e.link}" target="_blank" rel="noopener" title="${e.link}" style="text-decoration:none;font-size:14px">🔗</a>` : '—'}</td>
      <td class="dim">${e.data_verif || '—'}</td>
      <td class="wrap dim" style="font-size:10.5px">${e.obs || '—'}</td>
    </tr>
  `).join('');
  document.querySelectorAll('#tbl-full thead th').forEach(th => {
    th.classList.remove('sorted', 'asc');
    if (th.dataset.col === sortColFull) { th.classList.add('sorted'); if (sortAscFull) th.classList.add('asc'); }
  });
}

document.querySelectorAll('#tab-dados .filters select, #tab-dados .filters input').forEach(el => {
  el.addEventListener('input', applyFullFilters); el.addEventListener('change', applyFullFilters);
});
document.querySelectorAll('#tbl-full thead th').forEach(th => {
  th.addEventListener('click', () => {
    const col = th.dataset.col;
    if (sortColFull === col) sortAscFull = !sortAscFull; else { sortColFull = col; sortAscFull = false; }
    applyFullFilters();
  });
});

// ─── Tab switcher ─────────────────────────────────────────────────────────
document.querySelectorAll('.tab-btn').forEach(btn => {
  btn.addEventListener('click', () => {
    document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    document.querySelectorAll('.tab-panel').forEach(p => p.classList.remove('active'));
    document.getElementById('tab-' + btn.dataset.tab).classList.add('active');
    // Ajusta tamanho do mapa quando reabre Panorama
    if (btn.dataset.tab === 'panorama' && map) {
      setTimeout(() => map.invalidateSize(), 100);
    }
  });
});

// Contadores no topo das abas
document.getElementById('tab-count-panorama').textContent = DATA.length;
document.getElementById('tab-count-dados').textContent = ALL_DATA.length;

// Inicializa ambas (mapa removido v6.4)

// ─── DASHBOARD (v6.4) ─────────────────────────────────────────────────────
const SEG_ORDER = ['Popular', 'Médio', 'Médio-alto', 'Alto', 'Luxo'];
const SEG_COLORS = {
  'Popular': '#B8DBC5', 'Médio': '#8C8C8C', 'Médio-alto': '#E8D5A3',
  'Alto': '#C9A84C', 'Luxo': '#000000', '—': '#F2F2F2'
};
const TIPO_ORDER = ['Vertical', 'Horizontal', 'Loteamento'];

// ═══════════════════════════════════════════════════════════════
// DASHBOARD v7.2.0 — análise por bairro enxuta (foco em decisão)
//   · seção tem 3 elementos: bubble posicionamento → bubble oferta → tabela
//   · removidos: timeline semestral, 3 heatmaps (bairro × incorp/tip/seg)
//   · NOVO bubble de oferta: 1 bolha por (bairro × tipologia)
//       X = área média (m²) · Y = unidades disponíveis · cor = bairro · tamanho = nº emp.
//       Responde: onde estão concentradas unidades de cada tamanho? quem é mono-oferta?
//   · tabela movida pro final da seção (resumo, não destaque)
// ═══════════════════════════════════════════════════════════════
let DASH_FILTER = { periodo_meses: 24, seg: '', tipo: '', bairro: '', inc: '' };
const DASH_CHARTS = {};
const TIPO_ORDER_ENUM = ['Studio', '1D', '2D', '3D', '4D', 'Lote'];

function destroyDashChart(id) {
  if (DASH_CHARTS[id]) { DASH_CHARTS[id].destroy(); delete DASH_CHARTS[id]; }
}

function applyDashFilters() {
  let data = ALL_DATA.slice();
  if (DASH_FILTER.periodo_meses > 0) {
    const cutoff = new Date();
    cutoff.setMonth(cutoff.getMonth() - DASH_FILTER.periodo_meses);
    data = data.filter(e => {
      const m = String(e.lancamento || '').match(/(\d{1,2})\/(\d{4})/);
      if (!m) return false;
      const date = new Date(parseInt(m[2]), parseInt(m[1]) - 1, 1);
      return date >= cutoff;
    });
  }
  if (DASH_FILTER.seg) data = data.filter(e => e.segmento === DASH_FILTER.seg);
  if (DASH_FILTER.tipo) data = data.filter(e => e.tipo === DASH_FILTER.tipo);
  if (DASH_FILTER.bairro) data = data.filter(e => e.bairro === DASH_FILTER.bairro);
  if (DASH_FILTER.inc) data = data.filter(e => e.incorporadora === DASH_FILTER.inc);
  return data;
}

function expandTipologia(data) {
  const out = [];
  data.forEach(e => {
    const dorms = (e.dorms || '').split(';').map(s => s.trim()).filter(s => s && s !== '—');
    if (dorms.length === 0) {
      out.push({...e, _tipologia: '—', _is_mono: false});
    } else if (dorms.length === 1) {
      out.push({...e, _tipologia: dorms[0], _is_mono: true});
    } else {
      dorms.forEach(d => out.push({...e, _tipologia: d, _is_mono: false}));
    }
  });
  return out;
}

function avgSimple(items, value) {
  const vals = items.map(value).filter(v => v != null && !isNaN(v));
  return vals.length ? vals.reduce((a,b) => a+b, 0) / vals.length : null;
}

function median(items, value) {
  const vals = items.map(value).filter(v => v != null && !isNaN(v)).sort((a,b) => a-b);
  if (!vals.length) return null;
  const m = Math.floor(vals.length / 2);
  return vals.length % 2 ? vals[m] : (vals[m-1] + vals[m]) / 2;
}

function rangeStats(items, value) {
  const vals = items.map(value).filter(v => v != null && !isNaN(v));
  if (!vals.length) return null;
  return {
    min: Math.min(...vals),
    max: Math.max(...vals),
    avg: vals.reduce((a,b) => a+b, 0) / vals.length,
    n: vals.length
  };
}

function renderDashboard() {
  if (typeof Chart === 'undefined') return;
  Chart.defaults.font.family = "Calibri, Arial, sans-serif";
  Chart.defaults.color = '#4D4D4D';
  const data = applyDashFilters();
  document.getElementById('dash-res-count').textContent = data.length;
  renderDashKPIs(data);
  renderDashBairros(data);
  renderDashTipologias(data);
  renderDashIncorporadoras(data);
  renderDashSegmentos(data);
  renderDashHeatmaps(data);
}

function renderDashKPIs(data) {
  const total = data.length;
  const vgvTotal = data.reduce((a,e) => a + (e.vgv || 0), 0);
  const ticketMed = median(data, e => (e.ticket_min && e.ticket_max) ? (e.ticket_min + e.ticket_max)/2 : null);
  const rsm2Med = median(data, e => e.rsm2);
  const incs = new Set(data.map(e => e.incorporadora));
  const bairros = new Set(data.map(e => e.bairro).filter(b => b && b !== 'São Luís' && b !== 'Não identificado'));
  const absorcaoMed = avgSimple(data.filter(e => e.vendido != null), e => e.vendido);

  const tipoCount = {};
  expandTipologia(data).forEach(e => {
    if (e._tipologia !== '—') tipoCount[e._tipologia] = (tipoCount[e._tipologia] || 0) + 1;
  });
  const topTipo = Object.entries(tipoCount).sort((a,b) => b[1]-a[1])[0];

  const kpis = [
    {label: 'VGV mapeado', value: 'R$ ' + formatBRL(vgvTotal, true), gold: true},
    {label: 'Empreendimentos', value: total, gold: false},
    {label: 'Incorporadoras', value: incs.size, gold: false},
    {label: 'Bairros', value: bairros.size, gold: false},
    {label: 'Ticket mediano', value: ticketMed ? 'R$ ' + formatBRL(ticketMed, true) : '—', gold: false},
    {label: 'R$/m² mediano', value: rsm2Med ? 'R$ ' + formatBRL(rsm2Med) : '—', gold: false},
    {label: '% Absorção médio', value: absorcaoMed != null ? Math.round(absorcaoMed*100) + '%' : '—', gold: false},
    {label: 'Tipologia top', value: topTipo ? topTipo[0] + ' (' + topTipo[1] + ')' : '—', gold: false},
  ];

  document.getElementById('dash-kpis').innerHTML = kpis.map(k =>
    '<div class="kpi ' + (k.gold ? 'kpi-gold' : '') + '"><div class="kpi-label">' + k.label + '</div><div class="kpi-value">' + k.value + '</div></div>'
  ).join('');
}

function renderDashBairros(data) {
  const groups = {};
  data.forEach(e => {
    const b = e.bairro;
    if (!b || b === 'São Luís' || b === 'Não identificado') return;
    if (!groups[b]) groups[b] = [];
    groups[b].push(e);
  });

  const rows = Object.entries(groups).map(([bairro, items]) => {
    const totalUnid = items.reduce((a,e) => a + (e.unidades || 0), 0);
    const dispon = items.reduce((a,e) => {
      if (e.unidades && e.vendido != null) return a + Math.round(e.unidades * (1 - e.vendido));
      return a;
    }, 0);
    const absorcao = avgSimple(items.filter(e => e.vendido != null), e => e.vendido);
    const ticketMed = avgSimple(items.filter(e => e.ticket_min && e.ticket_max), e => (e.ticket_min + e.ticket_max)/2);
    const rsm2Med = avgSimple(items, e => e.rsm2);
    const vgvTotal = items.reduce((a,e) => a + (e.vgv || 0), 0);
    const segs = [...new Set(items.map(e => e.segmento).filter(s => s && s !== '—'))];
    return { bairro, n: items.length, totalUnid, dispon, absorcao, ticketMed, rsm2Med, vgvTotal, segs };
  }).sort((a,b) => b.vgvTotal - a.vgvTotal);

  const tbody = rows.map(r =>
    '<tr><td><strong>' + r.bairro + '</strong></td>' +
    '<td class="num">' + r.n + '</td>' +
    '<td class="num">' + (r.totalUnid || '—') + '</td>' +
    '<td class="num">' + (r.dispon || '—') + '</td>' +
    '<td class="num">' + (r.absorcao != null ? Math.round(r.absorcao*100)+'%' : '—') + '</td>' +
    '<td class="num">' + (r.ticketMed ? 'R$ '+formatBRL(r.ticketMed, true) : '—') + '</td>' +
    '<td class="num">' + (r.rsm2Med ? 'R$ '+formatBRL(r.rsm2Med) : '—') + '</td>' +
    '<td class="num"><strong>R$ ' + formatBRL(r.vgvTotal, true) + '</strong></td>' +
    '<td>' + r.segs.map(s => '<span class="chip ' + segClass(s) + '" style="font-size:10px">' + s + '</span>').join(' ') + '</td>' +
    '</tr>'
  ).join('');
  document.getElementById('tbl-bairros-body').innerHTML = tbody || '<tr><td colspan="9" style="text-align:center;color:var(--dom-gray-mid);padding:20px">Nenhum bairro com filtros aplicados</td></tr>';

  // ─── Bubble: posicionamento ticket × R$/m² (cada bolha = bairro, tamanho = nº empreend.) ───
  destroyDashChart('ch-bairro-bubble');
  const bubblePts = rows.filter(r => r.ticketMed && r.rsm2Med);
  if (bubblePts.length > 0) {
    const maxN = Math.max(...bubblePts.map(r => r.n));
    DASH_CHARTS['ch-bairro-bubble'] = new Chart(document.getElementById('ch-bairro-bubble'), {
      type: 'bubble',
      data: { datasets: bubblePts.map(r => ({
        label: r.bairro,
        data: [{
          x: r.rsm2Med,
          y: r.ticketMed,
          r: 6 + (r.n / maxN) * 22,  // raio 6–28 px conforme nº empreend.
          _bairro: r.bairro,
          _n: r.n,
          _vgv: r.vgvTotal
        }],
        backgroundColor: 'rgba(201,168,76,0.55)',
        borderColor: '#8B6914',
        borderWidth: 1.5
      })) },
      options: { responsive: true, maintainAspectRatio: false,
        plugins: {
          legend: { display: false },
          title: { display: true, text: 'Posicionamento competitivo dos bairros: Ticket × R$/m² (tamanho = nº empreendimentos)' },
          tooltip: { callbacks: { label: ctx => {
            const d = ctx.raw;
            return d._bairro + ' · ' + d._n + ' empreend. · ticket méd. R$ ' + formatBRL(d.y, true) + ' · R$/m² ' + formatBRL(d.x) + ' · VGV R$ ' + formatBRL(d._vgv, true);
          }}},
          datalabels: false
        },
        scales: {
          x: { title: { display: true, text: 'R$/m² médio do bairro' }, ticks: { callback: v => 'R$ ' + (v/1000).toFixed(0) + 'k' } },
          y: { title: { display: true, text: 'Ticket médio do bairro' }, ticks: { callback: v => 'R$ ' + (v/1000).toFixed(0) + 'k' } }
        }
      },
      plugins: [{
        id: 'bubble-labels',
        afterDatasetsDraw(chart) {
          const {ctx} = chart;
          chart.data.datasets.forEach((ds, i) => {
            const meta = chart.getDatasetMeta(i);
            meta.data.forEach((pt, j) => {
              const d = ds.data[j];
              ctx.save();
              ctx.fillStyle = '#1A1A1A';
              ctx.font = '600 10px Calibri, Arial, sans-serif';
              ctx.textAlign = 'center';
              ctx.textBaseline = 'middle';
              ctx.fillText(d._bairro, pt.x, pt.y);
              ctx.restore();
            });
          });
        }
      }]
    });
  }

  // ─── Bubble: oferta por bairro × tamanho do apto (v8.0.0 — granularidade de PLANTA) ───
  // v8.0.0: cada bolha = (bairro × tipologia × planta) ao invés de (bairro × tipologia).
  // Permite enxergar absorção por TICKET REAL (3D 100m² vs 3D 125m² no mesmo bairro).
  destroyDashChart('ch-bairro-oferta');

  // 1. lookup empreendimento → {bairro, total, vendido}
  const empMap = {};
  data.forEach(e => { empMap[e.empreendimento] = e; });

  // 2. agrega por (bairro, tipologia, planta_label, area_round 1dec):
  //    soma Total_planta (oferta), área única (não mais média ponderada), contagem distinta de empreend.
  const ofertaPair = {};  // 'bairro|tip|planta|area' → {bairro, tipologia, planta, area, totalPlanta, empSet}
  COMP_DATA.forEach(c => {
    const e = empMap[c.empreendimento];
    if (!e) return;
    const b = e.bairro;
    if (!b || b === 'São Luís' || b === 'Não identificado') return;
    if (!c.tipologia || c.tipologia === '—') return;
    // v7.0: usa c.area (valor único). Fallback se vier de v6.2 (média min/max).
    const area = c.area != null ? c.area
                : ((c.area_min && c.area_max) ? (c.area_min + c.area_max) / 2
                                              : (c.area_min || c.area_max || null));
    if (!area) return;
    // v7.0: usa Total_planta (granular); fallback unidades (Total tipologia)
    const totalPlanta = c.total_planta != null ? c.total_planta : (c.unidades || 0);
    if (totalPlanta <= 0) return;
    const planta = c.planta || '';
    const areaRound = Math.round(area * 10) / 10;
    const key = b + '|' + c.tipologia + '|' + planta + '|' + areaRound;
    if (!ofertaPair[key]) {
      ofertaPair[key] = { bairro: b, tipologia: c.tipologia, planta: planta, area: areaRound,
                          totalPlanta: 0, empSet: new Set() };
    }
    const slot = ofertaPair[key];
    slot.totalPlanta += totalPlanta;
    slot.empSet.add(c.empreendimento);
  });
  const pairs = Object.values(ofertaPair).map(p => ({
    bairro: p.bairro, tipologia: p.tipologia, planta: p.planta,
    areaMed: p.area,
    unidDisp: p.totalPlanta,  // mantém nome p/ compat com resto do código (representa OFERTA agora, não disp)
    nEmp: p.empSet.size
  })).filter(p => p.unidDisp > 0);

  // 3. Top N bairros por unidades disponíveis (cor distinta) · demais → "Outros"
  const bairroDisp = {};
  pairs.forEach(p => { bairroDisp[p.bairro] = (bairroDisp[p.bairro] || 0) + p.unidDisp; });
  const TOP_BAIRROS = 7;
  const topBairros = Object.entries(bairroDisp).sort((a,b) => b[1]-a[1]).slice(0, TOP_BAIRROS).map(x => x[0]);
  const bairroDisplay = b => topBairros.includes(b) ? b : 'Outros';

  // Paleta DOM (ouro como destaque) — 1 cor por bairro
  const palette = ['#C9A84C', '#1B4584', '#0F7B6C', '#8B2E2E', '#5D6E3C', '#6A4C93', '#B87333', '#8C8C8C'];
  const bairroColors = {};
  topBairros.forEach((b, i) => bairroColors[b] = palette[i % palette.length]);
  bairroColors['Outros'] = '#BFBFBF';

  // 4. Agrupa por bairro_display em datasets distintos (1 dataset = 1 cor na legenda)
  const bairrosOrder = topBairros.slice();
  if (pairs.some(p => bairroDisplay(p.bairro) === 'Outros')) bairrosOrder.push('Outros');

  const datasetsOferta = bairrosOrder.map(bDisp => {
    const ptsB = pairs.filter(p => bairroDisplay(p.bairro) === bDisp);
    const maxNEmp = Math.max(1, ...pairs.map(p => p.nEmp));
    return {
      label: bDisp,
      data: ptsB.map(p => ({
        x: p.areaMed,
        y: p.unidDisp,
        r: 6 + (p.nEmp / maxNEmp) * 18,  // raio 6–24 px conforme nº empreend.
        _bairro: p.bairro,
        _tip: p.tipologia,
        _planta: p.planta,
        _nEmp: p.nEmp
      })),
      backgroundColor: bairroColors[bDisp] + 'AA',
      borderColor: bairroColors[bDisp],
      borderWidth: 1.5
    };
  }).filter(ds => ds.data.length > 0);

  if (datasetsOferta.length > 0) {
    DASH_CHARTS['ch-bairro-oferta'] = new Chart(document.getElementById('ch-bairro-oferta'), {
      type: 'bubble',
      data: { datasets: datasetsOferta },
      options: { responsive: true, maintainAspectRatio: false,
        plugins: {
          legend: { position: 'right', labels: { boxWidth: 12, font: {size: 11} } },
          title: { display: true, text: 'Mapa de oferta por planta (v7.0): tamanho do apto × oferta total (cor = bairro · tamanho = nº empreend.)' },
          tooltip: { callbacks: { label: ctx => {
            const d = ctx.raw;
            const lbl = d._planta ? (' [' + d._planta + ']') : '';
            return d._bairro + ' · ' + d._tip + lbl + ' · ' + d.x.toFixed(1) + ' m² · ' + d.y + ' unid. (oferta total) · ' + d._nEmp + ' empreend.';
          }}}
        },
        scales: {
          x: { title: { display: true, text: 'Área média do apto (m²)' }, ticks: { callback: v => v + ' m²' } },
          y: { title: { display: true, text: 'Oferta total (Total planta) no cruzamento bairro × tipologia × planta' }, beginAtZero: true, ticks: { precision: 0 } }
        }
      },
      plugins: [{
        id: 'oferta-tip-labels',
        afterDatasetsDraw(chart) {
          const {ctx} = chart;
          chart.data.datasets.forEach((ds, i) => {
            const meta = chart.getDatasetMeta(i);
            meta.data.forEach((pt, j) => {
              const d = ds.data[j];
              if (d.r < 12) return;  // não desenha label em bolhas pequenas (poluição)
              ctx.save();
              ctx.fillStyle = '#000';
              ctx.font = '600 9px Calibri, Arial, sans-serif';
              ctx.textAlign = 'center';
              ctx.textBaseline = 'middle';
              ctx.fillText(d._tip, pt.x, pt.y);
              ctx.restore();
            });
          });
        }
      }]
    });
  } else {
    // Sem dados — limpar canvas
    const c = document.getElementById('ch-bairro-oferta');
    if (c) c.getContext('2d').clearRect(0,0,c.width,c.height);
  }
}

function renderDashTipologias(data) {
  // v8.0: dados precisos da aba Composição
  const empAtivos = new Set(data.map(e => e.empreendimento));
  const compFiltrada = COMP_DATA.filter(c => empAtivos.has(c.empreendimento));

  const groups = {};
  TIPO_ORDER_ENUM.forEach(t => groups[t] = []);
  compFiltrada.forEach(c => {
    if (groups[c.tipologia] !== undefined) groups[c.tipologia].push(c);
  });

  const rows = TIPO_ORDER_ENUM.filter(t => groups[t].length > 0).map(t => {
    const items = groups[t];
    const totalUnid = items.reduce((a,c) => a + (c.unidades || 0), 0);
    const nEmpreend = new Set(items.map(c => c.empreendimento)).size;
    const rsm2_vals = items.map(c => c.rsm2).filter(v => v != null);
    const rsm2 = rsm2_vals.length ? { min: Math.min(...rsm2_vals), max: Math.max(...rsm2_vals), avg: rsm2_vals.reduce((a,b)=>a+b,0)/rsm2_vals.length } : null;
    const area_mins = items.map(c => c.area_min).filter(v => v != null);
    const area_maxs = items.map(c => c.area_max).filter(v => v != null);
    const area = {
      min: area_mins.length ? Math.min(...area_mins) : null,
      max: area_maxs.length ? Math.max(...area_maxs) : null,
      avg: avgSimple(items, c => c.area_min && c.area_max ? (c.area_min + c.area_max)/2 : null),
    };
    const ticket_mins = items.map(c => c.ticket_min).filter(v => v != null);
    const ticket_maxs = items.map(c => c.ticket_max).filter(v => v != null);
    const ticket = {
      min: ticket_mins.length ? Math.min(...ticket_mins) : null,
      max: ticket_maxs.length ? Math.max(...ticket_maxs) : null,
      avg: avgSimple(items, c => c.ticket_min && c.ticket_max ? (c.ticket_min + c.ticket_max)/2 : null),
    };
    return { t, nEmpreend, totalUnid, rsm2, area, ticket };
  });

  const tbody = rows.map(r =>
    '<tr><td><strong>' + r.t + '</strong></td>' +
    '<td class="num">' + r.nEmpreend + '</td>' +
    '<td class="num"><strong>' + r.totalUnid + '</strong></td>' +
    '<td class="num">' + (r.rsm2 ? 'R$ '+formatBRL(r.rsm2.min) : '—') + '</td>' +
    '<td class="num"><strong>' + (r.rsm2 ? 'R$ '+formatBRL(r.rsm2.avg) : '—') + '</strong></td>' +
    '<td class="num">' + (r.rsm2 ? 'R$ '+formatBRL(r.rsm2.max) : '—') + '</td>' +
    '<td class="num">' + (r.area.min != null ? r.area.min.toFixed(0)+'m²' : '—') + '</td>' +
    '<td class="num"><strong>' + (r.area.avg ? r.area.avg.toFixed(0)+'m²' : '—') + '</strong></td>' +
    '<td class="num">' + (r.area.max != null ? r.area.max.toFixed(0)+'m²' : '—') + '</td>' +
    '<td class="num">' + (r.ticket.avg ? 'R$ '+formatBRL(r.ticket.avg, true) : '—') + '</td>' +
    '</tr>'
  ).join('');
  document.getElementById('tbl-tipologias-body').innerHTML = tbody || '<tr><td colspan="10" style="text-align:center;color:var(--dom-gray-mid);padding:20px">Sem dados de composição. Empreendimentos sem tabela detalhada extraível ainda — Lote 2/3 do roadmap.</td></tr>';

  destroyDashChart('ch-tipo-rsm2');
  if (rows.length > 0) {
    DASH_CHARTS['ch-tipo-rsm2'] = new Chart(document.getElementById('ch-tipo-rsm2'), {
      type: 'bar',
      data: {
        labels: rows.map(r => r.t),
        datasets: [
          { label: 'Mín', data: rows.map(r => r.rsm2 ? r.rsm2.min : 0), backgroundColor: '#E8D5A3' },
          { label: 'Médio', data: rows.map(r => r.rsm2 ? r.rsm2.avg : 0), backgroundColor: '#C9A84C' },
          { label: 'Máx', data: rows.map(r => r.rsm2 ? r.rsm2.max : 0), backgroundColor: '#8B6914' },
        ]
      },
      options: { responsive: true, maintainAspectRatio: false,
        plugins: { title: { display: true, text: 'R$/m² por tipologia (mín / médio / máx)' } },
        scales: { y: { ticks: { callback: v => 'R$ '+(v/1000).toFixed(0)+'k' } } } }
    });
  }
}

function renderDashIncorporadoras(data) {
  const groups = {};
  data.forEach(e => {
    if (!groups[e.incorporadora]) groups[e.incorporadora] = [];
    groups[e.incorporadora].push(e);
  });

  const rows = Object.entries(groups).map(([inc, items]) => {
    const vgvTotal = items.reduce((a,e) => a + (e.vgv || 0), 0);
    const bairros = [...new Set(items.map(e => e.bairro).filter(b => b && b !== 'São Luís'))];
    const segs = [...new Set(items.map(e => e.segmento).filter(s => s && s !== '—'))];
    const rsm2Med = avgSimple(items, e => e.rsm2);
    const ticketMed = avgSimple(items.filter(e => e.ticket_min && e.ticket_max), e => (e.ticket_min + e.ticket_max)/2);
    const absorcao = avgSimple(items.filter(e => e.vendido != null), e => e.vendido);
    return { inc, n: items.length, vgvTotal, bairros, segs, rsm2Med, ticketMed, absorcao };
  }).sort((a,b) => b.vgvTotal - a.vgvTotal);

  const tbody = rows.map(r =>
    '<tr><td><span class="inc-name" style="color:' + getColor(r.inc) + '">●</span> <strong>' + r.inc + '</strong></td>' +
    '<td class="num">' + r.n + '</td>' +
    '<td class="num"><strong>R$ ' + formatBRL(r.vgvTotal, true) + '</strong></td>' +
    '<td class="num">' + (r.ticketMed ? 'R$ '+formatBRL(r.ticketMed, true) : '—') + '</td>' +
    '<td class="num">' + (r.rsm2Med ? 'R$ '+formatBRL(r.rsm2Med) : '—') + '</td>' +
    '<td class="num">' + (r.absorcao != null ? Math.round(r.absorcao*100)+'%' : '—') + '</td>' +
    '<td>' + r.segs.map(s => '<span class="chip ' + segClass(s) + '" style="font-size:10px">' + s + '</span>').join(' ') + '</td>' +
    '<td>' + r.bairros.slice(0,3).join(', ') + (r.bairros.length > 3 ? '...' : '') + '</td>' +
    '</tr>'
  ).join('');
  document.getElementById('tbl-incorps-body').innerHTML = tbody || '<tr><td colspan="8" style="text-align:center;color:var(--dom-gray-mid);padding:20px">Sem dados</td></tr>';

  destroyDashChart('ch-inc-vgv');
  const incsComVgv = rows.filter(r => r.vgvTotal > 0);
  if (incsComVgv.length > 0) {
    DASH_CHARTS['ch-inc-vgv'] = new Chart(document.getElementById('ch-inc-vgv'), {
      type: 'bar',
      data: { labels: incsComVgv.map(r => r.inc),
        datasets: [{ label: 'VGV (R$ M)', data: incsComVgv.map(r => r.vgvTotal / 1e6),
          backgroundColor: incsComVgv.map(r => getColor(r.inc)) }] },
      options: { indexAxis: 'y', responsive: true, maintainAspectRatio: false,
        plugins: { legend: { display: false }, title: { display: true, text: 'VGV total por incorporadora (R$ Milhões)' } },
        scales: { x: { ticks: { callback: v => 'R$ ' + v + 'M' } } } }
    });
  }

  const points = data.filter(e => e.rsm2 && e.ticket_min && e.ticket_max).map(e => ({
    x: e.rsm2, y: (e.ticket_min + e.ticket_max) / 2,
    r: e.unidades ? Math.max(5, Math.min(20, Math.sqrt(e.unidades / 2))) : 8,
    label: e.empreendimento + ' (' + e.incorporadora + ')', inc: e.incorporadora
  }));
  destroyDashChart('ch-inc-scatter');
  if (points.length > 0) {
    const incsList = [...new Set(points.map(p => p.inc))];
    DASH_CHARTS['ch-inc-scatter'] = new Chart(document.getElementById('ch-inc-scatter'), {
      type: 'bubble',
      data: { datasets: incsList.map(inc => ({
        label: inc, data: points.filter(p => p.inc === inc),
        backgroundColor: getColor(inc) + 'CC', borderColor: '#000', borderWidth: 1 })) },
      options: { responsive: true, maintainAspectRatio: false,
        plugins: { legend: { position: 'right', labels: { boxWidth: 12, font: {size: 10} } },
          title: { display: true, text: 'Posicionamento: Ticket × R$/m² (cada bolha = 1 empreend.)' },
          tooltip: { callbacks: { label: ctx => {
            const d = ctx.raw;
            return d.label + ': R$/m² ' + formatBRL(d.x) + ' · Ticket ' + formatBRL(d.y, true);
          }}} },
        scales: {
          x: { title: { display: true, text: 'R$/m²' }, ticks: { callback: v => 'R$ ' + (v/1000).toFixed(0) + 'k' } },
          y: { title: { display: true, text: 'Ticket' }, ticks: { callback: v => 'R$ ' + (v/1000).toFixed(0) + 'k' } } } }
    });
  }
}

function renderDashSegmentos(data) {
  const groups = {};
  SEG_ORDER.forEach(s => groups[s] = []);
  data.forEach(e => { if (e.segmento && groups[e.segmento]) groups[e.segmento].push(e); });

  const rows = SEG_ORDER.filter(s => groups[s].length > 0).map(s => {
    const items = groups[s];
    const totalUnid = items.reduce((a,e) => a + (e.unidades || 0), 0);
    const dispon = items.reduce((a,e) => {
      if (e.unidades && e.vendido != null) return a + Math.round(e.unidades * (1 - e.vendido));
      return a;
    }, 0);
    const absorcao = avgSimple(items.filter(e => e.vendido != null), e => e.vendido);
    const ticket = rangeStats(items.filter(e => e.ticket_min && e.ticket_max), e => (e.ticket_min + e.ticket_max)/2);
    const rsm2 = rangeStats(items, e => e.rsm2);
    const area = rangeStats(items, e => e.area_med);
    const vgvTotal = items.reduce((a,e) => a + (e.vgv || 0), 0);
    const bairros = [...new Set(items.map(e => e.bairro).filter(b => b && b !== 'São Luís'))];
    return { s, n: items.length, totalUnid, dispon, absorcao, ticket, rsm2, area, vgvTotal, bairros };
  });

  const tbody = rows.map(r =>
    '<tr><td><span class="chip ' + segClass(r.s) + '">' + r.s + '</span></td>' +
    '<td class="num">' + r.n + '</td>' +
    '<td class="num">' + (r.totalUnid || '—') + '</td>' +
    '<td class="num">' + (r.dispon || '—') + '</td>' +
    '<td class="num">' + (r.absorcao != null ? Math.round(r.absorcao*100)+'%' : '—') + '</td>' +
    '<td class="num">' + (r.ticket ? 'R$ '+formatBRL(r.ticket.avg, true) : '—') + '</td>' +
    '<td class="num">' + (r.rsm2 ? 'R$ '+formatBRL(r.rsm2.avg) : '—') + '</td>' +
    '<td class="num">' + (r.area ? r.area.avg.toFixed(0)+'m²' : '—') + '</td>' +
    '<td class="num"><strong>R$ ' + formatBRL(r.vgvTotal, true) + '</strong></td>' +
    '<td>' + r.bairros.slice(0,3).join(', ') + (r.bairros.length > 3 ? '...' : '') + '</td>' +
    '</tr>'
  ).join('');
  document.getElementById('tbl-segmentos-body').innerHTML = tbody || '<tr><td colspan="10" style="text-align:center;color:var(--dom-gray-mid);padding:20px">Sem dados</td></tr>';

  destroyDashChart('ch-seg-pie');
  const segsAtivos = rows.filter(r => r.vgvTotal > 0);
  if (segsAtivos.length > 0) {
    DASH_CHARTS['ch-seg-pie'] = new Chart(document.getElementById('ch-seg-pie'), {
      type: 'doughnut',
      data: { labels: segsAtivos.map(r => r.s),
        datasets: [{ data: segsAtivos.map(r => r.vgvTotal),
          backgroundColor: segsAtivos.map(r => SEG_COLORS[r.s]),
          borderColor: '#fff', borderWidth: 2 }] },
      options: { responsive: true, maintainAspectRatio: false,
        plugins: { legend: { position: 'right', labels: { boxWidth: 14, font: { size: 11 } } },
          title: { display: true, text: 'VGV por segmento' },
          tooltip: { callbacks: { label: ctx => ctx.label + ': R$ ' + formatBRL(ctx.raw, true) } } } }
    });
  }
}

function renderDashHeatmaps(data) {
  // v7.2.0: heatmap-bt e heatmap-bs foram removidos do DOM.
  // Função preservada como stub safe (early-return) caso o painel volte a precisar
  // de heatmaps no futuro — basta restaurar os <div id="heatmap-bt"> no HTML.
  if (!document.getElementById('heatmap-bt') && !document.getElementById('heatmap-bs')) return;
}

function setupDashFilters() {
  ['dash-f-periodo', 'dash-f-seg', 'dash-f-tipo', 'dash-f-bairro', 'dash-f-inc'].forEach(id => {
    const el = document.getElementById(id);
    if (el) el.addEventListener('change', () => {
      DASH_FILTER.periodo_meses = parseInt(document.getElementById('dash-f-periodo').value);
      DASH_FILTER.seg = document.getElementById('dash-f-seg').value;
      DASH_FILTER.tipo = document.getElementById('dash-f-tipo').value;
      DASH_FILTER.bairro = document.getElementById('dash-f-bairro').value;
      DASH_FILTER.inc = document.getElementById('dash-f-inc').value;
      renderDashboard();
    });
  });
}

function resetDashFilters() {
  DASH_FILTER = { periodo_meses: 24, seg: '', tipo: '', bairro: '', inc: '' };
  document.getElementById('dash-f-periodo').value = '24';
  document.getElementById('dash-f-seg').value = '';
  document.getElementById('dash-f-tipo').value = '';
  document.getElementById('dash-f-bairro').value = '';
  document.getElementById('dash-f-inc').value = '';
  renderDashboard();
}

function populateDashFilters() {
  const segs = SEG_ORDER.filter(s => ALL_DATA.some(e => e.segmento === s));
  const tipos = [...new Set(ALL_DATA.map(e => e.tipo).filter(t => t))].sort();
  const bairrs = [...new Set(ALL_DATA.map(e => e.bairro).filter(b => b))].sort();
  const incs = [...new Set(ALL_DATA.map(e => e.incorporadora))].sort();

  document.getElementById('dash-f-seg').innerHTML = '<option value="">Todos</option>' + segs.map(s => '<option>' + s + '</option>').join('');
  document.getElementById('dash-f-tipo').innerHTML = '<option value="">Todos</option>' + tipos.map(t => '<option>' + t + '</option>').join('');
  document.getElementById('dash-f-bairro').innerHTML = '<option value="">Todos</option>' + bairrs.map(b => '<option>' + b + '</option>').join('');
  document.getElementById('dash-f-inc').innerHTML = '<option value="">Todas</option>' + incs.map(i => '<option>' + i + '</option>').join('');
}

// v7.0 fix: inicializar aba Panorama (sem isso, KPIs e tabelas A/B ficam vazios)
populateFilters(); buildLegend(); applyFilters();
// v7.0: inicializar aba Dashboard
populateDashFilters();
setupDashFilters();
renderDashboard();
// Aba Dados Completos
populateFullFilters(); applyFullFilters();
</script>
</body>
</html>"""


# ─── Pipeline ────────────────────────────────────────────────────────────────
def build(include_all: bool = False) -> None:
    from datetime import datetime
    print("─" * 70)
    print("DOM Incorporação · Inteligência de Mercado · build_panorama")
    print("─" * 70)

    planilha = find_latest_planilha()
    print(f"📊 Planilha: {planilha.name}")

    rows = read_planilha(planilha)
    print(f"📋 Linhas lidas: {len(rows)}")

    composicao = read_composicao(planilha)
    if composicao:
        # v7.0: somar Total_planta (granular) ao invés de unidades (Total tipologia, repetido por planta)
        total_unid_comp = sum((c.get("total_planta") or c.get("unidades") or 0) for c in composicao)
        n_emp_comp = len(set((c.get("incorporadora"), c.get("empreendimento")) for c in composicao))
        print(f"📊 Composição: {len(composicao)} linhas / {total_unid_comp} unidades / {n_emp_comp} empreend.")
    else:
        print("📊 Composição: aba ausente (planilha pré-v8.0)")

    enriched = enrich(rows, include_all=include_all)
    active = [e for e in enriched if e.get("is_active")]
    print(f"📐 Total enriquecido: {len(enriched)} (aba Dados Completos)")
    print(f"🎯 Ativos no ciclo: {len(active)} (aba Panorama)")

    # Info por incorporadora (só ativos)
    from collections import Counter
    incs = Counter(e["incorporadora"] for e in active)
    print("   Por incorporadora (panorama):")
    for inc, n in incs.most_common():
        print(f"      {inc}: {n}")

    # Gera HTML
    data_json = json.dumps(enriched, ensure_ascii=False, separators=(",", ":"), default=str)
    comp_json = json.dumps(composicao, ensure_ascii=False, separators=(",", ":"), default=str)
    inc_colors_json = json.dumps(INC_COLORS, ensure_ascii=False)
    version_match = re.search(r"v([\d.]+)", planilha.stem)
    version = "v" + version_match.group(1) if version_match else "v?"
    hoje = datetime.now().strftime("%d/%m/%Y")
    ciclo = "Todos os empreendimentos" if include_all else "2025/2026"

    # Carrega logo DOM como data URI para embutir no HTML
    logo_path = SCRIPT_DIR / "dom_logo.png"
    logo_b64 = ""
    if logo_path.exists():
        import base64
        logo_b64 = "data:image/png;base64," + base64.b64encode(logo_path.read_bytes()).decode('ascii')

    html = (HTML_TEMPLATE
            .replace("__DATA_PLACEHOLDER__", data_json)
            .replace("__COMP_PLACEHOLDER__", comp_json)
            .replace("__INC_COLORS_PLACEHOLDER__", inc_colors_json)
            .replace("__VERSAO__", version)
            .replace("__DATA_UPDATE__", hoje)
            .replace("__CICLO__", ciclo)
            .replace("__PLANILHA_NAME__", planilha.name)
            .replace("__LOGO_B64__", logo_b64))

    OUTPUT_HTML.write_text(html, encoding="utf-8")
    print(f"\n✅ HTML gerado: {OUTPUT_HTML.name}")
    print(f"   Tamanho: {len(html):,} caracteres")
    print(f"   Caminho: {OUTPUT_HTML}")
    print("─" * 70)


def main():
    parser = argparse.ArgumentParser(description="Gera o Panorama de Lançamentos HTML")
    parser.add_argument("--all", action="store_true", help="Inclui todos os empreendimentos (sem filtro de ciclo)")
    args = parser.parse_args()
    build(include_all=args.all)


if __name__ == "__main__":
    main()
